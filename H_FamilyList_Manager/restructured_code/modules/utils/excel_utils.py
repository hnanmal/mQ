# modules/utils/excel_utils.py

from openpyxl import load_workbook
import tkinter as tk

def display_excel_data_openpyxl(app):
    # Load the Excel file using openpyxl
    file_path = "WorkMaster_DB.xlsx"  # Assuming the file is in the project directory
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Set up the Treeview columns
    columns = [cell.value for cell in sheet[1]]  # Assuming the first row is the header
    app.excel_treeview["columns"] = columns

    # Columns to be hidden (2, 3, 5, 7, 9, 11, 13)
    hidden_columns = [1, 2, 4, 6, 8, 10, 12]  # Zero-based index

    for idx, col in enumerate(columns):
        if idx in hidden_columns:
            app.excel_treeview.column(col, width=0, stretch=tk.NO)
        else:
            app.excel_treeview.heading(col, text=col)
            app.excel_treeview.column(
                col, anchor="w", width=150
            )  # Set a default width for visible columns

    # Insert rows into the Treeview
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip rows 2 through 5
        if 2 <= i <= 6:
            continue
        app.excel_treeview.insert("", "end", values=row)
