# Load the Python Standard and DesignScript Libraries
import os
python_rootpath = os.getenv('LOCALAPPDATA')+'\Programs\Python'
file_list = os.listdir(python_rootpath)
python_version = file_list[-1]
import sys
sys.path.append(os.getenv('LOCALAPPDATA').replace('\\','\\\\') + f'\Programs\Python\{python_version}\Lib\site-packages')
import clr


clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *

import random

from collections import Iterable
import datetime
import openpyxl

from itertools import chain
from functools import reduce

curry = lambda f: lambda a,*args: f(a, *args) if (len(args)) else lambda *args: f(a, *args)

add = curry(lambda a,b: a + b)

filter = curry(filter)
map = curry(map)

def _take(length, iter):
    res = []
    for a in iter:
        res.append(a)
        if len(res) == length:
            return res
take = curry(_take)
reduce = curry(reduce)
go = lambda *args: reduce(lambda a,f: f(a), args)

# The inputs to this node will be stored as a list in the IN variables.
dataEnteringNode = IN
inputFilePath = IN[0]
inputSheetName = IN[1]

def importExcel(path, SheetNameIn):
    wb_obj = openpyxl.load_workbook(str(path), data_only=True)
    sheet_obj = wb_obj.get_sheet_by_name(str(SheetNameIn))

    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    listOut = []

    for i in range(1, m_row +1):
        listTemp = []
        for j in range(1, m_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            listTemp.append(cell_obj.value)
            if len(listTemp) == m_col:
                listOut.append(listTemp)
    return listOut

def checkAllElement(list, target):
    tmp=[]
    for i in list:
        if i == target:
            tmp.append(True)
        else:
            tmp.append(False)
    return all(tmp)

# Place your code below this line
rm_Null = lambda list: go(list,
    filter(lambda a: not checkAllElement(a, None)),
    )
    
s1 = importExcel(inputFilePath, inputSheetName)

s2 = rm_Null(s1)  ## null만 있는 Excel Row 제거

result_tr = list(zip(*s2))

s3 = rm_Null(result_tr) ## null만 있는 Excel Column 제거

result = list(zip(*s3))

# Assign your output to the OUT variable.
OUT = result
