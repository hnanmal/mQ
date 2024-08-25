from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk

class SearchManager:
    def __init__(self, app, file_path="WorkMaster_DB.xlsx"):
        self.app = app
        self.file_path = file_path

    def setup_search(self, parent_frame):
        """Sets up the search box and button on the given parent frame."""
        search_label = ttk.Label(parent_frame, text="Search:")
        search_label.pack(side="left", padx=(0, 5))

        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(parent_frame, textvariable=self.search_var)
        search_entry.pack(side="left", fill="x", expand=True)

        search_button = ttk.Button(
            parent_frame,
            text="Search",
            command=self.execute_search
        )
        search_button.pack(side="left", padx=(5, 0))

        # Bind the Enter key to trigger the search when pressed in the search box
        search_entry.bind("<Return>", lambda event: self.execute_search())

    def execute_search(self):
        """Executes the search and updates the Treeview with filtered data."""
        search_keyword = self.search_var.get()
        self.filter_excel_data(search_keyword)

    def load_excel_data(self):
        """Loads data from the Excel file."""
        workbook = load_workbook(filename=self.file_path)
        sheet = workbook.active
        return sheet

    def clear_treeview(self):
        """Clears all items in the Treeview."""
        for item in self.app.excel_treeview.get_children():
            self.app.excel_treeview.delete(item)

    def insert_filtered_data(self, sheet, search_keyword):
        """Filters and inserts data into the Treeview based on the search keyword."""
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if 2 <= i <= 6:  # Skip rows 2 through 6
                continue
            if any(search_keyword.lower() in str(cell).lower() for cell in row):
                self.app.excel_treeview.insert("", "end", values=row)

    def filter_excel_data(self, search_keyword):
        """Filters the Excel data and updates the Treeview."""
        sheet = self.load_excel_data()
        self.clear_treeview()
        self.insert_filtered_data(sheet, search_keyword)
