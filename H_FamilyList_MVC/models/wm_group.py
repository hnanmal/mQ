# models/wm_group.py

import json
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook


def load_wm_group_match_data(app):
    try:
        with open("wm_group_match.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            app.wm_group_match_data = data.get("matches", {})
            # Update lock status based on the JSON data
            app.lock_status = data.get("lock_status", {})
            # Apply the lock status to the treeview items
            for item in app.wm_group_treeview.get_children():
                item_name = app.wm_group_treeview.item(item, "values")[0]
                is_locked = app.lock_status.get(item_name, False)
                if is_locked:
                    app.wm_group_treeview.item(item, tags=("locked",))
                else:
                    app.wm_group_treeview.item(item, tags=("unlocked",))
            return data
    except FileNotFoundError:
        app.wm_group_match_data = {}
        app.lock_status = {}


def save_lock_status_to_json(app):
    try:
        with open("wm_group_match.json", "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        data = {}

    if "lock_status" not in data:
        data["lock_status"] = {}

    # Update the lock status for the current item
    data["lock_status"] = app.lock_status

    with open("wm_group_match.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)


def save_current_matching_to_json(app, item_name):

    # Load the existing JSON data
    try:
        with open("wm_group_match.json", "r", encoding="utf-8") as file:
            wm_group_match_data = json.load(file)
    except FileNotFoundError:
        wm_group_match_data = {}

    """Save the current matching data to JSON for the specified item."""
    current_matches = list(
        app.drop_area.get(0, tk.END)
    )  # Get all items in the drop area

    # Update the JSON data with the new match
    # wm_group_match_data[item_name] = current_matches
    wm_group_match_data[item_name] = list(current_matches)

    # Save the updated data back to the file
    with open("wm_group_match.json", "w", encoding="utf-8") as file:
        json.dump(wm_group_match_data, file, indent=4, ensure_ascii=False)


def filter_excel_data(app, search_keyword):
    # Get all rows in the Treeview
    all_items = app.excel_treeview.get_children()

    # Clear the Treeview
    for item in all_items:
        app.excel_treeview.delete(item)

    # Reinsert only the rows that match the search keyword
    file_path = "WorkMaster_DB.xlsx"  # Assuming the file is in the project directory
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if 2 <= i <= 6:  # Skip rows 2 through 5
            continue
        if any(search_keyword.lower() in str(cell).lower() for cell in row):
            app.excel_treeview.insert("", "end", values=row)


def save_configuration(app):
    tree_data = []
    for item in app.tree.get_children():
        tree_data.append(app.config_manager.get_item_data(app, item))
    # Save to JSON file
    file_path = filedialog.asksaveasfilename(
        defaultextension=".json", filetypes=[("JSON files", "*.json")]
    )
    if file_path:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(tree_data, f, indent=4, ensure_ascii=False)


# def load_configuration(app, file_path=None):
#     if not file_path:
#         file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
#     if file_path:
#         with open(file_path, "r", encoding="utf-8") as f:
#             tree_data = json.load(f)
#         # Clear existing tree
#         for item in app.tree.get_children():
#             app.tree.delete(item)
#         # Load new tree
#         for item_data in tree_data:
#             app.config_manager.insert_item_data(app, "", item_data)

#     # After populating the treeview, collect level 6 items
#     level_6_items = app.treeview_operations.collect_level_6_items()

#     # Update the Listbox in the "WM 그룹별 매칭" tab
#     app.update_wm_group_matching_treeview(level_6_items)


def load_configuration(app, file_path=None):
    if not file_path:
        file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if file_path:
        with open(file_path, "r", encoding="utf-8") as f:
            tree_data = json.load(f)
        # Clear existing tree
        for item in app.tree.get_children():
            app.tree.delete(item)
        # Load new tree
        for item_data in tree_data:
            app.config_manager.insert_item_data(app, "", item_data)

    # After populating the treeview, collect level 6 items
    level_6_items = app.treeview_operations.collect_level_6_items()

    # Update the Listbox in the "WM 그룹별 매칭" tab
    app.update_wm_group_matching_treeview(level_6_items)

    # Reapply visual styles and lock status in WM Group Matching tab
    app.ui_manager.reapply_lock_status(app)
