import os
import subprocess
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tksheet import Sheet, num2alpha as n2a
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

import ttkbootstrap as ttk
from ttkbootstrap.constants import *

from src.core.fp_utils import *
from src.views.widget.widget import StateObserver

# from src.views.widget.sheet_utils import SheetSearchManager


class SearchManager:
    def __init__(self, parent_frame, sheet_widget):
        """
        Initialize the SearchManager.
        :param parent_frame: The parent frame where the search UI components will be added.
        :param sheet_widget: The tksheet widget to which search functionality will be applied.
        """
        self.sheet = sheet_widget
        self.parent_frame = parent_frame

        # Create search UI components
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(
            parent_frame, textvariable=self.search_var, width=30
        )
        self.search_entry.pack(side=tk.LEFT, padx=5)

        self.search_button = ttk.Button(
            parent_frame, text="Search", command=self.search_sheet
        )
        self.search_button.pack(side=tk.LEFT, padx=5)

        self.clear_button = ttk.Button(
            parent_frame, text="Clear Search", command=self.clear_search
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

        self.noti_label = ttk.Label(
            parent_frame,
            text="Dynamo 에서 물량 산출을 하기 전, 반드시 B-note의 현재 상태를 저장해 주세요!",
            font=("맑은고딕", 14, "bold"),
            foreground="red",
        )
        self.noti_label.pack(side=tk.LEFT, padx=5)

        # Bind Enter key to search functionality
        self.search_entry.bind("<Return>", lambda event: self.search_sheet())

    def search_sheet(self):
        """Search for rows containing the search term."""
        search_term = self.search_var.get().strip().lower()
        if not search_term:
            self.sheet.display_rows("all")  # Show all rows if search is empty
            self.sheet.redraw()  # Refresh the sheet immediately
            return

        sheet_data = self.sheet.get_sheet_data()
        matching_rows = [
            rn
            for rn, row in enumerate(sheet_data)
            if any(search_term in str(cell).lower() for cell in row)
        ]

        self.sheet.display_rows(rows=matching_rows, all_displayed=False)
        self.sheet.redraw()  # Refresh the sheet immediately
        print(f"Search term '{search_term}' found in rows: {matching_rows}")

    def clear_search(self):
        """Clear the search field and reset the displayed rows."""
        self.search_var.set("")
        self.sheet.display_rows("all")
        self.sheet.redraw()  # Refresh the sheet immediately


class SearchManager_report_Group(SearchManager):
    def __init__(self, parent_frame, sheet_widget):
        # super().__init__(parent_frame, sheet_widget)
        self.sheet = sheet_widget
        self.parent_frame = parent_frame

        # Create search UI components
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(
            parent_frame, textvariable=self.search_var, width=30
        )
        self.search_entry.pack(side=tk.LEFT, padx=5)

        self.search_button = ttk.Button(
            # parent_frame, text="Search", command=self.search_and_scroll
            parent_frame,
            text="Search",
            command=self.cycle_through_matches,
        )
        self.search_button.pack(side=tk.LEFT, padx=5)

        self.clear_button = ttk.Button(
            parent_frame, text="Clear Search", command=self.clear_search
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

        # Bind Enter key to search and cycle through matches
        self.search_entry.bind("<Return>", lambda event: self.cycle_through_matches())

        # Initialize match tracking variables
        self.matched_rows = []
        self.current_match_index = 0
        self.previous_query = ""

    def search_and_scroll(self):
        query = self.search_entry.get().strip().lower()
        if not query:
            return  # No search if the query is empty

        self.previous_query = query

        data = self.sheet.get_sheet_data()
        # self.matched_rows.clear()
        # self.current_match_index = 0  # Reset index

        # Clear previous highlights
        # self.sheet.highlight_rows(rows=[], bg="white")
        self.sheet.dehighlight_all()

        # Find matching rows
        for index, row in enumerate(data):
            if any(query in str(cell).lower() for cell in row):
                self.matched_rows.append(index)

        if self.matched_rows:
            # Highlight matched rows
            self.sheet.highlight_rows(rows=self.matched_rows, bg="lightblue")

            # Scroll to the first match
            # self.sheet.set_yview(self.matched_rows[0])
        else:
            print("No matches found.")

    def cycle_through_matches(self):
        if self.previous_query != self.search_entry.get().strip().lower():
            self.matched_rows.clear()
            self.current_match_index = 0  # Reset index
            self.sheet.dehighlight_all()

            self.search_and_scroll()
        if not self.matched_rows:
            self.sheet.set_yview(0)
            return

        # Scroll to the next matched row
        self.sheet.set_yview(
            (self.matched_rows[self.current_match_index] - 1)
            / len(self.sheet.get_sheet_data())
        )

        # Update the index for the next scroll, wrapping around if necessary
        self.current_match_index = (self.current_match_index + 1) % len(
            self.matched_rows
        )
        # self.current_match_index = self.current_match_index + 1

    def clear_search(self):
        """Clear the search field and reset the displayed rows."""
        self.search_var.set("")
        self.matched_rows.clear()
        self.current_match_index = 0  # Reset index
        self.previous_query = ""
        self.sheet.display_rows("all")
        self.sheet.dehighlight_all()
        self.sheet.redraw()  # Refresh the sheet immediately
        self.sheet.set_yview(0)


class ReportMember_SheetWidget(ttk.Frame):
    def __init__(self, state, parent, data_kind=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.state = state
        self.data_kind = data_kind
        self.data = []
        self.column_headers = []

        # Load data
        self.load_data()
        try:
            self.manual_data = self.get_manula_item_data()
        except:
            self.manual_data = [[]]
        self.total_data = self.data + self.manual_data

        # Create a frame for the top controls (button)
        self.control_frame = ttk.Frame(parent)
        self.control_frame.pack(fill=tk.X, pady=5)

        # Add "Clear Filters" button at the top
        self.clear_button = ttk.Button(
            self.control_frame, text="Clear Filters", command=self.clear_filters
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

        # Initialize tksheet widget
        self.sheet = Sheet(
            self,
            # data=self.data,
            data=self.total_data,
            column_width=180,
            # theme="dark",
            # theme="light green",
            height=500,
            width=1000,
        )
        self.sheet.enable_bindings(
            "select_all",
            "single_select",
            "drag_select",
            "row_select",
            "copy",
            "rc_select",
            "arrowkeys",
            "double_click_column_resize",
            "column_width_resize",
            "column_select",
            "row_height_resize",
        )

        self.sheet.set_options(header_font=("Arial Narrow", 8, "normal"))
        self.sheet.set_options(font=("Arial Narrow", 8, "normal"))
        self.sheet.set_options(set_all_heights_and_widths=True)
        self.set_colums_widths()
        # self.sheet.set_all_cell_sizes_to_text()

        self.sheet.pack(fill=tk.BOTH, expand=True)

        # Add header dropdowns
        self.add_dropdowns()

        # Add search functionality
        self.search_manager = SearchManager(self.control_frame, self.sheet)

        # Create Extract Excel Button
        export_btn = ttk.Button(
            self.control_frame,
            text="Export to Excel",
            command=self.export_tksheet_to_excel,
            bootstyle="success-outline",
        )
        export_btn.pack(side="left", padx=100)

        # 상태 변경 감지를 위한 옵저버 설정
        self.state_observer = StateObserver(state, lambda e: self.update(e))

    def set_colums_widths(self, hdr_widths=None):
        # self.sheet.headers(self.column_headers)
        if hdr_widths:
            self.sheet.set_column_widths(hdr_widths)
        else:
            self.sheet.set_column_widths(
                [
                    60,  # 카테고리
                    70,  # 표준타입번호
                    140,  # 표준타입
                    40,  # 분류
                    200,  # name
                    40,  # GUID
                    # 150,  # GUID
                    180,  # 상세분류
                    100,  # wm_code
                    50,  # gauge
                    150,  # description
                    360,  # Spec.
                    # 540,  # Spec.
                    180,  # Add_spec.
                    150,  # 수식
                    200,  # 대입수식
                    50,  # 산출결과
                    50,  # 단위
                    50,  # 산출유형
                    200,  # Note
                    200,  # 산출로그
                ]
            )
        self.sheet.set_options(
            default_row_height=33,
        )

    def get_manula_item_data(self):
        ## dynamo 로직 준용
        def calc_formula(formula, calc_param_list):

            try:
                tmp_formula = go(
                    formula.split("\n"),
                    filter(lambda x: "#" not in x),
                    list,
                )[0]
            except:
                tmp_formula = ""

            if (
                tmp_formula != "=Exca"
                and tmp_formula != "=Back"
                and tmp_formula != "=Disp"
            ):
                for n, v in calc_param_list:
                    tmp_formula = tmp_formula.replace(n, str(v))
                try:
                    calc_result = eval(tmp_formula.strip("=")), "계산 성공"
                    # calc_result = float(eval(tmp_formula.strip("="))), "계산 성공"
                except:
                    calc_result = 0, "계산 실패 - 수식 혹은 파라미터가 유효하지 않음"
            else:
                tmp_formula = formula
                calc_result = (
                    0,
                    "토공 항목 - [H_PAB.RT.Q2A]_Revit 토공 물량 자동산출.dyn에서 산출 요망",
                )

            return {
                "수식": "'" + formula,
                "대입수식": "'" + tmp_formula,
                "산출결과": calc_result[0],
                "산출로그": calc_result[1],
            }

        state = self.state

        maunal_node = go(
            state.team_std_info["project-assigntype"],
            lambda x: x.get("children"),
            filter(lambda x: x["values"][1] == state.current_building.get()),
            filter(lambda x: "14." in x["values"][2].split("|")[1].strip()),
            # map(lambda x: x["children"]),
            list,
        )
        print(f"maunal_node:: {maunal_node}")

        wm_dict_hdrs = [
            "분류",
            "표준타입",
            "상세분류",
            "wm",
            "gauge",
            # "add_spec",
            "Add_spec.",
            "수식",
            "단위",
            "산출유형",
            "Note",
        ]
        flatten_hdrs = [
            "카테고리",
            "표준타입번호",
            "표준타입",
            "분류",
            "name",
            "GUID",
            "상세분류",
            "wm_code",
            "gauge",
            "Description",
            "Spec.",
            "Add_spec.",
            "수식",
            "대입수식",
            "산출결과",
            "단위",
            "산출유형",
            "Note",
            "산출로그",
        ]

        def calc_and_set_data(node_):
            node = deepcopy(node_)

            cat = node["values"][2].split("|")[1].strip()
            std_type_no = node["values"][2].split("|")[2].strip()
            name = node["name"]

            wm_list = go(
                node["children"],
                list,
            )

            wm_dicts = go(
                wm_list,
                deepcopy,
                map(lambda x: dict(zip(wm_dict_hdrs, x))),
                list,
            )
            for wm_dict in wm_dicts:
                if "Note" not in wm_dict.keys():
                    wm_dict.update({"Note": ""})
            # print(f"wm_dicts:: {wm_dicts}")

            res = []
            for wm_dict in wm_dicts:
                calcType_no = wm_dict["산출유형"].strip()
                calcdict = go(
                    state.team_std_info["std-calcdict"]["children"],
                    filter(lambda x: x["name"] == calcType_no),
                    list,
                )[0]
                # print(f"calcdict:: {calcdict}")
                param_list = go(
                    calcdict["children"],
                    list,
                    map(lambda x: x["values"]),
                    list,
                    map(lambda x: list(x)[1:]),
                    lambda x: sorted(x, key=lambda x: len(str(x[0])), reverse=True),
                    list,
                )

                wm_dict.update(calc_formula(wm_dict["수식"], param_list))
                wm_code = wm_dict["wm"].split("|")[0].strip()

                try:
                    wm_desc = wm_dict["wm"].split(" | ")[7]
                except:
                    wm_desc = ""

                wm_spec = go(
                    wm_dict["wm"].split(" | ")[9:-4],
                    filter(lambda x: not x.isnumeric()),
                    filter(
                        lambda x: not (
                            ("(   )" in x) or ("(   )" in x) or ("(  )" in x)
                        )
                    ),
                    lambda x: "\n".join(x),
                )

                wm_dict.update(
                    {
                        "name": name,
                        "GUID": "수동산출항목",
                        "카테고리": cat,
                        "표준타입번호": std_type_no,
                        "wm_code": wm_code,
                        "Description": wm_desc,
                        "Spec.": wm_spec,
                        "산출로그": "계산 성공",
                    }
                )
                row = []
                try:
                    for i in flatten_hdrs:
                        row.append(wm_dict[i])
                except:
                    row = [
                        "",
                        "",
                        wm_dict["표준타입"],
                        "",
                        wm_dict["name"],
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                    ]
                res.append(row)
            return res
            # res_.append(wm_dict)

        res = []
        for node_ in maunal_node:
            try:
                rows = calc_and_set_data(node_)
                for row in rows:
                    res.append(row)
            except:
                log = [
                    "",
                    "",
                    "",
                    "",
                    node_["name"],
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                ]
                res.append(log)
        return res

    def update(self, event=None):
        """Update the Sheet widget whenever the state changes."""
        state = self.state
        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")
        print(f"{self.__class__.__name__} > update 메소드 시작")

        # Reload data
        self.load_data()

        try:
            self.manual_data = self.get_manula_item_data()
        except:
            self.manual_data = [[]]
        self.total_data = self.data + self.manual_data

        # Set headers and data
        self.sheet.set_sheet_data(
            # self.data, reset_col_positions=True, reset_row_positions=True
            self.total_data,
            reset_col_positions=True,
            reset_row_positions=True,
        )
        self.sheet.headers(self.column_headers)

        # Reinitialize dropdowns
        self.add_dropdowns()

        self.clear_filters()

        # Ensure cell sizes and redraw
        # self.sheet.set_all_cell_sizes_to_text()
        self.set_colums_widths()
        self.sheet.redraw()

        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")
        print(f"{self.__class__.__name__} > update 메소드 종료")

    def load_data(self):
        state = self.state
        """Load data from Excel file and set headers and data."""
        try:
            # workbook = openpyxl.load_workbook(self.excel_file, data_only=True)
            # sheet = workbook.active
            data = self.data
            # Extract headers
            src_data = state.team_std_info[self.data_kind].get(
                state.current_building.get()
            )
            self.column_headers = src_data[0]
            # self.sheet.headers(self.column_headers)

            # Extract data
            self.data = src_data[1:]
            # self.data = src_data

            # self.sheet.set_sheet_data(self.data)
            print("ok~")
        except Exception as e:
            # self.sheet.set_sheet_data([])
            self.data = []
            print(f"Error loading data: {e}")

    def add_dropdowns(self, target_col_index=None):
        """Add dropdowns dynamically to each header based on the displayed rows."""
        current_headers = self.sheet.headers()  # Get current header values
        displayed_data = self.sheet.get_sheet_data(
            get_displayed=True
        )  # Get only displayed rows
        # print(f"Displayed data for dropdown update: {displayed_data}")  # Debugging

        if target_col_index is not None:
            # Update only the dropdown for the selected column
            header = self.column_headers[target_col_index]
            unique_values = ["all"] + sorted(
                set(
                    row[target_col_index]
                    for row in displayed_data
                    if row[target_col_index] is not None
                )
            )
            print(
                f"Updating dropdown for column {header} with values: {unique_values}"
            )  # Debugging
            current_value = (
                current_headers[target_col_index]
                if target_col_index < len(current_headers)
                else "all"
            )
            self.sheet.dropdown(
                self.sheet.span(n2a(target_col_index), header=True, table=False),
                values=unique_values,
                set_value=current_value,
                selection_function=lambda event, col=target_col_index: self.header_dropdown_selected(
                    event, col
                ),
                text=header,
            )
        else:
            # Update all dropdowns (initial setup or full refresh)
            for col_idx, header in enumerate(self.column_headers):
                unique_values = ["all"] + sorted(
                    set(
                        row[col_idx]
                        for row in displayed_data
                        if row[col_idx] is not None
                    )
                )
                # print(
                #     f"Updating dropdown for column {header} with values: {unique_values}"
                # )  # Debugging
                current_value = (
                    current_headers[col_idx]
                    if col_idx < len(current_headers)
                    else "all"
                )
                self.sheet.dropdown(
                    self.sheet.span(n2a(col_idx), header=True, table=False),
                    values=unique_values,
                    set_value=current_value,
                    selection_function=lambda event, col=col_idx: self.header_dropdown_selected(
                        event, col
                    ),
                    text=header,
                )

    def header_dropdown_selected(self, event, col_index):
        """Filter rows based on header dropdown selection."""
        hdrs = self.sheet.headers()
        hdrs[event.loc] = event.value

        if all(dd == "all" for dd in hdrs):  # No filters applied
            self.sheet.display_rows("all")
        else:  # Apply filters
            rows = [
                rn
                for rn, row in enumerate(self.total_data)
                if all(row[c] == e or e == "all" for c, e in enumerate(hdrs))
            ]
            self.sheet.display_rows(rows=rows, all_displayed=False)
            # self.sheet.display_rows(rows=rows, all_rows_displayed=False)
            self.add_dropdowns()
            self.sheet.redraw()

        # Recompute only the affected dropdown
        # self.add_dropdowns()
        # Redraw only once
        self.set_colums_widths()
        self.sheet.redraw()

    def clear_filters(self):
        """Clear all filters and reset the table to show all data."""
        # Reset all headers to "all"
        hdrs = ["all"] * len(self.column_headers)
        self.sheet.headers(hdrs)

        # Display all rows
        self.sheet.display_rows(
            "all", reset_col_positions=True, reset_row_positions=True
        )

        # Recompute dropdown options based on the full dataset
        # self.update()
        self.add_dropdowns()

        # Ensure everything is redrawn
        self.set_colums_widths()
        self.sheet.redraw()

    def export_tksheet_to_excel(self):
        """tksheet 위젯 데이터를 엑셀 파일로 저장하며 서식도 적용"""
        state = self.state
        sheet = self.sheet

        brief_pjtName = state.team_std_info.get("project-info").get("abbr")
        brief_buildingName = go(
            state.current_building.get(),
            lambda x: x.split(" "),
            map(lambda x: x[0]),
            lambda x: "".join(x),
        )

        # 파일 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save as Excel",
            initialfile=f"{brief_pjtName}_{brief_buildingName}-Members Report",
        )

        if not file_path:  # 사용자가 취소한 경우
            return

        # 새로운 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = f"{brief_buildingName}-Members Report"

        # 헤더 가져오기
        headers = sheet.headers()

        # 1. 엑셀 첫 번째 행에 헤더 입력
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)

            # 헤더 스타일 (Bold, 가운데 정렬)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # 헤더 배경색 (회색)
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

        # 2. tksheet 데이터 가져오기 (헤더 제외)
        data = sheet.get_sheet_data()
        column_widths = sheet.get_column_widths()

        # tksheet 셀 스타일 가져오기 (배경색, 글꼴, 정렬)
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 정렬 설정
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                # 기본 폰트 설정 (Calibri, Bold)
                cell.font = Font(name="Calibri", size=9, bold=False)

        # 컬럼 너비 조정
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
                width / 7
            )  # Excel 너비 조정

        # 엑셀 파일 저장
        wb.save(file_path)
        print(f"엑셀 파일 저장 완료: {file_path}")

        # 6. 메시지 박스 띄우기
        open_file = messagebox.askyesno(
            "저장 완료", "엑셀 파일 저장이 완료되었습니다.\n파일을 여시겠습니까?"
        )

        # 7. 사용자가 "예(Yes)"를 선택한 경우 엑셀 파일 열기
        if open_file:
            try:
                os.startfile(file_path)  # Windows에서 엑셀 파일 열기
                # subprocess.Popen(["start", file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다.\n{e}")


class ReportGroup_SheetWidget(ttk.Frame):
    def __init__(self, state, parent, data_kind=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.state = state
        self.data_kind = data_kind
        self.data = []
        self.column_headers = []

        # Load data
        self.load_data()
        try:
            self.manual_data = self.get_manula_item_data()
        except:
            self.manual_data = [[]]
        self.total_data = self.data + self.manual_data
        print(f"self.total_data::{self.total_data[0]}")

        # Formatting to Grouped form
        self.grped_total_data = self.format_data(self.total_data)

        # Create a frame for the top controls (button)
        self.control_frame = ttk.Frame(parent)
        self.control_frame.pack(fill=tk.X, pady=5)

        # Initialize tksheet widget
        self.sheet = Sheet(
            self,
            # data=[],
            # data=self.total_data,
            data=self.grped_total_data,
            column_width=180,
            # theme="dark",
            # theme="light green",
            height=500,
            width=1000,
        )
        self.sheet.enable_bindings(
            "select_all",
            "single_select",
            "drag_select",
            "row_select",
            "copy",
            "rc_select",
            "arrowkeys",
            "double_click_column_resize",
            "column_width_resize",
            "column_select",
            "row_height_resize",
        )

        self.sheet.set_options(header_font=("Arial Narrow", 8, "normal"))
        self.sheet.set_options(font=("Arial Narrow", 9, "normal"))

        self.sheet.pack(fill=tk.BOTH, expand=True)

        # Add search functionality
        # self.search_manager = SearchManager(self.control_frame, self.sheet)
        self.search_manager = SearchManager_report_Group(self.control_frame, self.sheet)

        self.only_types_btn = ttk.Button(
            self.control_frame,
            text="show only types",
            command=lambda: self.update(mode="only_types"),
            bootstyle="info",
        )
        self.only_types_btn.pack(side="left", anchor="w", padx=5)
        self.reset_btn = ttk.Button(
            self.control_frame,
            text="reset view",
            command=lambda: self.update(),
            bootstyle="info",
        )
        self.reset_btn.pack(side="left", anchor="w", padx=5)

        # Create Extract Excel Button
        export_btn = ttk.Button(
            self.control_frame,
            text="Export to Excel",
            command=self.export_tksheet_to_excel,
            bootstyle="success-outline",
        )
        export_btn.pack(side="left", padx=100)

        # 상태 변경 감지를 위한 옵저버 설정
        self.state_observer = StateObserver(state, lambda e: self.update(e))

        self.column_headers = [
            "카테고리",
            "표준타입번호",
            "표준타입",
            "name",
            "타입별 부재수",
            "분류",
            "상세분류",
            "wm_code",
            "gauge",
            "description",
            "Spec.",
            "Add_spec.",
            "수식",
            "대입수식",
            "산출결과",
            "단위",
            "산출유형",
            "Note",
            "산출로그",
        ]

        self.sheet.headers(self.column_headers)

        # # Ensure cell sizes and redraw
        self.set_colums_widths()

    def set_colums_widths(self, hdr_widths=None):
        # self.sheet.headers(self.column_headers)
        if hdr_widths:
            self.sheet.set_column_widths(hdr_widths)
        else:
            self.sheet.set_column_widths(
                [
                    60,  # 카테고리
                    70,  # 표준타입번호
                    140,  # 표준타입
                    200,  # name
                    70,  # 타입별 부재수
                    # 150,  # 타입별 부재수
                    40,  # 분류
                    180,  # 상세분류
                    100,  # wm_code
                    50,  # gauge
                    160,  # description
                    380,  # Spec.
                    # 540,  # Spec.
                    180,  # Add_spec.
                    150,  # 수식
                    200,  # 대입수식
                    50,  # 산출결과
                    50,  # 단위
                    50,  # 산출유형
                    200,  # Note
                    200,  # 산출로그
                ]
            )
        self.sheet.set_options(
            default_row_height=33,
        )

    def update(self, event=None, mode=None):
        """Update the Sheet widget whenever the state changes."""
        state = self.state
        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")
        print(f"{self.__class__.__name__} > update 메소드 시작")

        # Reload data
        self.load_data()

        try:
            self.manual_data = self.get_manula_item_data()
        except:
            self.manual_data = [[]]
        self.total_data = self.data + self.manual_data

        # Formatting to Grouped form
        if mode == "only_types":
            self.grped_total_data = self.format_data_onlyTypes(self.total_data)
        elif not mode:
            self.grped_total_data = self.format_data(self.total_data)

        # Set headers and data
        self.sheet.set_sheet_data(
            # self.data, reset_col_positions=True, reset_row_positions=True
            self.grped_total_data,
            reset_col_positions=True,
            reset_row_positions=True,
        )
        # self.sheet.headers(self.column_headers)

        # # Ensure cell sizes and redraw
        self.set_colums_widths()
        self.sheet.redraw()

        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")
        print(f"{self.__class__.__name__} > update 메소드 종료")

    def format_data_onlyTypes(self, data):
        """
        Processes hierarchical tabular data and groups it by the first two columns.
        """
        try:
            sorted_data = go(
                data,
                lambda x: sorted(x, key=lambda c: float(c[1])),
            )
            # print(f"sorted_data:: {sorted_data}")
            grouped_data = {}

            # for row in structured_data:
            for row in sorted_data:
                (category, sub1, sub2, sub3, sub4, sub5) = (
                    row[0],  # 카테고리
                    row[1],  # 표준타입번호
                    row[2],  # 표준타입 명
                    row[4],  # revit type
                    row[5],  # GUID
                    row[3],  # 분류(GWM/SWM)
                )
                if category not in grouped_data:
                    grouped_data[category] = {}
                if sub1 not in grouped_data[category]:
                    grouped_data[category][sub1] = {}
                if sub2 not in grouped_data[category][sub1]:
                    grouped_data[category][sub1][sub2] = {}
                if sub3 not in grouped_data[category][sub1][sub2]:
                    grouped_data[category][sub1][sub2][sub3] = {}
                if sub4 not in grouped_data[category][sub1][sub2][sub3]:
                    grouped_data[category][sub1][sub2][sub3][sub4] = {}
                if sub5 not in grouped_data[category][sub1][sub2][sub3][sub4]:
                    grouped_data[category][sub1][sub2][sub3][sub4][sub5] = []
                grouped_data[category][sub1][sub2][sub3][sub4][sub5].append(row[6:])

        except:
            grouped_data = {}

        # Populate Excel with hierarchical data
        sheet = []
        for category, sub1s in grouped_data.items():
            sheet.append([category])  # First level header

            for sub1, sub2s in sub1s.items():
                sub2 = list(sub2s.keys())[0]
                sheet.append(["", sub1, sub2])  # Second level header

                # for sub2, sub3s in sub2s.items():
                for sub2, sub3s in sub2s.items():
                    # sheet.append(["", "", "", sub2])  # Third level header

                    for sub3, sub4s in sub3s.items():
                        sheet.append(
                            ["", "", "", sub3, len(sub4s.keys())]
                        )  # Third level header

        return sheet

    def format_data(self, data):
        """
        Processes hierarchical tabular data and groups it by the first two columns.
        """
        try:
            sorted_data = go(
                data,
                lambda x: sorted(x, key=lambda c: float(c[1])),
            )
            # print(f"sorted_data:: {sorted_data}")
            grouped_data = {}

            # for row in structured_data:
            for row in sorted_data:
                (category, sub1, sub2, sub3, sub4, sub5) = (
                    row[0],  # 카테고리
                    row[1],  # 표준타입번호
                    row[2],  # 표준타입 명
                    row[4],  # revit type
                    row[5],  # GUID
                    row[3],  # 분류(GWM/SWM)
                )
                if category not in grouped_data:
                    grouped_data[category] = {}
                if sub1 not in grouped_data[category]:
                    grouped_data[category][sub1] = {}
                if sub2 not in grouped_data[category][sub1]:
                    grouped_data[category][sub1][sub2] = {}
                if sub3 not in grouped_data[category][sub1][sub2]:
                    grouped_data[category][sub1][sub2][sub3] = {}
                if sub4 not in grouped_data[category][sub1][sub2][sub3]:
                    grouped_data[category][sub1][sub2][sub3][sub4] = {}
                    # grouped_data[category][sub1][sub2][sub3][sub4] = []
                if sub5 not in grouped_data[category][sub1][sub2][sub3][sub4]:
                    grouped_data[category][sub1][sub2][sub3][sub4][sub5] = []
                grouped_data[category][sub1][sub2][sub3][sub4][sub5].append(row[6:])

        except:
            grouped_data = {}

        print(f"grouped_data:: {grouped_data}")

        # Populate Excel with hierarchical data
        sheet = []
        for category, sub1s in grouped_data.items():
            sheet.append([category])  # First level header

            for sub1, sub2s in sub1s.items():
                sub2 = list(sub2s.keys())[0]
                sheet.append(["", sub1, sub2])  # Second level header

                # for sub2, sub3s in sub2s.items():
                for sub2, sub3s in sub2s.items():
                    # sheet.append(["", "", "", sub2])  # Third level header
                    for sub3, sub4s in sub3s.items():
                        sheet.append(
                            ["", "", "", sub3, len(sub4s.keys())]
                        )  # Third level header
                        sub4, sub5s = list(sub4s.items())[0]
                        for sub5, rows in sub5s.items():
                            sheet.append(["", "", "", "", "", sub5])  # Thir
                            tmp = []
                            for row in rows:
                                res_row = [
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    "",
                                    *row[:7],
                                    "",
                                    "",
                                    *row[9:11],
                                ]
                                if res_row not in tmp:
                                    tmp.append(res_row)
                                    sheet.append(res_row)  # Data rows indented

        return sheet

    # def format_data(self, data):
    #     """
    #     Processes hierarchical tabular data and groups it by the first two columns.
    #     """
    #     try:
    #         sorted_data = go(
    #             data,
    #             lambda x: sorted(x, key=lambda c: float(c[1])),
    #         )
    #         print(f"sorted_data:: {sorted_data}")
    #         grouped_data = {}

    #         # for row in structured_data:
    #         for row in sorted_data:
    #             category, sub1, sub2, sub3, sub4, sub5 = (
    #                 row[0],
    #                 row[1],
    #                 row[2],
    #                 row[4],
    #                 row[5],
    #                 row[3],
    #             )
    #             if category not in grouped_data:
    #                 grouped_data[category] = {}
    #             if sub1 not in grouped_data[category]:
    #                 grouped_data[category][sub1] = {}
    #             if sub2 not in grouped_data[category][sub1]:
    #                 grouped_data[category][sub1][sub2] = {}
    #             if sub3 not in grouped_data[category][sub1][sub2]:
    #                 grouped_data[category][sub1][sub2][sub3] = {}
    #             if sub4 not in grouped_data[category][sub1][sub2][sub3]:
    #                 grouped_data[category][sub1][sub2][sub3][sub4] = {}
    #             if sub5 not in grouped_data[category][sub1][sub2][sub3][sub4]:
    #                 grouped_data[category][sub1][sub2][sub3][sub4][sub5] = []
    #             grouped_data[category][sub1][sub2][sub3][sub4][sub5].append(row[6:])

    #     except:
    #         grouped_data = {}

    #     # Populate Excel with hierarchical data
    #     sheet = []
    #     for category, sub1s in grouped_data.items():
    #         sheet.append([category])  # First level header

    #         for sub1, sub2s in sub1s.items():
    #             sub2 = list(sub2s.keys())[0]
    #             sheet.append(["", sub1, sub2])  # Second level header

    #             # for sub2, sub3s in sub2s.items():
    #             for sub2, sub3s in sub2s.items():
    #                 # sheet.append(["", "", "", sub2])  # Third level header

    #                 for sub3, sub4s in sub3s.items():
    #                     sheet.append(["", "", "", sub3])  # Third level header

    #                     for sub4, sub5s in sub4s.items():
    #                         sheet.append(["", "", "", "", sub4])  # Third level header

    #                         for sub5, rows in sub5s.items():
    #                             sheet.append(
    #                                 ["", "", "", "", "", sub5]
    #                             )  # Third level header

    #                             for row in rows:
    #                                 sheet.append(
    #                                     ["", "", "", "", "", "", *row]
    #                                 )  # Data rows indented

    #     return sheet

    def get_manula_item_data(self):
        ## dynamo 로직 준용
        def calc_formula(formula, calc_param_list):

            try:
                tmp_formula = go(
                    formula.split("\n"),
                    filter(lambda x: "#" not in x),
                    list,
                )[0]
            except:
                tmp_formula = ""

            if (
                tmp_formula != "=Exca"
                and tmp_formula != "=Back"
                and tmp_formula != "=Disp"
            ):
                for n, v in calc_param_list:
                    tmp_formula = tmp_formula.replace(n, str(v))
                try:
                    calc_result = eval(tmp_formula.strip("=")), "계산 성공"
                except:
                    calc_result = 0, "계산 실패 - 수식 혹은 파라미터가 유효하지 않음"
            else:
                tmp_formula = formula
                calc_result = (
                    0,
                    "토공 항목 - [H_PAB.RT.Q2A]_Revit 토공 물량 자동산출.dyn에서 산출 요망",
                )

            return {
                "수식": "'" + formula,
                "대입수식": "'" + tmp_formula,
                "산출결과": calc_result[0],
                "산출로그": calc_result[1],
            }

        state = self.state

        maunal_node = go(
            state.team_std_info["project-assigntype"],
            lambda x: x.get("children"),
            filter(lambda x: x["values"][1] == state.current_building.get()),
            filter(lambda x: "14." in x["values"][2].split("|")[1].strip()),
            # map(lambda x: x["children"]),
            list,
        )
        print(f"maunal_node:: {maunal_node}")

        wm_dict_hdrs = [
            "분류",
            "표준타입",
            "상세분류",
            "wm",
            "gauge",
            # "add_spec",
            "Add_spec.",
            "수식",
            "단위",
            "산출유형",
            "Note",
        ]
        flatten_hdrs = [
            "카테고리",
            "표준타입번호",
            "표준타입",
            "분류",
            "name",
            "GUID",
            "상세분류",
            "wm_code",
            "gauge",
            "Description",
            "Spec.",
            "Add_spec.",
            "수식",
            "대입수식",
            "산출결과",
            "단위",
            "산출유형",
            "Note",
            "산출로그",
        ]

        def calc_and_set_data(node_):
            node = deepcopy(node_)

            cat = node["values"][2].split("|")[1].strip()
            std_type_no = node["values"][2].split("|")[2].strip()
            name = node["name"]

            wm_list = go(
                node["children"],
                list,
            )

            wm_dicts = go(
                wm_list,
                deepcopy,
                map(lambda x: dict(zip(wm_dict_hdrs, x))),
                list,
            )
            for wm_dict in wm_dicts:
                if "Note" not in wm_dict.keys():
                    wm_dict.update({"Note": ""})
            # print(f"wm_dicts:: {wm_dicts}")

            res = []
            for wm_dict in wm_dicts:
                calcType_no = wm_dict["산출유형"]
                calcdict = go(
                    state.team_std_info["std-calcdict"]["children"],
                    filter(lambda x: x["name"] == calcType_no),
                    list,
                )[0]
                # print(f"calcdict:: {calcdict}")
                param_list = go(
                    calcdict["children"],
                    list,
                    map(lambda x: x["values"]),
                    list,
                    map(lambda x: list(x)[1:]),
                    lambda x: sorted(x, key=lambda x: len(str(x[0])), reverse=True),
                    list,
                )

                wm_dict.update(calc_formula(wm_dict["수식"], param_list))
                wm_code = wm_dict["wm"].split("|")[0].strip()

                try:
                    wm_desc = wm_dict["wm"].split(" | ")[7]
                except:
                    wm_desc = ""

                wm_spec = go(
                    wm_dict["wm"].split(" | ")[9:-4],
                    filter(lambda x: not x.isnumeric()),
                    filter(
                        lambda x: not (
                            ("(   )" in x) or ("(   )" in x) or ("(  )" in x)
                        )
                    ),
                    lambda x: "\n".join(x),
                )

                wm_dict.update(
                    {
                        "name": name,
                        "GUID": "수동산출항목",
                        "카테고리": cat,
                        "표준타입번호": std_type_no,
                        "wm_code": wm_code,
                        "Description": wm_desc,
                        "Spec.": wm_spec,
                        "산출로그": "계산 성공",
                    }
                )
                row = []
                try:
                    for i in flatten_hdrs:
                        row.append(wm_dict[i])
                except:
                    row = [
                        "",
                        "",
                        wm_dict["표준타입"],
                        "",
                        wm_dict["name"],
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                    ]
                res.append(row)
            return res
            # res_.append(wm_dict)

        res = []
        for node_ in maunal_node:
            try:
                rows = calc_and_set_data(node_)
                for row in rows:
                    res.append(row)
            except:
                log = [
                    "",
                    "",
                    "",
                    "",
                    node_["name"],
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                ]
                res.append(log)
        return res

    def load_data(self):
        state = self.state
        """Load data from Excel file and set headers and data."""
        try:
            # Extract headers
            src_data = state.team_std_info[self.data_kind].get(
                state.current_building.get()
            )
            # self.column_headers = src_data[0]
            # self.sheet.headers(self.column_headers)

            # Extract data
            self.data = src_data[1:]
            # self.data = src_data

            # self.sheet.set_sheet_data(self.data)
            print("ok~")
        except Exception as e:
            # self.sheet.set_sheet_data([])
            self.data = []
            print(f"Error loading data: {e}")

    def export_tksheet_to_excel(self):
        """tksheet 위젯 데이터를 엑셀 파일로 저장하며 서식도 적용"""
        state = self.state
        sheet = self.sheet

        brief_pjtName = state.team_std_info.get("project-info").get("abbr")
        brief_buildingName = go(
            state.current_building.get(),
            lambda x: x.split(" "),
            map(lambda x: x[0]),
            lambda x: "".join(x),
        )

        # 파일 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save as Excel",
            initialfile=f"{brief_pjtName}_{brief_buildingName}-Types Report",
        )

        if not file_path:  # 사용자가 취소한 경우
            return

        # 새로운 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = f"{brief_buildingName}-Types Report"

        # 헤더 가져오기
        headers = sheet.headers()

        # 1. 엑셀 첫 번째 행에 헤더 입력
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)

            # 헤더 스타일 (Bold, 가운데 정렬)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # 헤더 배경색 (회색)
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

        # 2. tksheet 데이터 가져오기 (헤더 제외)
        data = sheet.get_sheet_data()
        column_widths = sheet.get_column_widths()

        # tksheet 셀 스타일 가져오기 (배경색, 글꼴, 정렬)
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 정렬 설정
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                # 기본 폰트 설정 (Calibri, Bold)
                cell.font = Font(name="Calibri", size=9, bold=False)

        # 컬럼 너비 조정
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
                width / 7
            )  # Excel 너비 조정

        # 엑셀 파일 저장
        wb.save(file_path)
        print(f"엑셀 파일 저장 완료: {file_path}")

        # 6. 메시지 박스 띄우기
        open_file = messagebox.askyesno(
            "저장 완료", "엑셀 파일 저장이 완료되었습니다.\n파일을 여시겠습니까?"
        )

        # 7. 사용자가 "예(Yes)"를 선택한 경우 엑셀 파일 열기
        if open_file:
            try:
                os.startfile(file_path)  # Windows에서 엑셀 파일 열기
                # subprocess.Popen(["start", file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다.\n{e}")


class ReportBuildingTotal_SheetWidget(ttk.Frame):
    def __init__(self, state, parent, data_kind=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.state = state
        self.data_kind = data_kind
        self.data = []
        self.column_headers = []

        # Load data
        self.load_data()
        try:
            self.manual_data = self.get_manula_item_data()
        except:
            self.manual_data = [[]]
        self.total_data = self.data + self.manual_data

        # Formatting to Total BOQ form
        self.building_total_data = self.format_data(self.total_data)

        # Create a frame for the top controls (button)
        self.control_frame = ttk.Frame(parent)
        self.control_frame.pack(fill=tk.X, pady=5)

        # Initialize tksheet widget
        self.sheet = Sheet(
            self,
            # data=[],
            # data=self.total_data,
            data=self.building_total_data,
            column_width=180,
            height=500,
            width=1000,
        )
        self.sheet.enable_bindings(
            "select_all",
            "single_select",
            "drag_select",
            "row_select",
            "copy",
            "rc_select",
            "arrowkeys",
            "double_click_column_resize",
            "column_width_resize",
            "column_select",
            "row_height_resize",
        )

        self.sheet.set_options(header_font=("Arial Narrow", 8, "normal"))
        self.sheet.set_options(font=("Arial Narrow", 9, "normal"))

        self.sheet.pack(fill=tk.BOTH, expand=True)

        # Add search functionality
        self.search_manager = SearchManager(self.control_frame, self.sheet)

        # Create Extract Excel Button
        export_btn = ttk.Button(
            self.control_frame,
            text="Export to Excel",
            command=self.export_tksheet_to_excel,
            bootstyle="success-outline",
        )
        export_btn.pack(side="left", padx=100)

        # 상태 변경 감지를 위한 옵저버 설정
        self.state_observer = StateObserver(state, lambda e: self.update(e))

        self.column_headers = [
            "Work Master Code",
            "Gauge Code",
            "Description",
            "Spec.",
            "Additional Spec.",
            "Reference to",
            "UoM",
            f"{state.current_building.get()}",
        ]

        self.sheet.headers(self.column_headers)

        # # Ensure cell sizes and redraw
        self.set_colums_widths()

        self.highlight_category_row()

    def set_colums_widths(self, hdr_widths=None):
        # self.sheet.headers(self.column_headers)
        if hdr_widths:
            self.sheet.set_column_widths(hdr_widths)
        else:
            self.sheet.set_column_widths(
                [
                    110,  # Work Master Code
                    50,  # Gauge Code
                    250,  # Description
                    380,  # Spec.
                    250,  # Additional Spec.
                    50,  # Reference to
                    50,  # UoM
                    200,  # Current Building
                ]
            )
        self.sheet.set_options(
            default_row_height=33,
        )

    def load_data(self):
        state = self.state
        """Load data from Excel file and set headers and data."""
        if state.current_building.get() == "All":
            all_buildings = go(
                state.team_std_info["project-buildinglist"]["children"],
                map(lambda x: x["name"]),
                list,
            )
            print(f"all_buildings_load_data::{all_buildings}")
            self.data = []
            # src_data = []
            for bd in all_buildings:
                try:
                    self.data.append(
                        state.team_std_info[self.data_kind].get(bd)[1:],
                    )
                except:
                    self.data.append([])
        else:
            try:
                # Extract headers
                src_data = state.team_std_info[self.data_kind].get(
                    state.current_building.get()
                )

                # Extract data
                self.data = src_data[1:]
                # self.data = src_data

                # self.sheet.set_sheet_data(self.data)
                print("ok~")
            except Exception as e:
                # self.sheet.set_sheet_data([])
                self.data = []
                print(f"Error loading data: {e}")

    def get_manula_item_data(self, cur_bd=None):
        ## dynamo 로직 준용
        def calc_formula(formula, calc_param_list):

            try:
                tmp_formula = go(
                    formula.split("\n"),
                    filter(lambda x: "#" not in x),
                    list,
                )[0]
            except:
                tmp_formula = ""

            if (
                tmp_formula != "=Exca"
                and tmp_formula != "=Back"
                and tmp_formula != "=Disp"
            ):
                for n, v in calc_param_list:
                    tmp_formula = tmp_formula.replace(n, str(v))
                try:
                    calc_result = eval(tmp_formula.strip("=")), "계산 성공"
                except:
                    calc_result = 0, "계산 실패 - 수식 혹은 파라미터가 유효하지 않음"
            else:
                tmp_formula = formula
                calc_result = (
                    0,
                    "토공 항목 - [H_PAB.RT.Q2A]_Revit 토공 물량 자동산출.dyn에서 산출 요망",
                )

            return {
                "수식": "'" + formula,
                "대입수식": "'" + tmp_formula,
                "산출결과": calc_result[0],
                "산출로그": calc_result[1],
            }

        state = self.state

        maunal_node = go(
            state.team_std_info["project-assigntype"],
            lambda x: x.get("children"),
            filter(
                lambda x: (
                    x["values"][1]
                    == (cur_bd if cur_bd else state.current_building.get())
                )
            ),
            filter(lambda x: "14." in x["values"][2].split("|")[1].strip()),
            # map(lambda x: x["children"]),
            list,
        )
        print(f"maunal_node:: {maunal_node}")

        wm_dict_hdrs = [
            "분류",
            "표준타입",
            "상세분류",
            "wm",
            "gauge",
            # "add_spec",
            "Add_spec.",
            "수식",
            "단위",
            "산출유형",
            "Note",
        ]
        flatten_hdrs = [
            "카테고리",
            "표준타입번호",
            "표준타입",
            "분류",
            "name",
            "GUID",
            "상세분류",
            "wm_code",
            "gauge",
            "Description",
            "Spec.",
            "Add_spec.",
            "수식",
            "대입수식",
            "산출결과",
            "단위",
            "산출유형",
            "Note",
            "산출로그",
        ]

        def calc_and_set_data(node_):
            node = deepcopy(node_)

            cat = node["values"][2].split("|")[1].strip()
            std_type_no = node["values"][2].split("|")[2].strip()
            name = node["name"]

            wm_list = go(
                node["children"],
                list,
            )

            wm_dicts = go(
                wm_list,
                deepcopy,
                map(lambda x: dict(zip(wm_dict_hdrs, x))),
                list,
            )
            for wm_dict in wm_dicts:
                if "Note" not in wm_dict.keys():
                    wm_dict.update({"Note": ""})
            # print(f"wm_dicts:: {wm_dicts}")

            res = []
            for wm_dict in wm_dicts:
                calcType_no = wm_dict["산출유형"]
                calcdict = go(
                    state.team_std_info["std-calcdict"]["children"],
                    filter(lambda x: x["name"] == calcType_no),
                    list,
                )[0]
                # print(f"calcdict:: {calcdict}")
                param_list = go(
                    calcdict["children"],
                    list,
                    map(lambda x: x["values"]),
                    list,
                    map(lambda x: list(x)[1:]),
                    lambda x: sorted(x, key=lambda x: len(str(x[0])), reverse=True),
                    list,
                )

                wm_dict.update(calc_formula(wm_dict["수식"], param_list))
                wm_code = wm_dict["wm"].split("|")[0].strip()

                try:
                    wm_desc = wm_dict["wm"].split(" | ")[7]
                except:
                    wm_desc = ""

                wm_spec = go(
                    wm_dict["wm"].split(" | ")[9:-4],
                    filter(lambda x: not x.isnumeric()),
                    filter(
                        lambda x: not (
                            ("(   )" in x) or ("(   )" in x) or ("(  )" in x)
                        )
                    ),
                    lambda x: "\n".join(x),
                )

                wm_dict.update(
                    {
                        "name": name,
                        "GUID": "수동산출항목",
                        "카테고리": cat,
                        "표준타입번호": std_type_no,
                        "wm_code": wm_code,
                        "Description": wm_desc,
                        "Spec.": wm_spec,
                        "산출로그": "계산 성공",
                    }
                )
                row = []
                try:
                    for i in flatten_hdrs:
                        row.append(wm_dict[i])
                except:
                    row = [
                        "",
                        "",
                        wm_dict["표준타입"],
                        "",
                        wm_dict["name"],
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                    ]
                res.append(row)
            return res
            # res_.append(wm_dict)

        res = []
        for node_ in maunal_node:
            try:
                rows = calc_and_set_data(node_)
                for row in rows:
                    res.append(row)
            except:
                log = [
                    "",
                    "",
                    "",
                    "",
                    node_["name"],
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "계산 실패 - 수식 혹은 파라미터가 유효하지 않음",
                ]
                res.append(log)
        return res

    def update(self, event=None):
        """Update the Sheet widget whenever the state changes."""
        state = self.state
        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")
        print(f"{self.__class__.__name__} > update 메소드 시작")

        if state.current_building.get() == "All":
            all_buildings = state.team_std_info["project-buildinglist"]["children"]
            self.column_headers = [
                "Work Master Code",
                "Gauge Code",
                "Description",
                "Spec.",
                "Additional Spec.",
                "Reference to",
                "UoM",
                "Total",
                *[i["name"] for i in all_buildings],
            ]
            self.sheet.headers(self.column_headers)
            self.sheet.set_column_widths(
                [
                    110,  # Work Master Code
                    50,  # Gauge Code
                    250,  # Description
                    380,  # Spec.
                    250,  # Additional Spec.
                    50,  # Reference to
                    50,  # UoM
                    70,  # Total
                    *[200] * (len(all_buildings)),  # Current Building
                ]
            )
        else:
            self.column_headers = [
                "Work Master Code",
                "Gauge Code",
                "Description",
                "Spec.",
                "Additional Spec.",
                "Reference to",
                "UoM",
                f"{state.current_building.get()}",
            ]
            self.sheet.headers(self.column_headers)
            self.set_colums_widths()

        # Reload data
        self.load_data()
        print(f"self.data길이 {len(self.data)}")

        if state.current_building.get() == "All":
            all_buildings = go(
                state.team_std_info["project-buildinglist"]["children"],
                map(lambda x: x["name"]),
                list,
            )
            print(f"all_buildings::{all_buildings}")
            self.total_data = []
            for idx, bd in enumerate(all_buildings):
                try:
                    self.manual_data = self.get_manula_item_data(bd)
                except:
                    self.manual_data = [[]]
                self.total_data.append(self.data[idx] + self.manual_data)

        else:
            try:
                self.manual_data = self.get_manula_item_data()
            except:
                self.manual_data = [[]]
            self.total_data = self.data + self.manual_data
            self.sheet.set_column_widths()

        # Formatting to Total BOQ form
        if state.current_building.get() == "All":
            all_buildings_data = [self.format_data(i) for i in self.total_data]

            # 스틸 플레이트 물량 분개 구간

            first_building, *others = all_buildings_data

            building_total_data_ = []
            for rIdx, row in enumerate(first_building):
                new_row = deepcopy(row)
                for bd_data in others:
                    bd_last = bd_data[rIdx][-1]
                    new_row.append(bd_last)
                new_row_forTotal = deepcopy(new_row)

                res_forSum = []
                for i in new_row_forTotal[7:]:
                    try:
                        res_forSum.append(float(i))
                    except:
                        pass
                if sum(res_forSum):
                    new_row_qty_sum = round(sum(res_forSum), 3)
                else:
                    new_row_qty_sum = ""

                new_row.insert(7, new_row_qty_sum)
                building_total_data_.append(new_row)

            self.building_total_data = building_total_data_
        else:
            # 스틸 플레이트 물량 분개 구간
            formated_data = self.format_data(self.total_data)
            plate_itmes_amount = go(
                formated_data,
                filter(lambda x: x[0] == "S01AA009-00001"),
                map(lambda x: x[7] if x[7] else 0),
                sum,
            )
            print(f"plate_itmes_amount {plate_itmes_amount}")
            indig_tgt_wmcodes = [
                "S03AA081-00001",
                "S03AA082-00001",
                "S03AA083-00001",
                "S03AA084-00001",
                "S03AA085-00001",
                # "S03AA007-00001", #PnG
            ]

            indig_tgt_rows_amount = go(
                formated_data,
                filter(lambda x: x[0] in indig_tgt_wmcodes),
                map(lambda x: x[7] if x[7] else 0),
                sum,
            )
            print(f"indig_tgt_rows_amount {indig_tgt_rows_amount}")

            if plate_itmes_amount and indig_tgt_rows_amount:
                for row in formated_data:
                    if row[0] in indig_tgt_wmcodes:
                        row_origin_qty = deepcopy(row[7]) if row[7] else 0
                        print(f"row_origin_qty {row_origin_qty}")
                        add_qty_forRow = (
                            float(row_origin_qty) / indig_tgt_rows_amount
                        ) * plate_itmes_amount
                        row[7] = (
                            round(row_origin_qty + add_qty_forRow, 3)
                            if (row_origin_qty + add_qty_forRow)
                            else ""
                        )

            # 클래스 변수에 등록
            self.building_total_data = formated_data

        # Set headers and data
        self.sheet.set_sheet_data(
            # self.data, reset_col_positions=True, reset_row_positions=True
            self.building_total_data,
            reset_col_positions=True,
            reset_row_positions=True,
        )

        # # Ensure cell sizes and redraw

        self.highlight_category_row()

        self.sheet.redraw()

        if state.current_building.get() == "All":
            self.sheet.set_column_widths(
                [
                    110,  # Work Master Code
                    50,  # Gauge Code
                    250,  # Description
                    380,  # Spec.
                    250,  # Additional Spec.
                    50,  # Reference to
                    50,  # UoM
                    70,  # Total
                    *[200] * (len(all_buildings)),  # Current Building
                ]
            )
        else:
            self.sheet.set_column_widths(
                [
                    110,  # Work Master Code
                    50,  # Gauge Code
                    250,  # Description
                    380,  # Spec.
                    250,  # Additional Spec.
                    50,  # Reference to
                    50,  # UoM
                    200,  # Current Building
                ]
            )

        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")
        print(f"{self.__class__.__name__} > update 메소드 종료")

    def format_data(self, _data=None):
        """
        Processes Total BOQ format
        """
        state = self.state

        if _data:
            data = _data
        else:
            data = self.total_data

        if state.team_std_info.get(self.data_kind):
            unique_wm_inBuilding = go(
                data,
                filter(lambda x: x[7] != ""),
                map(
                    lambda x: "|".join(
                        [
                            str(x[7]),
                            str(x[8]),
                            str(x[9]),
                            str(x[10]),
                            str(x[11]),
                            "",
                            str(x[15]),
                        ]
                    )
                ),
                set,
                list,
                sorted,
                map(lambda x: x.split("|")),
                list,
            )
        else:
            unique_wm_inBuilding = []

        if state.team_std_info.get("project-assigntype"):
            unique_wm_inPjt_preModi = go(
                state.team_std_info.get("project-assigntype"),
                lambda x: x.get("children"),
                map(lambda x: x.get("children")),
                lambda x: list(chain(*x)),
            )
            # print(f"unique_wm_inPjt_preModi::{unique_wm_inPjt_preModi}")
            for i in unique_wm_inPjt_preModi:
                if "trowel" in i:
                    print(i)

            unique_wm_inPjt_ = []
            for r_ in unique_wm_inPjt_preModi:
                r = deepcopy(r_)
                wm_spec = go(
                    r[3].split(" | ")[9:-4],
                    filter(lambda x: not x.isnumeric()),
                    filter(
                        lambda x: not (
                            ("(   )" in x) or ("(   )" in x) or ("(  )" in x)
                        )
                    ),
                    lambda x: "\n".join(x),
                )
                try:
                    wm_desc = r[3].split("|")[7].strip()
                except:
                    wm_desc = ""
                new_r = [
                    r[3].split("|")[0].strip(),  #
                    r[4],  #
                    wm_desc,  # Description
                    wm_spec,  # Spec.
                    r[5],  # Additional Spec.
                    # "",  # Additional Spec.
                    "",  # Reference to
                    r[7],  # UoM
                ]
                unique_wm_inPjt_.append(new_r)
                # if "SteelTrowel" in r:
                #     print(r)
            # print(f"unique_wm_inPjt::{unique_wm_inPjt_[0]}")
            unique_wm_inPjt = go(
                unique_wm_inPjt_,
                map(lambda x: "|".join(x)),
                set,
                list,
                sorted,
                map(lambda x: x.split("|")),
                list,
            )
            # print(f"ttt:{unique_wm_inPjt}")

        else:
            unique_wm_inPjt = []

        if unique_wm_inBuilding != []:
            unique_wm_inBuilding_withCalc = []
            # for i in unique_wm_inBuilding:
            for i in unique_wm_inPjt:
                calc_sum = 0
                for row in data:
                    if row[7] == i[0] and row[8] == i[1] and row[11] == i[4]:
                        try:
                            value = row[14] if row[14] != "M2" else row[15]
                            calc_sum = calc_sum + value
                        except:
                            calc_sum = ""
                    # else:
                    #     calc_sum = ""
                # unique_wm_inBuilding_withCalc.append(i + [round(calc_sum, 3)])
                if calc_sum != 0:
                    unique_wm_inBuilding_withCalc.append(i + [round(calc_sum, 3)])
                else:
                    unique_wm_inBuilding_withCalc.append(i + [""])
        else:
            unique_wm_inBuilding_withCalc = []

        WMs_data = state.team_std_info.get("WMs", [])

        unique_Cat_large = go(
            WMs_data,
            list,
            map(lambda x: "|".join([str(x[2]), str(x[3])])),
            filter(lambda x: "None" not in x),
            set,
            list,
            sorted,
            map(lambda x: x.split("|")),
            list,
        )
        state.unique_Cat_large = unique_Cat_large
        print(f"unique_Cat_large:: {unique_Cat_large}")

        unique_Cat_middle = go(
            WMs_data,
            list,
            map(lambda x: "|".join([str(x[2]), str(x[3]), str(x[4]), str(x[5])])),
            filter(lambda x: "None" not in x),
            set,
            list,
            sorted,
            map(lambda x: x.split("|")),
            list,
        )
        state.unique_Cat_middle = unique_Cat_middle
        # print(f"unique_Cat_middle:: {unique_Cat_middle}")

        category_slot = []
        for large in unique_Cat_large:
            middle_grp = []
            for middle in unique_Cat_middle:
                if large[0] == middle[0]:
                    grp_row = []
                    for wm_row in unique_wm_inBuilding_withCalc:
                        # print(wm_row)
                        if "".join([middle[0], middle[2]]) in wm_row[0]:
                            if (
                                len(wm_row) > 8
                            ):  ## 구버전 데이터로 M2|M2 가 남아있는 경우 해결
                                wm_row[7] = wm_row[8]
                            grp_row.append(wm_row[:8])
                    if grp_row:
                        middle_grp.append(["", "", middle[3], "", "", "", "", ""])
                        middle_grp.extend(grp_row)
            if middle_grp:
                category_slot.append(["", "", large[1], "", "", "", "", ""])
                category_slot.extend(middle_grp)

        return category_slot

    def highlight_category_row(self):
        """Update row highlights based on checkbox values."""
        state = self.state

        uniq_cat_large = go(
            state.unique_Cat_middle,
            map(lambda x: x[1]),
            set,
            list,
        )

        uniq_cat_middle = go(
            state.unique_Cat_middle,
            map(lambda x: x[3]),
            set,
            list,
        )

        cat_large_rows = []
        for row in range(self.sheet.get_total_rows()):
            if (
                self.sheet.get_cell_data(row, 2) in uniq_cat_large
            ):  # If checkbox is checked
                cat_large_rows.append(row)

        cat_middle_rows = []
        for row in range(self.sheet.get_total_rows()):
            if (
                self.sheet.get_cell_data(row, 2) in uniq_cat_middle
            ):  # If checkbox is checked
                cat_middle_rows.append(row)

        # Clear existing highlights
        self.highlighted_rows = {}

        self.sheet.dehighlight_all()
        self.sheet.highlight_rows(
            cat_large_rows, highlight_index=False, bg="#D3F9D8"
        )  # Light green
        for row in cat_large_rows:
            self.highlighted_rows[row] = "#D3F9D8"
        for row in cat_large_rows:
            self.sheet.row_height(row, height=15)

        self.sheet.highlight_rows(
            cat_middle_rows, highlight_index=False, bg="#fffbd9"
        )  # Light yellow
        for row in cat_middle_rows:
            self.highlighted_rows[row] = "#fffbd9"
        for row in cat_middle_rows:
            self.sheet.row_height(row, height=15)

    def export_tksheet_to_excel(self):
        """tksheet 위젯 데이터를 엑셀 파일로 저장하며 서식도 적용"""
        state = self.state
        sheet = self.sheet

        brief_pjtName = state.team_std_info.get("project-info").get("abbr")
        brief_buildingName = go(
            state.current_building.get(),
            lambda x: x.split(" "),
            map(lambda x: x[0]),
            lambda x: "".join(x),
        )

        # 파일 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save as Excel",
            initialfile=f"{brief_pjtName}_{brief_buildingName}-BOQ",
        )

        if not file_path:  # 사용자가 취소한 경우
            return

        # 새로운 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = f"{brief_buildingName}-BOQ"

        # 헤더 가져오기
        headers = sheet.headers()

        # 1. 엑셀 첫 번째 행에 헤더 입력
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)

            # 헤더 스타일 (Bold, 가운데 정렬)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # 헤더 배경색 (회색)
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

        # 2. tksheet 데이터 가져오기 (헤더 제외)
        data = sheet.get_sheet_data()
        column_widths = sheet.get_column_widths()

        # tksheet 셀 스타일 가져오기 (배경색, 글꼴, 정렬)
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 정렬 설정
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

                # 기본 폰트 설정 (Calibri, Bold)
                cell.font = Font(name="Calibri", size=9, bold=False)

                # 배경색 설정 (tksheet 스타일 유지)
                bg_color = self.highlighted_rows.get(row_idx - 2)

                if bg_color and bg_color != "":  # 배경색이 존재하면 적용
                    cell.fill = PatternFill(
                        start_color=bg_color.replace("#", ""),
                        end_color=bg_color.replace("#", ""),
                        fill_type="solid",
                    )

        # 컬럼 너비 조정
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
                width / 7
            )  # Excel 너비 조정

        # 엑셀 파일 저장
        wb.save(file_path)
        print(f"엑셀 파일 저장 완료: {file_path}")

        # 6. 메시지 박스 띄우기
        open_file = messagebox.askyesno(
            "저장 완료", "엑셀 파일 저장이 완료되었습니다.\n파일을 여시겠습니까?"
        )

        # 7. 사용자가 "예(Yes)"를 선택한 경우 엑셀 파일 열기
        if open_file:
            try:
                os.startfile(file_path)  # Windows에서 엑셀 파일 열기
                # subprocess.Popen(["start", file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다.\n{e}")
