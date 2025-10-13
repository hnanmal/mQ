# src/views/widget/treeview_utils.py
import json
import os
from tkinter import filedialog, messagebox

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.controllers.tree_data_navigator import TreeDataManager_treeview
from src.core.fp_utils import *
import tkinter as tk
from tkinter import (
    simpledialog,
)
import re
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

from src.views.widget.treeview_editor import (
    TreeviewEditor,
    TreeviewEditor_forBuildingList,
    TreeviewEditor_forFamilyList,
    TreeviewEditor_stdGWMSWM,
)
from src.views.widget.widget import StateObserver
from src.core.fp_utils import *
from src.views.widget.treeview_contextmenu import (
    TreeViewContextMenu,
    TreeViewContextMenu_FamilyList,
    TreeViewContextMenu_GWMSWM,
)


# Composition for Style Management
class DefaultTreeViewStyleManager:

    @staticmethod
    def apply_style(treeview, rowheight=30):
        style = ttk.Style()
        style.configure(
            # "Treeview",
            "Treeview",
            font=("Arial Narrow", 10),
            highlightthickness=1,
            background="white",
            foreground="black",
            # rowheight=30,
            rowheight=rowheight,
            # height=20,
        )
        style.map(
            "Treeview",
            background=[("selected", "#fffec0")],  # Set the selection background color
            foreground=[("selected", "blue")],
        )  # Set the selection foreground color
        # Configure hover style
        style.configure(
            "Hover.Treeview", background="#d3d3d3", padding=(2, 2)
        )  # Light gray hover color
        style.configure("Treeview.Heading", font=("Arial Narrow", 8, "normal"))
        treeview.configure(style="Treeview")

    @staticmethod
    def apply_locked_style(treeview):
        style = ttk.Style()
        style.configure(
            "Custom_locked.Treeview",
            font=("Arial Narrow", 10),
            highlightthickness=1,
            background="#fafafa",
            foreground="black",
            height=50,
        )
        style.map(
            "Custom_locked.Treeview",
            background=[("selected", "#fffec0")],  # Set the selection background color
            foreground=[("selected", "blue")],
        )  # Set the selection foreground color
        # Configure hover style
        style.configure(
            "Hover.Custom_locked.Treeview", background="#d3d3d3", padding=(2, 2)
        )  # Light gray hover color
        style.configure(
            "Custom_locked.Treeview.Heading", font=("Arial Narrow", 10, "normal")
        )
        treeview.configure(style="Custom_locked.Treeview")

    @staticmethod
    def apply_alternate_row_colors(treeview):
        """Apply alternate row colors to the Treeview."""
        for i, item in enumerate(treeview.get_children("")):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            treeview.item(item, tags=(tag,))
            DefaultTreeViewStyleManager.apply_alternate_row_colors_recursive(
                treeview, item, i % 2 == 0
            )

        treeview.tag_configure("evenrow", background="white")
        treeview.tag_configure("oddrow", background="lightgray")

    @staticmethod
    def apply_alternate_row_colors_recursive(tree, item, is_even):
        """Apply alternate row colors recursively to child items."""
        children = tree.get_children(item)
        for i, child in enumerate(children):
            tag = (
                "evenrow"
                if (is_even and i % 2 == 0) or (not is_even and i % 2 != 0)
                else "oddrow"
            )
            tree.item(child, tags=(tag,))
            DefaultTreeViewStyleManager.apply_alternate_row_colors_recursive(
                tree, child, tag == "evenrow"
            )

    @staticmethod
    def apply_dynamic_alternate_row_colors(treeview):
        """현재 보이는 항목 기준으로 교차 색상을 적용"""
        visible_items = treeview.get_children("")  # 최상위 항목들 가져오기

        # 현재 보이는 모든 항목을 순회하는 함수
        def get_visible_items_recursively(items, result):
            for item in items:
                result.append(item)
                # 자식 항목 중 확장된 항목만 추가
                if treeview.item(item, "open"):  # 이 항목이 열려 있을 때만 자식 탐색
                    children = treeview.get_children(item)
                    if children:
                        get_visible_items_recursively(children, result)

        # 현재 트리뷰에서 보이는 모든 항목들을 가져옴
        all_visible_items = []
        get_visible_items_recursively(visible_items, all_visible_items)

        # 교차 색상 적용
        for i, item in enumerate(all_visible_items):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            treeview.item(item, tags=(tag,))

        # 태그에 대한 색상 정의
        treeview.tag_configure("evenrow", background="white")
        treeview.tag_configure("oddrow", background="#ebf0e6")


class ScrollbarWidget:
    def __init__(self, parent, target_widget, x=True, y=True):
        # self.frame = ttk.Frame(parent)
        self.frame = parent
        self.frame.pack(side="top", expand=True, fill="both")

        # The target widget should be provided by the user of this class.
        # self.target_widget = target_widget(self.frame)
        self.target_widget = target_widget

        commands_scroll = []
        if y and x:
            # Add Vertical Scrollbar
            self.v_scrollbar = ttk.Scrollbar(
                self.frame, orient="vertical", command=self.target_widget.yview
            )
            self.v_scrollbar.pack(side="right", fill="y")
            commands_scroll.append(self.v_scrollbar.set)

            # Add Horizontal Scrollbar
            self.h_scrollbar = ttk.Scrollbar(
                self.frame, orient="horizontal", command=self.target_widget.xview
            )
            self.h_scrollbar.pack(side="bottom", fill="x")
            # self.h_scrollbar.pack(fill="x")
            commands_scroll.append(self.h_scrollbar.set)
        elif y and not x:
            # Add Vertical Scrollbar
            self.v_scrollbar = ttk.Scrollbar(
                self.frame, orient="vertical", command=self.target_widget.yview
            )
            # self.v_scrollbar.pack(side="right", fill="y", anchor="ne")
            commands_scroll.append(self.v_scrollbar.set)
        elif not y and x:
            # Add Horizontal Scrollbar
            self.h_scrollbar = ttk.Scrollbar(
                self.frame, orient="horizontal", command=self.target_widget.xview
            )
            # self.h_scrollbar.pack(side="bottom", fill="x", anchor="sw")
            # self.h_scrollbar.pack(fill="x")
            commands_scroll.append(self.h_scrollbar.set)
        else:
            pass

        # Attach scrollbars to the target widget
        if commands_scroll:
            argNames = ["yscrollcommand", "xscrollcommand"]
            args = dict(zip(argNames, commands_scroll))
            self.target_widget.configure(
                **args
                # yscrollcommand=self.v_scrollbar.set, xscrollcommand=self.h_scrollbar.set
            )


class TreeviewSearchManager:
    def __init__(self, treeview: ttk.Treeview, container: tk.Widget):
        self.tree = treeview
        self.parent = container

        # UI
        self.search_frame = tk.Frame(self.parent)
        self.search_frame.pack(fill="x", pady=5)

        self.search_imoji = ttk.Label(self.search_frame, text="🔎 항목 검색:")
        self.search_imoji.pack(side="left", padx=(5, 5))

        self.search_entry = tk.Entry(self.search_frame)
        # self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.search_entry.pack(side="left", padx=(0, 10))
        self.search_entry.bind("<Return>", lambda event: self.search_or_next())

        self.search_button = tk.Button(
            self.search_frame, text="Search", command=self.search_or_next
        )
        self.search_button.config(
            bg="#b6d1c5",
            activebackground="#81998e",
        )
        self.search_button.pack(side="left")

        # 검색 상태
        self.matched_items = []
        self.current_index = -1
        self.previous_search = ""

        # 마커를 위한 캔버스
        self.marker_canvas = ttk.Canvas(
            self.parent,
            width=10,
            # bg="white",
            # bg="gray",
            highlightthickness=0,
            # bootstyle="info",
        )
        self.marker_canvas.config(bg="#fafafa")
        self.marker_canvas.update()
        self.marker_canvas.pack(
            side="right",
            fill="y",
            # padx=(2, 0),
        )  # 트리뷰 오른쪽 끝에 붙이기

        self.tree.bind(
            "<Configure>", lambda e: self.draw_markers()
        )  # 리사이즈 시 마커 업데이트

    def search_or_next(self):
        """처음엔 검색, 이후엔 다음 결과로 이동"""
        search_text = self.search_entry.get().strip().lower()

        if not search_text:
            # 👉 검색어 없을 때: 결과 초기화 + 마커 제거
            self.previous_search = ""
            self.matched_items = []
            self.current_index = -1
            self.marker_canvas.delete("all")  # 마커 리셋
            self.tree.selection_remove(self.tree.selection())  # 선택 제거 (선택사항)
            return

        # 검색어가 바뀌었으면 새로 검색
        if search_text != self.previous_search:
            self.previous_search = search_text
            self.matched_items = []
            self.current_index = -1
            self.find_matches(search_text)

            if self.matched_items:
                self.current_index = 0
                self.highlight_and_focus(self.matched_items[0])
        else:
            # 동일 검색어이면 다음 결과로 이동
            if not self.matched_items:
                return

            self.current_index += 1
            if self.current_index >= len(self.matched_items):
                self.current_index = 0  # 순환

            self.highlight_and_focus(self.matched_items[self.current_index])

        self.draw_markers()

    def find_matches(self, search_text):
        """트리 전체에서 검색어 포함 항목 찾기"""
        for item in self.tree.get_children():
            self._recursive_match(item, search_text)

    def _recursive_match(self, item, search_text):
        text = self.tree.item(item, "text").lower()
        values = [str(v).lower() for v in self.tree.item(item, "values")]

        if search_text in text or any(search_text in val for val in values):
            self.matched_items.append(item)

        for child in self.tree.get_children(item):
            self._recursive_match(child, search_text)

    def highlight_and_focus(self, item):
        """포커스 이동 및 강조"""
        try:
            self.tree.selection_set(item)
            self.tree.focus(item)
            self.tree.see(item)
        except:
            # 검색상태 초기화
            self.matched_items = []
            self.current_index = -1
            self.previous_search = ""

        DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(self.tree)

    def draw_markers(self):
        """검색 결과 위치 마커 표시"""
        self.marker_canvas.delete("all")
        if not self.matched_items:
            return

        all_items = self._get_all_items()
        canvas_height = int(self.marker_canvas.winfo_height())
        self.marker_canvas.update_idletasks()

        for match_id in self.matched_items:
            if match_id in all_items:
                index = all_items.index(match_id)
                y = int(canvas_height * (index / max(1, len(all_items))))
                self.marker_canvas.create_line(0, y, 10, y, fill="#d17e0a", width=2)

    def _get_all_items(self):
        items = []

        def walk(node):
            items.append(node)
            for child in self.tree.get_children(node):
                walk(child)

        for root in self.tree.get_children():
            walk(root)

        return items


# class TreeviewSearchManager:
#     def __init__(self, treeview: ttk.Treeview, container: tk.Widget):
#         self.tree = treeview
#         self.parent = container

#         # UI
#         self.search_frame = tk.Frame(self.parent)
#         self.search_frame.pack(pady=5)

#         self.search_entry = tk.Entry(self.search_frame)
#         self.search_entry.pack(side="left", padx=5)
#         self.search_entry.bind(
#             "<Return>", lambda event: self.search_or_next()
#         )  # ⏎ 바인드

#         self.search_button = tk.Button(
#             self.search_frame, text="Search", command=self.search_or_next
#         )
#         self.search_button.pack(side="left")

#         # 검색 상태
#         self.matched_items = []
#         self.current_index = -1
#         self.previous_search = ""

#     def search_or_next(self):
#         """처음엔 검색, 이후엔 다음 결과로 이동"""
#         search_text = self.search_entry.get().strip().lower()

#         if not search_text:
#             return

#         # 검색어가 바뀌었으면 새로 검색
#         if search_text != self.previous_search:
#             self.previous_search = search_text
#             self.matched_items = []
#             self.current_index = -1
#             self.find_matches(search_text)

#             if self.matched_items:
#                 self.current_index = 0
#                 self.highlight_and_focus(self.matched_items[0])
#         else:
#             # 동일 검색어이면 다음 결과로 이동
#             if not self.matched_items:
#                 return

#             self.current_index += 1
#             if self.current_index >= len(self.matched_items):
#                 self.current_index = 0  # 순환

#             self.highlight_and_focus(self.matched_items[self.current_index])

#     def find_matches(self, search_text):
#         """트리 전체에서 검색어 포함 항목 찾기"""
#         for item in self.tree.get_children():
#             self._recursive_match(item, search_text)

#     def _recursive_match(self, item, search_text):
#         text = self.tree.item(item, "text").lower()
#         values = [str(v).lower() for v in self.tree.item(item, "values")]

#         if search_text in text or any(search_text in val for val in values):
#             self.matched_items.append(item)

#         for child in self.tree.get_children(item):
#             self._recursive_match(child, search_text)

#     def highlight_and_focus(self, item):
#         """포커스 이동 및 강조"""
#         self.tree.selection_set(item)
#         self.tree.focus(item)
#         self.tree.see(item)


class BaseTreeView:
    def __init__(self, state, parent, headers):
        self.state = state
        self.tree = ttk.Treeview(
            parent,
            columns=headers,
            show=(
                "headings",
                # "tree",
            ),
        )
        # self.tree = Tableview(parent, columns=headers, show="headings")
        self.setup_columns(headers)
        # self.tree.column("#0", width=20, stretch=False)
        self.parent = parent

        # Configure tag styles
        self.tree.tag_configure(
            "hover",
            background="#d3d3d3",
            foreground="blue",
            font=("Arial Narrow", 10, "bold"),
            # padding=(2, 2),
        )  # Light gray background for hover

        self.tree.tag_configure(
            "normal", font=("Arial Narrow", 10)
        )  # Default white background

        # Track the last selected item
        self.last_selected_item = None

        self.tree.bind("<Control-MouseWheel>", lambda e: zoom_tree(e))

        def zoom_tree(event):
            step = 1 if event.delta > 0 else -1  # 확대(+) or 축소(-)
            if state._rowheight.get() < 21:
                if step > 0:
                    state._rowheight.set(state._rowheight.get() + step)
            elif state._rowheight.get() > 34:
                if step < 0:
                    state._rowheight.set(state._rowheight.get() + step)
            else:
                state._rowheight.set(state._rowheight.get() + step)
            # print(f"_rowheight: {state._rowheight.get()}")
            DefaultTreeViewStyleManager.apply_style(
                self.tree,
                state._rowheight.get(),
            )
            self.tree.update_idletasks()

            return "break"  # 🛑 기본 스크롤 방지

    # 모든 항목에서 hover 효과 제거 함수 (재귀적으로 모든 하위 항목 포함)
    def clear_all_hover(self, tree):
        def clear_tags_recursive(item):
            if self.last_selected_item == item:
                children = tree.get_children(item)
                for child in children:
                    clear_tags_recursive(child)
            else:
                # 현재 항목의 태그를 'normal'로 변경
                tree.item(item, tags=("normal",))
                # 자식 항목들에 대해서도 동일하게 적용
                children = tree.get_children(item)
                for child in children:
                    clear_tags_recursive(child)

        # 루트 항목부터 시작하여 모든 항목에 대해 재귀적으로 태그 설정
        for root_item in tree.get_children():
            clear_tags_recursive(root_item)

    # Function to apply hover effect
    def on_mouse_motion(self, event):
        # 모든 항목의 hover 상태 초기화
        self.clear_all_hover(self.tree)
        # self.tree.config(cursor="circle")  # Change to a hand cursor

        # 마우스 커서 아래에 있는 아이템의 ID 가져오기
        item_id = self.tree.identify_row(event.y)
        if item_id:
            # hover 상태를 적용
            self.tree.item(item_id, tags=("hover",))

    # Function to reset all tags when the mouse leaves the widget
    def on_mouse_leave(self, event):
        self.clear_all_hover(self.tree)
        # self.tree.config(cursor="")  # Reset to default

    def setup_columns(self, headers, hdr_widths=None):
        for idx, header in enumerate(headers):
            self.tree.heading(header, text=header)
            if hdr_widths:
                self.tree.column(header, width=hdr_widths[idx])

    def insert_data(self, data):
        for row in data:
            self.tree.insert("", "end", values=row)

    def clear_treeview(self):
        """Clear all items from the Treeview."""
        try:
            self.tree.delete(*self.tree.get_children())
        except:
            pass

    def get_tree_data(self):
        """Get the current data from the Treeview as a list of dictionaries."""
        data = []
        for item in self.tree.get_children():
            values = self.tree.item(item, "values")
            data.append(dict(zip(self.tree["columns"], values)))
        return data

    def prevent_expansion(self, event):
        # Prevent expansion or collapsing based on the lock state
        if self.locked:
            item_id = self.tree.focus()
            self.tree.item(item_id, open=False)

    def toggle_tree_lock(self, lock=True):
        """Locks or unlocks the tree from being expanded."""
        self.locked = lock
        if self.locked:
            self.tree.bind("<<TreeviewOpen>>", self.prevent_expansion)
            self.tree.bind("<<TreeviewClose>>", self.prevent_expansion)
        else:
            self.tree.unbind("<<TreeviewOpen>>")
            self.tree.unbind("<<TreeviewClose>>")

    def ensure_item_visible_at_level(self, item_id, level):
        """
        Ensure the given item or its closest ancestor is visible at the specified level.

        Parameters:
            item_id (str): The ID of the item to check.
            level (int): The target level to ensure visibility.

        Returns:
            str: The ID of the closest visible ancestor or the item itself.
        """
        tree = self.tree
        current_level = 0
        current_item = item_id

        while current_item:
            # Calculate the item's depth in the hierarchy
            parent = tree.parent(current_item)
            if not parent:
                current_level = 1
            else:
                current_level += 1

            # If the item's level is within the visible range, return it
            if current_level <= level:
                tree.see(current_item)
                return current_item

            # Otherwise, move up to the parent
            current_item = parent

        # If no visible ancestor is found, return None
        return None

    def bring_item_to_top(self, item_id):
        """
        Ensure the given item_id is shown at the top of the Treeview view.

        Parameters:
            item_id (str): The ID of the item to bring to the top.
        """
        treeview = self.tree
        if not item_id:
            return  # No item selected

        # Make sure the item is visible
        treeview.see(item_id)

        def get_visible_items(tree, parent=""):
            """Recursively collect only visible items (expanded items)."""
            visible_items = []

            # Get direct children of the current node
            for child in tree.get_children(parent):
                # Check if the item is visible (parent must be expanded)
                if tree.item(child, "open"):
                    visible_items.append(child)  # Store the visible item's text
                    # Recursively check for children if the parent is expanded
                    visible_items.extend(get_visible_items(tree, child))
                else:
                    # If the parent is collapsed, do not check its children
                    visible_items.append(child)  # Include the collapsed parent itself

            return visible_items

        visible_items = get_visible_items(treeview)

        # Find the index of the selected item
        if item_id in visible_items:
            item_index = visible_items.index(item_id)

            # Calculate the scroll fraction and move the view
            total_items = len(visible_items)
            if total_items > 0:
                fraction = (item_index - 9) / total_items
                # treeview.see("")
                treeview.yview_moveto(fraction)

    def collapse_all_items(self):
        """Recursively collapse all items in the Treeview."""

        def collapse_item(item):
            # Set the item to closed (collapsed)
            self.tree.item(item, open=False)
            # Get the children of the current item
            children = self.tree.get_children(item)
            for child in children:
                collapse_item(child)

        # Get all root items
        root_items = self.tree.get_children("")
        for item in root_items:
            collapse_item(item)

    def expand_all_items(self):
        """Recursively expand all items in the Treeview."""
        treeview = self.tree
        self.collapse_all_items()

        def expand_item(item):
            self.state.log_widget.write(f"Expanding item: {item}")
            treeview.item(item, open=True)
            # Get the children of the current item
            children = treeview.get_children(item)
            for child in children:
                expand_item(child)

        # Get all root items
        root_items = treeview.get_children("")
        if not root_items:
            self.state.log_widget.write("No root items found.")
        for item in root_items:
            expand_item(item)

    def expand_tree_to_level(self, level):
        """Expand the treeview up to a given level."""
        # First, collapse all items to ensure fresh expansion
        self.collapse_all_items()

        def expand_item(item, current_level):
            # If the current level exceeds the specified level, stop expanding
            if current_level > level:
                return
            # Try to expand the current item
            try:
                self.tree.item(item, open=True)
                children = self.tree.get_children(item)
                for child in children:
                    expand_item(child, current_level + 1)
            except Exception as e:
                self.state.log_widget.write(
                    f"Error expanding tree at level {current_level}: {e}"
                )

        # Expand all root items initially
        root_items = self.tree.get_children("")
        for root in root_items:
            try:
                expand_item(root, 1)
            except:
                pass

    def remove_items_with_rule(self, _data_list, _depth, _rule):
        """
        Removes items with an underscore in their 'name' at depth 2, along with their children.

        Parameters:
            data_list (list): A list of hierarchical data structures.

        Returns:
            list: Cleaned data with specified items removed.
        """
        data_list = deepcopy(_data_list)

        def clean_children(children, depth):
            cleaned_children = []
            for child in children:
                # At _depth , check for _rule in the 'name' field
                if depth == _depth and _rule in child.get("name", ""):
                    continue  # Skip this item and its children
                # Recursively clean if 'children' field exists
                if "children" in child:
                    child["children"] = clean_children(child["children"], depth + 1)
                cleaned_children.append(child)
            return cleaned_children

        # Process each top-level item in the list
        for item in data_list:
            if "children" in item:
                # item["children"] = clean_children(item["children"], _depth)
                item["children"] = clean_children(item["children"], 1)

        return data_list

    def insert_data_with_levels(self, data, parent_id=""):
        """Insert data into the TreeView with levels based on the new data structure."""
        for node in data:
            if isinstance(node, dict):
                # Extract the last value in the 'values' list to be displayed in the TreeView
                display_value = node["values"]

                # Insert the current node with its name and the last value from 'values'
                node_id = self.tree.insert(
                    parent_id, "end", text=node["name"], values=(display_value)
                )
                self.tree.item(
                    node_id,  # open=True
                )  # Keep the tree open to display all items

                # Recursively insert children if there are any
                if "children" in node and isinstance(node["children"], list):
                    # self.insert_data_with_levels(node["children"], node_id)
                    self.insert_data_with_levels(
                        sorted(node["children"], key=sort_func), node_id
                    )

            elif isinstance(node, str):
                # Insert the string as a leaf node without additional children
                # self.tree.insert(parent_id, "end", text=node, values=(node,))
                pass

    def insert_data_with_levels_withTag(self, data, parent_id=""):
        """Insert data into the TreeView with levels based on the new data structure."""
        for node in data:
            if isinstance(node, dict):
                # Extract the last value in the 'values' list to be displayed in the TreeView
                display_value = node["values"]

                # Insert the current node with its name and the last value from 'values'
                if node["children"]:
                    node_id = self.tree.insert(
                        parent_id, "end", text=node["name"], values=(display_value)
                    )
                else:
                    node_id = self.tree.insert(
                        parent_id,
                        "end",
                        text=node["name"],
                        values=(display_value),
                        tags="red_text",
                    )

                # Recursively insert children if there are any
                if "children" in node and isinstance(node["children"], list):
                    # self.insert_data_with_levels(node["children"], node_id)
                    self.insert_data_with_levels_withTag(
                        sorted(node["children"], key=sort_func), node_id
                    )

            elif isinstance(node, str):
                # Insert the string as a leaf node without additional children
                # self.tree.insert(parent_id, "end", text=node, values=(node,))
                pass

    def get_item_indices(self, selected_item_id):
        indices = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.tree.parent(current_item)
            siblings = self.tree.get_children(parent_id)

            try:
                # Get the index of the current item among its siblings
                item_index = siblings.index(current_item)
            except ValueError:
                # If the item is not found among siblings, there's an issue
                self.state.log_widget.write(
                    f"Error: Item '{current_item}' not found among siblings {siblings}"
                )
                return []

            # Debug: print current processing details
            self.state.log_widget.write(
                f"Current item: {current_item}, Parent ID: {parent_id}, Index: {item_index}"
            )

            # Insert the item index at the beginning of the list to maintain top-down order
            indices.insert(0, item_index)

            # Move to the parent item
            current_item = parent_id

        # Debug: print the final indices
        # self.state.log_widget.write(f"Final Indices: {indices}")
        return indices

    def select_item_by_indices(self, indices):
        """
        Select an item in the Treeview using the provided indices.
        """
        current_parent = ""
        target_item = None

        for index in indices:
            children = self.tree.get_children(current_parent)
            if index < len(children):
                target_item = children[index]
                current_parent = target_item
            else:
                self.state.log_widget.write(
                    f"Index {index} is out of bounds for the current parent."
                )
                return

        if target_item:
            self.tree.selection_set(target_item)
            self.tree.focus(target_item)
            self.tree.see(target_item)


class TeamStd_GWMTreeView:
    def __init__(self, state, parent, showmode="team", view_level=2):
        self.state = state
        self.data_kind = "std-GWM"
        self.showmode = showmode
        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(
            state, lambda e: self.update(e, view_level), tag="GWM"
        )

        self.selected_item = tk.StringVar()
        self.selected_item.trace_add("write", state._notify_selected_change)
        headers = ["분류", "G-WM", "Item"]
        hdr_widths = [100, 140, 180]

        def set_tree_row(event=None):
            # rowheight = round(rowheight_scalebar.get())
            # state._rowheight = tk.IntVar()
            state._rowheight.set(
                round(rowheight_scalebar.get()),
            )

            DefaultTreeViewStyleManager.apply_style(
                self.treeview.tree,
                state._rowheight.get(),
            )
            self.treeview.tree.update_idletasks()

        rowheight_scalebar = ttk.Scale(
            parent,
            bootstyle="info",
            variable=state._rowheight,
            command=set_tree_row,
            from_=20,
            to=35,
        )
        if showmode == "tmp_team_fl" or showmode == "project_fl":
            pass
        else:
            rowheight_scalebar.pack(padx=10, pady=5, side="top")

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(parent, width=600, height=2000)
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=3000)

        # config selection mode
        self.treeview.tree.config(selectmode="browse")
        # Tag styles
        # self.treeview.tree.tag_configure("bold", font=("Arial", 10, "bold"))
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(tree_frame)

        # 검색 매니저 붙이기
        self.search = TreeviewSearchManager(self.treeview.tree, tree_frame)

        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        # self.treeviewEditor = TreeviewEditor(state, self)
        self.treeviewEditor = TreeviewEditor_stdGWMSWM(state, self)

        # Track the last selected item with an instance attribute
        self.last_selected_item = None

        self.treeview.tree.bind("<<TreeviewSelect>>", self.on_item_selected)

        # Create and integrate context menu
        self.context_menu = TreeViewContextMenu_GWMSWM(
            state,
            self.treeview,
            data_kind=self.data_kind,
            add_top=self.add_top_item,
            add=self.add_item,
            delete=self.delete_item,
            copy_Topitem=self.copy_Topitem,
            copy_GWM=self.copy_GWM,
            copy_item=self.copy_item,
            edit_item=self.treeviewEditor.append_suffix_via_popup,
        )

        self.treeview.tree.tag_configure(
            "bold", foreground="#a000cc", font=("맑은 고딕", 10, "bold")
        )

    def set_title(self, parent):
        title_font = ttk.font.Font(family="맑은 고딕", size=12)
        title_label = ttk.Label(parent, text="Standard Types for GWM", font=title_font)
        title_label.pack(padx=5, pady=5, anchor="w")

    def should_highlight(self, item_text):
        state = self.state
        # print(f"used_names !!")
        used_names = go(
            state.team_std_info["project-assigntype"]["children"],
            map(lambda x: x["children"]),
            lambda x: list(chain(*x)),
            filter(lambda x: list(x)[0] == "GWM"),
            map(lambda x: x[2]),
            list,
        )
        # print(f"used_names !!{used_names}[0]")
        used_WM_names = go(
            used_names,
            map(lambda x: x.split(" | ")),
            list,
            map(lambda x: x[0]),
            set,
            list,
        )
        # print(f"used_WM_names !!{used_WM_names}")

        used_item_names = go(
            used_names,
            map(lambda x: x.split(" | ")),
            list,
            map(lambda x: x[1]),
            set,
            list,
        )
        # print(f"used_item_names !!{used_item_names}")

        return item_text in used_WM_names  # or (item_text in used_item_names)

    def apply_highlight_to_items(self):
        for item_id in self.treeview.tree.get_children(""):
            self._apply_highlight_recursive(item_id)

    def _apply_highlight_recursive(self, item_id):
        item_text = self.treeview.tree.item(item_id, "text")
        if self.should_highlight(item_text):
            self.treeview.tree.item(item_id, tags=("bold",))
            # print("하이라이트!if")
        else:
            self.treeview.tree.item(item_id, tags=("normal",))
            # print("하이라이트!else")

        for child_id in self.treeview.tree.get_children(item_id):
            self._apply_highlight_recursive(child_id)

    def update(self, event=None, view_level=None):
        state = self.state
        # self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")

        selected_item_id = self.treeview.tree.focus()
        origin_indices = None

        # Extract origin indices if a valid item is focused
        if selected_item_id:
            try:
                origin_indices = self.treeview.get_item_indices(selected_item_id)
                print(origin_indices, "!!!")
            except Exception as e:
                print(f"origin_indices 추출 실패: {e}")

        """Update the TreeView whenever the state changes."""
        if self.data_kind in state.team_std_info:
            if self.showmode == "team":
                data = self.treeview.remove_items_with_rule(
                    state.team_std_info[self.data_kind]["children"],
                    _depth=1,
                    _rule="::",
                )
            else:
                data = state.team_std_info[self.data_kind]["children"]

            # Clear the TreeView and reload data from the updated state
            self.treeview.clear_treeview()
            self.treeview.insert_data_with_levels(data)

            self.treeview.expand_tree_to_level(level=view_level)

        # Reselect the item if possible
        if origin_indices:
            try:
                self.treeview.select_item_by_indices(origin_indices)
            except Exception as e:
                self.state.log_widget.write(f"Item selection failed: {e}")

        self.apply_highlight_to_items()
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        state = self.state
        treeDataManager = TreeDataManager_treeview(self.state)
        try:
            # Reset the tag for the previously selected item to 'normal'
            if self.last_selected_item:
                self.treeview.tree.item(self.last_selected_item, tags=("normal",))
        except Exception as e:
            self.state.log_widget.write(f"Error resetting last selected item tag: {e}")

        # Get the currently selected item
        selected_item_id = self.treeview.tree.focus()
        state.selected_GWMSWM = self.treeview.tree.item(selected_item_id, "values")
        self.treeview.last_selected_item = selected_item_id

        try:
            state_data = self.state.team_std_info[self.data_kind]["children"]
            target_node = treeDataManager.find_node_by_name_recur(
                state_data, self.treeview.tree.item(selected_item_id, "text")
            ).copy()
            self.state.log_widget.write(
                f"select tree item!!! {self.treeview.tree.item(selected_item_id, 'text')}"
            )

            self.state.selectedGWMitems = [target_node]
        except:
            pass

        # Get the parent and grandparent of the selected item
        parent_item_id = self.treeview.tree.parent(selected_item_id)
        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)

        # Extract the names from the values, ensuring each level is handled appropriately
        try:
            # Get the name of the selected item, parent, and grandparent
            selected_item_name = go(
                self.treeview.tree.item(selected_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            parent_item_name = go(
                self.treeview.tree.item(parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            grand_parent_item_name = go(
                self.treeview.tree.item(grand_parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )

            # Format the selected path as "grandparent | parent | selected"
            formatted_value = " | ".join(
                filter(
                    None, [grand_parent_item_name, parent_item_name, selected_item_name]
                )
            )

            # Update the selected item in the state
            self.selected_item.set(formatted_value)

            # Register the last selected item
            self.last_selected_item = selected_item_id

        except Exception as e:
            self.state.log_widget.write(f"Error processing selected item details: {e}")

        self.apply_highlight_to_items()

    def get_parent_ids(self, selected_item_id):
        parent_ids = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.treeview.tree.parent(current_item)
            if parent_id:  # 부모 항목이 있을 경우에만 리스트에 추가
                parent_ids.append(parent_id)
            current_item = parent_id

        # Debug: print the final list of parent IDs
        self.state.log_widget.write(
            f"Parent IDs for selected item '{selected_item_id}': {parent_ids}"
        )
        return list(reversed(parent_ids))

    def copy_Topitem(self):
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )

        print(f"path: {path}")

        self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=0,
            # name_depth=[1, 2],
        )
        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

    def copy_GWM(self):
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )
        print(f"path: {path}")

        copied_GWMSWM = self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=1,
            # name_depth=[1, 2],
        )

        for item_ in copied_GWMSWM["children"]:
            origin_path = deepcopy(path)
            origin_path.append(item_["name"])
            copied_path = deepcopy(origin_path)
            copied_path[1] = new_name
            print(f"[path] {origin_path}")
            print(f"[origin_path] {origin_path}")
            print(f"[copied_path] {copied_path}")

            self.apply_originWMassign_forItem(origin_path, copied_path)

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

    def update_deleting_stdType_GWMSWM_in(self):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        target_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = target_name

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.delete_node(
                "std-familylist",
                path,
            )

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])
        return target_parent_nodes

    def update_deleting_stdType_wmItem_in(self):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        target_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = self.treeview.tree.item(parent_id, "text")

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name, target_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.delete_node(
                "std-familylist",
                path,
            )

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])
        return target_parent_nodes

    def update_editing_stdType_GWMSWM_in(self, new_name):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        old_name = self.treeview.tree.item(selected_id, "text")

        # project-GWMSWM 업데이트
        pjt_GWMSWM_db = state.team_std_info[f"project-{mode}"]
        parent_name = self.treeview.tree.item(parent_id, "text")
        try:
            children_names = go(
                self.treeview.tree.get_children(selected_id),
                list,
                map(lambda x: self.treeview.tree.item(x, "text")),
                list,
            )
            exist_pathes = go(
                children_names,
                map(lambda x: " | ".join([parent_name, old_name, x])),
                list,
            )
            new_pathes = go(
                children_names,
                map(lambda x: " | ".join([parent_name, new_name, x])),
                list,
            )

            print(f"exist_pathes::{exist_pathes}")
            print(f"new_pathes::{new_pathes}")

            for exist_path, new_path in zip(exist_pathes, new_pathes):
                pjt_GWMSWM_db[new_path] = pjt_GWMSWM_db[exist_path]
        except:
            pass

        # std-familylist 업데이트
        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            old_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = old_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.update_node_value(
                "std-familylist",
                path,
                4,
                new_name,
            )
            self.treeDataManager.update_node_name(
                "std-familylist",
                path,
                new_name,
            )

        # project-assigntype 업데이트 필요
        pjt_assign_db = state.team_std_info["project-assigntype"]["children"]

        for idx, asgn_item in enumerate(pjt_assign_db):
            try:
                pjt_assign_db[idx] = self.update_GWMSWM_editing_to_pjtAssign(
                    exist_pathes, new_pathes, asgn_item
                )
                # print("asgn update succes")
            except:
                pass
                # print("asgn update failed")

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])
        return target_parent_nodes

    def update_GWMSWM_editing_to_pjtAssign(self, exist_pathes, new_pathes, asgn_item):
        mode = self.data_kind.split("-")[-1]
        state = self.state

        exist_brief_pathes = go(
            exist_pathes,
            map(lambda x: x.split(" | ")),
            map(lambda x: x[1:]),
            map(lambda x: " | ".join(x)),
            list,
        )
        # print(f"exist_pathes::{exist_pathes}")
        # print(f"exist_brief_pathes::{exist_brief_pathes}")
        new_brief_pathes = go(
            new_pathes,
            map(lambda x: x.split(" | ")),
            map(lambda x: x[1:]),
            map(lambda x: " | ".join(x)),
            list,
        )
        # print(f"new_pathes::{new_pathes}")
        # print(f"new_brief_pathes::{new_brief_pathes}")
        asgn_children = asgn_item["children"]

        for idx, item in enumerate(asgn_children):
            for old_brief_name, new_brief_name in zip(
                exist_brief_pathes, new_brief_pathes
            ):
                # print(f"item::{item}")
                # print(f"mode:{mode}")
                # print(f"old_brief_name:{old_brief_name}")
                if mode == item[0] and old_brief_name == item[2]:
                    print("찾았다!")
                    try:
                        print("asgn 변경시도")
                        item[2] = new_brief_name
                        # asgn_children[idx] = item
                        print(f"asgn 변경성공 {old_brief_name} to {new_brief_name}")
                    except:
                        print(f"asgn 변경실패 {old_brief_name} to {new_brief_name}")
        return asgn_item

    def update_editing_stdType_wmItem_in(self, new_name):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        old_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = self.treeview.tree.item(parent_id, "text")

        ## pjt-GWM의 키값에 바뀐 이름 업데이트 하기
        old_path_str = self.selected_item.get()
        old_path_forPjtGWMSWM = old_path_str.split(" | ")
        new_path_forPjtGWMSWM = deepcopy(old_path_forPjtGWMSWM)
        new_path_forPjtGWMSWM[-1] = new_name
        new_path_str = " | ".join(new_path_forPjtGWMSWM)
        print(f"[update editing] old_path_forPjtGWMSWM - {old_path_forPjtGWMSWM}")
        print(f"[update editing] new_path_forPjtGWMSWM - {new_path_forPjtGWMSWM}")

        # if state.team_std_info[f"project-{mode}"].get(new_path_str):
        if state.team_std_info[f"project-{mode}"].get(old_path_str):
            state.team_std_info[f"project-{mode}"][new_path_str] = state.team_std_info[
                f"project-{mode}"
            ].pop(old_path_str)

        # 패밀리리스트 업데이트 구간
        print(f"check GWM_parent_name::{GWM_parent_name}")

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name, old_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.update_node_value(
                "std-familylist",
                path,
                5,
                new_name,
            )
            self.treeDataManager.update_node_name(
                "std-familylist",
                path,
                new_name,
            )

        # project-assigntype 업데이트 필요
        pjt_assign_db = state.team_std_info["project-assigntype"]["children"]

        for idx, asgn_item in enumerate(pjt_assign_db):
            try:
                pjt_assign_db[idx] = self.update_GWMSWM_editing_to_pjtAssign(
                    [old_path_str], [new_path_str], asgn_item
                )
                # print("asgn update succes")
            except:
                pass
                # print("asgn update failed")

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])
        return target_parent_nodes

    def update_copying_stdType_wmItem_in(self, copied_node, parent_item_name):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            parent_item_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = parent_item_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: x + [target_GWM_name]),
            list,
        )

        for path in target_paths:
            self.treeDataManager.match_GWMitems_to_stdFam(
                "std-familylist",
                *path,
                [copied_node],
            )

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

        return target_parent_nodes

    # 복사 item 항목 마다 원본 GWM 할당 상황 조회 및 반영
    def apply_originWMassign_forItem(self, path, copied_path):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        ref_WM_assign_key = " | ".join(path)
        ref_WM_assign_val = state.team_std_info[f"project-{mode}"].get(
            ref_WM_assign_key
        )
        copied_WM_assign_key = " | ".join(copied_path)

        if ref_WM_assign_val:
            state.team_std_info[f"project-{mode}"].update(
                {copied_WM_assign_key: ref_WM_assign_val}
            )

    def copy_item(self):
        mode = self.data_kind.split("-")[-1]
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )
        print(f"path: {path}")

        copied_node = self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=2,
            # name_depth=[1, 2],
        )

        try:
            self.update_copying_stdType_wmItem_in(copied_node, parent_item_name)
        except:
            pass

        copied_path = deepcopy(path)
        copied_path[-1] = new_name

        self.apply_originWMassign_forItem(path, copied_path)

        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

    def add_top_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Top Item", "Enter the name of the new item:"
        )
        if new_item_name:
            self.treeDataManager.add_top_level_node(self.data_kind, new_item_name)
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

    def add_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])

    def delete_item(self):
        state = self.state

        # Get the selected item ID
        selected_item_id = self.treeview.tree.selection()
        if not selected_item_id:
            print("No item selected.")
            return
        selected_item_id = selected_item_id[0]  # Handle single selection case

        # Build the full path to the selected item
        def get_full_path(item_id):
            path = []
            current_item_id = item_id
            while current_item_id:
                # Get the name of the current item
                item_values = self.treeview.tree.item(current_item_id, "values")
                item_name = next(
                    (v for v in item_values if v), None
                )  # Get the first non-empty value
                if item_name:
                    path.insert(0, item_name)  # Add to the beginning of the path
                # Move to the parent item
                current_item_id = self.treeview.tree.parent(current_item_id)
            return path

        # Get the full path of the selected item
        full_path = get_full_path(selected_item_id)
        print(f"Deleting item with path: {full_path}")

        # Validate the constructed path
        if not full_path:
            print("Could not construct a valid path for the selected item.")
            return

        # self.current_item = selected_item_id
        self.current_column = go(
            self.treeview.tree.item(selected_item_id, "values"),
            lambda x: next((i for i, s in enumerate(x) if s), None),
        )

        # Pass the full path to the delete_node method
        try:
            self.treeDataManager.delete_node(self.data_kind, full_path)
            print(f"Successfully deleted item with path: {full_path}")
        except Exception as e:
            print(f"Error deleting item with path {full_path}: {e}")

        if self.current_column == 2:
            self.update_deleting_stdType_wmItem_in()
        elif self.current_column == 1:
            self.update_deleting_stdType_GWMSWM_in()

        # Notify observers about the state update
        state.observer_manager.notify_observers(state, targets=["GWM", "famlist"])


class PjtStd_GWMTreeView(TeamStd_GWMTreeView):
    def __init__(self, state, parent):
        super().__init__(state, parent, view_level=2)


class TeamStd_WMmatching_TreeView:
    def __init__(self, state, parent, relate_widget, data_kind=None, view_level=2):
        self.state = state
        self.data_kind = data_kind
        self.selected_item_relate_widget = relate_widget.selected_item
        headers = ["Work Master"]
        hdr_widths = [2000]

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(parent, width=3000, height=1000)
        tree_frame.pack(expand=True, fill="both")
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=10)
        self.state_observer = StateObserver(state, lambda e: self.update(e, view_level))

        # config selection mode
        self.treeview.tree.config(selectmode="extended")

        # Tag styles
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(tree_frame)
        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        self.treeviewEditor = TreeviewEditor(state, self)

        # Track the last selected item with an instance attribute
        self.last_selected_item = None
        # Bind selection events
        self.treeview.tree.bind("<<TreeviewSelect>>", self.on_item_selected)

    def set_title(self, parent):
        title_font = ttk.font.Font(family="맑은 고딕", size=12)
        title_label = ttk.Label(
            parent, text="Matched WMs for Selected Standard Types", font=title_font
        )
        title_label.pack(padx=5, pady=5, anchor="w")

    def update(self, event=None, view_level=None):
        state = self.state
        """Update the TreeView whenever the state changes."""
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")
        self.state.log_widget.write(
            f"선택아이템 출력 : {self.selected_item_relate_widget.get()}"
        )

        try:
            # Split the selected item path to find the grandparent, parent, and selected item names
            grand_parent_item_name, parent_item_name, selected_item_name = (
                self.selected_item_relate_widget.get().split(" | ")
            )

            # Ensure the data kind exists in the team standard information
            if self.data_kind in state.team_std_info:
                data = state.team_std_info[self.data_kind]["children"]

                # Find the grandparent node
                grand_parent_node = next(
                    (node for node in data if node["name"] == grand_parent_item_name),
                    None,
                )
                if grand_parent_node:
                    # Find the parent node
                    parent_node = next(
                        (
                            node
                            for node in grand_parent_node["children"]
                            if node["name"] == parent_item_name
                        ),
                        None,
                    )
                    if parent_node:
                        # Find the selected node
                        selected_node = next(
                            (
                                node
                                for node in parent_node["children"]
                                if node["name"] == selected_item_name
                            ),
                            None,
                        )
                        if selected_node:
                            # Clear the TreeView and insert the data for the selected node
                            self.treeview.clear_treeview()

                            # Wrap the children of the selected node for insertion
                            wrapped_data = go(
                                selected_node["children"],
                                map(lambda x: [x]),
                                list,
                            )
                            # self.treeview.insert_data_with_levels(wrapped_data)
                            self.treeview.insert_data(wrapped_data)
                        else:
                            self.state.log_widget.write(
                                f"Selected item '{selected_item_name}' not found."
                            )
                    else:
                        self.state.log_widget.write(
                            f"Parent item '{parent_item_name}' not found."
                        )
                else:
                    self.state.log_widget.write(
                        f"Grandparent item '{grand_parent_item_name}' not found."
                    )

        except Exception as e:
            self.state.log_widget.write(
                f"{self.__class__.__name__} > update 메소드 진입 안됩니다~: {e}"
            )

        self.treeview.expand_tree_to_level(level=view_level)

        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        # if self.last_selected_item:
        #     self.treeview.tree.item(self.treeview.last_selected_item, tags=("normal",))

        self.state.log_widget.write("on_item_selected_시작")
        selected_items_id = self.treeview.tree.selection()
        selected_items_name = go(
            selected_items_id,
            map(lambda x: list(self.treeview.tree.item(x, "values"))),
            list,
            map(lambda x: x[0]),
            list,
        )
        # print(selected_items_name)
        self.state.selected_matchedWMs = selected_items_name
        # # 마지막 선택항목으로 재등록
        # self.last_selected_item = selected_items_id[0]
        self.state.log_widget.write("on_item_selected_종료")


class TeamStd_SWMTreeView:
    def __init__(self, state, parent, showmode="team", view_level=2):
        self.state = state
        self.data_kind = "std-SWM"
        self.showmode = showmode
        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(
            state, lambda e: self.update(e, view_level), tag="SWM"
        )

        self.selected_item = tk.StringVar()
        self.selected_item.trace_add("write", state._notify_selected_change)
        headers = ["분류", "S-WM", "Item"]
        # headers = ["분류-1", "S-WM", "Item"]
        hdr_widths = [100, 90, 250]

        def set_tree_row(event=None):
            # rowheight = round(rowheight_scalebar.get())
            # state._rowheight = tk.IntVar()
            state._rowheight.set(
                round(rowheight_scalebar.get()),
            )

            DefaultTreeViewStyleManager.apply_style(
                self.treeview.tree,
                state._rowheight.get(),
            )
            self.treeview.tree.update_idletasks()

        rowheight_scalebar = ttk.Scale(
            parent,
            bootstyle="info",
            variable=state._rowheight,
            command=set_tree_row,
            from_=20,
            to=35,
        )
        if showmode == "tmp_team_fl" or showmode == "project_fl":
            pass
        else:
            rowheight_scalebar.pack(padx=10, pady=5, side="top")

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(parent, width=600, height=2000)
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=3000)

        # config selection mode
        self.treeview.tree.config(selectmode="browse")
        # Tag styles
        # self.treeview.tree.tag_configure("bold", font=("Arial", 10, "bold"))
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(tree_frame)

        # 검색 매니저 붙이기
        self.search = TreeviewSearchManager(self.treeview.tree, tree_frame)

        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        self.treeviewEditor = TreeviewEditor_stdGWMSWM(state, self)

        # Track the last selected item with an instance attribute
        self.last_selected_item = None
        # Bind selection events
        self.treeview.tree.bind("<<TreeviewSelect>>", self.on_item_selected)

        # Create and integrate context menu
        self.context_menu = TreeViewContextMenu_GWMSWM(
            state,
            self.treeview,
            self.data_kind,
            add_top=self.add_top_item,
            add=self.add_item,
            delete=self.delete_item,
            copy_Topitem=self.copy_Topitem,
            copy_SWM=self.copy_SWM,
            copy_item=self.copy_item,
            edit_item=self.treeviewEditor.append_suffix_via_popup,
        )

        self.treeview.tree.tag_configure(
            "bold", foreground="#a000cc", font=("맑은 고딕", 10, "bold")
        )

    def set_title(self, parent):
        title_font = ttk.font.Font(family="맑은 고딕", size=12)
        title_label = ttk.Label(parent, text="Standard Types for SWM", font=title_font)
        title_label.pack(padx=5, pady=5, anchor="w")

    def should_highlight(self, item_text):
        state = self.state
        # print(f"used_names !!")
        used_names = go(
            state.team_std_info["project-assigntype"]["children"],
            map(lambda x: x["children"]),
            lambda x: list(chain(*x)),
            filter(lambda x: x[0] == "SWM"),
            map(lambda x: x[2]),
            list,
        )
        # print(f"used_names !!{used_names}[0]")
        used_WM_names = go(
            used_names,
            map(lambda x: x.split(" | ")),
            list,
            map(lambda x: x[0]),
            set,
            list,
        )
        # print(f"used_WM_names !!{used_WM_names}")

        used_item_names = go(
            used_names,
            map(lambda x: x.split(" | ")),
            list,
            map(lambda x: x[1]),
            set,
            list,
        )
        # print(f"used_item_names !!{used_item_names}")

        return item_text in used_item_names

    def apply_highlight_to_items(self):
        for item_id in self.treeview.tree.get_children(""):
            self._apply_highlight_recursive(item_id)

    def _apply_highlight_recursive(self, item_id):
        item_text = self.treeview.tree.item(item_id, "text")
        if self.should_highlight(item_text):
            self.treeview.tree.item(item_id, tags=("bold",))
            # print("하이라이트!if")
        else:
            self.treeview.tree.item(item_id, tags=("normal",))
            # print("하이라이트!else")

        for child_id in self.treeview.tree.get_children(item_id):
            self._apply_highlight_recursive(child_id)

    def update(self, event=None, view_level=None):
        state = self.state
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")

        selected_item_id = self.treeview.tree.focus()
        origin_indices = None

        # Extract origin indices if a valid item is focused
        if selected_item_id:
            try:
                origin_indices = self.treeview.get_item_indices(selected_item_id)
                print(origin_indices, "!!!")
            except Exception as e:
                print(f"origin_indices 추출 실패: {e}")

        """Update the TreeView whenever the state changes."""
        if self.data_kind in state.team_std_info:
            if self.showmode == "team":
                data = self.treeview.remove_items_with_rule(
                    state.team_std_info[self.data_kind]["children"],
                    _depth=1,
                    _rule="::",
                )
            else:
                data = state.team_std_info[self.data_kind]["children"]

            # Clear the TreeView and reload data from the updated state
            self.treeview.clear_treeview()
            self.treeview.insert_data_with_levels(data)

            self.treeview.expand_tree_to_level(level=view_level)

        # Reselect the item if possible
        if origin_indices:
            try:
                self.treeview.select_item_by_indices(origin_indices)
            except Exception as e:
                print(f"Item selection failed: {e}")

        self.apply_highlight_to_items()
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        state = self.state
        treeDataManager = TreeDataManager_treeview(self.state)
        try:
            # Reset the tag for the previously selected item to 'normal'
            if self.last_selected_item:
                self.treeview.tree.item(self.last_selected_item, tags=("normal",))

        except Exception as e:
            self.state.log_widget.write(f"Error resetting last selected item tag: {e}")

        # Get the currently selected item
        selected_item_id = self.treeview.tree.focus()
        state.selected_GWMSWM = self.treeview.tree.item(selected_item_id, "values")
        self.treeview.last_selected_item = selected_item_id

        try:
            state_data = self.state.team_std_info[self.data_kind]["children"]
            target_node = treeDataManager.find_node_by_name_recur(
                state_data, self.treeview.tree.item(selected_item_id, "text")
            ).copy()
            self.state.log_widget.write(
                f"select tree item!!! {self.treeview.tree.item(selected_item_id, 'text')}"
            )

            self.state.selectedGWMitems = [target_node]
        except:
            pass

        # Get the parent and grandparent of the selected item
        parent_item_id = self.treeview.tree.parent(selected_item_id)
        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)

        # Extract the names from the values, ensuring each level is handled appropriately
        try:
            # Get the name of the selected item, parent, and grandparent
            selected_item_name = go(
                self.treeview.tree.item(selected_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            parent_item_name = go(
                self.treeview.tree.item(parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            grand_parent_item_name = go(
                self.treeview.tree.item(grand_parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )

            # Format the selected path as "grandparent | parent | selected"
            formatted_value = " | ".join(
                filter(
                    None, [grand_parent_item_name, parent_item_name, selected_item_name]
                )
            )

            # Update the selected item in the state
            self.selected_item.set(formatted_value)

            # Register the last selected item
            self.last_selected_item = selected_item_id

        except Exception as e:
            self.state.log_widget.write(f"Error processing selected item details: {e}")

        self.apply_highlight_to_items()

    def get_parent_ids(self, selected_item_id):
        parent_ids = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.treeview.tree.parent(current_item)
            if parent_id:  # 부모 항목이 있을 경우에만 리스트에 추가
                parent_ids.append(parent_id)
            current_item = parent_id

        # Debug: print the final list of parent IDs
        self.state.log_widget.write(
            f"Parent IDs for selected item '{selected_item_id}': {parent_ids}"
        )
        return list(reversed(parent_ids))

    def copy_Topitem(self):
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )

        print(f"path: {path}")

        self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=0,
            # name_depth=[1, 2],
        )
        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])

    def copy_SWM(self):
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )
        print(f"path: {path}")

        copied_GWMSWM = self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=1,
            # name_depth=[1, 2],
        )

        for item_ in copied_GWMSWM["children"]:
            origin_path = deepcopy(path)
            origin_path.append(item_["name"])
            copied_path = deepcopy(origin_path)
            copied_path[1] = new_name
            print(f"[path] {origin_path}")
            print(f"[origin_path] {origin_path}")
            print(f"[copied_path] {copied_path}")

            self.apply_originWMassign_forItem(origin_path, copied_path)

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])

    def update_deleting_stdType_GWMSWM_in(self):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        target_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = target_name

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.delete_node(
                "std-familylist",
                path,
            )

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])
        return target_parent_nodes

    def update_deleting_stdType_wmItem_in(self):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        target_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = self.treeview.tree.item(parent_id, "text")

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name, target_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.delete_node(
                "std-familylist",
                path,
            )

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])
        return target_parent_nodes

    def update_editing_stdType_GWMSWM_in(self, new_name):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        old_name = self.treeview.tree.item(selected_id, "text")

        # project-GWMSWM 업데이트
        pjt_GWMSWM_db = state.team_std_info[f"project-{mode}"]
        parent_name = self.treeview.tree.item(parent_id, "text")
        children_names = go(
            self.treeview.tree.get_children(selected_id),
            list,
            map(lambda x: self.treeview.tree.item(x, "text")),
            list,
        )
        exist_pathes = go(
            children_names,
            map(lambda x: " | ".join([parent_name, old_name, x])),
            list,
        )
        new_pathes = go(
            children_names,
            map(lambda x: " | ".join([parent_name, new_name, x])),
            list,
        )

        print(f"exist_pathes::{exist_pathes}")

        for exist_path, new_path in zip(exist_pathes, new_pathes):
            pjt_GWMSWM_db[new_path] = pjt_GWMSWM_db[exist_path]

        # std-familylist 업데이트
        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            old_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = old_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.update_node_value(
                "std-familylist",
                path,
                4,
                new_name,
            )
            self.treeDataManager.update_node_name(
                "std-familylist",
                path,
                new_name,
            )

        # project-assigntype 업데이트 필요
        pjt_assign_db = state.team_std_info["project-assigntype"]["children"]

        for idx, asgn_item in enumerate(pjt_assign_db):
            try:
                pjt_assign_db[idx] = self.update_GWMSWM_editing_to_pjtAssign(
                    exist_pathes, new_pathes, asgn_item
                )
                # print("asgn update succes")
            except:
                pass
                # print("asgn update failed")

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])
        return target_parent_nodes

    def update_GWMSWM_editing_to_pjtAssign(self, exist_pathes, new_pathes, asgn_item):
        mode = self.data_kind.split("-")[-1]
        state = self.state

        exist_brief_pathes = go(
            exist_pathes,
            map(lambda x: x.split(" | ")),
            map(lambda x: x[1:]),
            map(lambda x: " | ".join(x)),
            list,
        )
        # print(f"exist_pathes::{exist_pathes}")
        # print(f"exist_brief_pathes::{exist_brief_pathes}")
        new_brief_pathes = go(
            new_pathes,
            map(lambda x: x.split(" | ")),
            map(lambda x: x[1:]),
            map(lambda x: " | ".join(x)),
            list,
        )
        # print(f"new_pathes::{new_pathes}")
        # print(f"new_brief_pathes::{new_brief_pathes}")
        asgn_children = asgn_item["children"]

        for idx, item in enumerate(asgn_children):
            for old_brief_name, new_brief_name in zip(
                exist_brief_pathes, new_brief_pathes
            ):
                # print(f"item::{item}")
                # print(f"mode:{mode}")
                # print(f"old_brief_name:{old_brief_name}")
                if mode == item[0] and old_brief_name == item[2]:
                    try:
                        print("asgn 변경시도")
                        item[2] = new_brief_name
                        # asgn_children[idx] = item
                        print(f"asgn 변경성공 {old_brief_name} to {new_brief_name}")
                    except:
                        print(f"asgn 변경실패 {old_brief_name} to {new_brief_name}")
        return asgn_item

    def update_editing_stdType_wmItem_in(self, new_name):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]
        selected_id = self.treeview.tree.selection()
        parent_id = self.treeview.tree.parent(selected_id)
        old_name = self.treeview.tree.item(selected_id, "text")
        GWM_parent_name = self.treeview.tree.item(parent_id, "text")

        ## pjt-GWM의 키값에 바뀐 이름 업데이트 하기
        old_path_str = self.selected_item.get()
        old_path_forPjtGWMSWM = old_path_str.split(" | ")
        new_path_forPjtGWMSWM = deepcopy(old_path_forPjtGWMSWM)
        new_path_forPjtGWMSWM[-1] = new_name
        new_path_str = " | ".join(new_path_forPjtGWMSWM)

        print(f"[update editing] old_path_forPjtGWMSWM - {old_path_forPjtGWMSWM}")
        print(f"[update editing] new_path_forPjtGWMSWM - {new_path_forPjtGWMSWM}")

        if state.team_std_info[f"project-{mode}"].get(old_path_str):
            state.team_std_info[f"project-{mode}"][new_path_str] = state.team_std_info[
                f"project-{mode}"
            ].pop(old_path_str)

        # 패밀리리스트 업데이트 구간
        print(f"check SWM_parent_name::{GWM_parent_name}")

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            GWM_parent_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = GWM_parent_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: ["Top"] + x + [target_GWM_name, old_name]),
            list,
        )
        print(f"target_paths::{target_paths}")

        for path in target_paths:
            self.treeDataManager.update_node_value(
                "std-familylist",
                path,
                5,
                new_name,
            )
            self.treeDataManager.update_node_name(
                "std-familylist",
                path,
                new_name,
            )

        # project-assigntype 업데이트 필요
        pjt_assign_db = state.team_std_info["project-assigntype"]["children"]

        for idx, asgn_item in enumerate(pjt_assign_db):
            try:
                pjt_assign_db[idx] = self.update_GWMSWM_editing_to_pjtAssign(
                    [old_path_str], [new_path_str], asgn_item
                )
                # print("asgn update succes")
            except:
                pass
                # print("asgn update failed")

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])
        return target_parent_nodes

    def update_copying_stdType_wmItem_in(self, copied_node, parent_item_name):
        state = self.state
        std_familylist_db = state.team_std_info["std-familylist"]["children"][0][
            "children"
        ]

        target_parent_nodes = self.treeDataManager.find_parentnodes_by_childname_recur(
            std_familylist_db,
            parent_item_name,
        )
        target_parent_names = go(
            target_parent_nodes,
            map(lambda x: x["name"]),
            list,
        )
        target_grand_names = go(
            target_parent_names,
            map(
                lambda x: self.treeDataManager.find_parentnodes_by_childname_recur(
                    std_familylist_db,
                    x,
                )
            ),
            map(lambda x: list(x)[0]["name"]),
            list,
        )
        target_GWM_name = parent_item_name

        target_paths = go(
            zip(target_grand_names, target_parent_names),
            map(list),
            list,
            map(lambda x: x + [target_GWM_name]),
            list,
        )

        for path in target_paths:
            self.treeDataManager.match_GWMitems_to_stdFam(
                "std-familylist",
                *path,
                [copied_node],
            )

        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])

        return target_parent_nodes

    # 복사 item 항목 마다 원본 GWM 할당 상황 조회 및 반영
    def apply_originWMassign_forItem(self, path, copied_path):
        mode = self.data_kind.split("-")[-1]
        state = self.state
        ref_WM_assign_key = " | ".join(path)
        ref_WM_assign_val = state.team_std_info[f"project-{mode}"].get(
            ref_WM_assign_key
        )
        copied_WM_assign_key = " | ".join(copied_path)

        if ref_WM_assign_val:
            state.team_std_info[f"project-{mode}"].update(
                {copied_WM_assign_key: ref_WM_assign_val}
            )

    def copy_item(self):
        state = self.state

        selected_item_id = self.treeview.tree.selection()
        selected_item_name = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )[0]
        # new_name = selected_item_name + "::copy"
        new_name = (
            selected_item_name
            + f"::[{state.team_std_info.get("project-info").get("abbr", "??")}]"
        )

        parent_item_id = self.treeview.tree.parent(selected_item_id)
        if parent_item_id:
            parent_item_name = go(
                parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            parent_item_name = ""

        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        if grand_parent_item_id:
            grand_parent_item_name = go(
                grand_parent_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
        else:
            grand_parent_item_name = ""

        path = go(
            [grand_parent_item_name, parent_item_name, selected_item_name],
            filter(lambda x: x != ""),
            list,
        )
        print(f"path: {path}")

        copied_node = self.treeDataManager.copy_node(
            data_kind=self.data_kind,
            path=path,
            new_name=new_name,
            name_depth=2,
            # name_depth=[1, 2],
        )
        try:
            self.update_copying_stdType_wmItem_in(copied_node, parent_item_name)
        except:
            pass

        copied_path = deepcopy(path)
        copied_path[-1] = new_name

        self.apply_originWMassign_forItem(path, copied_path)

        state.observer_manager.notify_observers(state)

    def add_top_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Top Item", "Enter the name of the new item:"
        )
        if new_item_name:
            self.treeDataManager.add_top_level_node(self.data_kind, new_item_name)
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])

    def add_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])

    def delete_item(self):
        state = self.state

        # Get the selected item ID
        selected_item_id = self.treeview.tree.selection()
        if not selected_item_id:
            print("No item selected.")
            return
        selected_item_id = selected_item_id[0]  # Handle single selection case

        # Build the full path to the selected item
        def get_full_path(item_id):
            path = []
            current_item_id = item_id
            while current_item_id:
                # Get the name of the current item
                item_values = self.treeview.tree.item(current_item_id, "values")
                item_name = next(
                    (v for v in item_values if v), None
                )  # Get the first non-empty value
                if item_name:
                    path.insert(0, item_name)  # Add to the beginning of the path
                # Move to the parent item
                current_item_id = self.treeview.tree.parent(current_item_id)
            return path

        # Get the full path of the selected item
        full_path = get_full_path(selected_item_id)
        print(f"Deleting item with path: {full_path}")

        # Validate the constructed path
        if not full_path:
            print("Could not construct a valid path for the selected item.")
            return

        self.current_column = go(
            self.treeview.tree.item(selected_item_id, "values"),
            lambda x: next((i for i, s in enumerate(x) if s), None),
        )

        # Pass the full path to the delete_node method
        try:
            self.treeDataManager.delete_node(self.data_kind, full_path)
            print(f"Successfully deleted item with path: {full_path}")
        except Exception as e:
            print(f"Error deleting item with path {full_path}: {e}")

        if self.current_column == 2:
            self.update_deleting_stdType_wmItem_in()
        elif self.current_column == 1:
            self.update_deleting_stdType_GWMSWM_in()

        # Notify observers about the state update
        state.observer_manager.notify_observers(state, targets=["SWM", "famlist"])


###################for common_input################################################################


class TeamStd_CommonInputTreeView:
    def __init__(self, state, parent, data_kind="common-input", view_level=2):
        self.state = state
        # self.data_kind = "common-input"
        self.data_kind = data_kind

        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(
            state, lambda e: self.update(e, view_level), tag="common-input"
        )

        self.selected_item = tk.StringVar()
        self.selected_item.trace_add("write", state._notify_selected_change)
        headers = ["분류", "Abbreviation", "Description", "Input", "Unit", "Remark"]
        hdr_widths = [127, 125, 200, 50, 50, 200]

        # Compose TreeView, Style Manager, and State Observer
        title_frame = ttk.Frame(parent, width=600, height=500)
        title_frame.pack()

        tree_frame = ttk.Frame(parent, width=600, height=2000)
        tree_frame.pack()
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=3000)

        # config selection mode
        self.treeview.tree.config(selectmode="browse")
        # Tag styles
        # self.treeview.tree.tag_configure("bold", font=("Arial", 10, "bold"))
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(title_frame)
        self.set_bulkUpdateBtn(title_frame)
        self.set_rollbackBtn(title_frame)
        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        self.treeviewEditor = TreeviewEditor(state, self)
        # self.add_item = self.treeviewEditor.add_item

        # Track the last selected item with an instance attribute
        self.last_selected_item = None

        # Bind selection events
        self.treeview.tree.bind(
            "<<TreeviewSelect>>", lambda e: self.on_item_selected(e)
        )

        # Create and integrate context menu
        self.context_menu = TreeViewContextMenu(
            state,
            self.treeview,
            data_kind=self.data_kind,
            add_top=self.add_top_item,
            add=self.add_item,
            delete=self.delete_item,
        )
        # state.edit_mode_manager.register_widgets(treeCtxtMenu=[self.context_menu])

    def set_title(self, parent):
        title_font = ttk.font.Font(family="맑은 고딕", size=12)
        title_label = ttk.Label(
            parent, text="Standard Common Input Setting", font=title_font
        )
        title_label.pack(side="left", padx=5, pady=5, anchor="nw")

    def set_bulkUpdateBtn(self, parent):
        state = self.state
        bulkUp_btn = ttk.Button(
            parent,
            text="변경내용 → 산출타입 사전 전체 업데이트",
            bootstyle="info-outline",
            command=self.bulkUpdate_commoninfo,
        )
        bulkUp_btn.pack(side="left", padx=5, pady=5, anchor="nw")

    def set_rollbackBtn(self, parent):
        state = self.state
        bulkUp_btn = ttk.Button(
            parent,
            text="Common Input 값 팀표준으로 복원하기",
            bootstyle="warning-outline",
            command=self.rollback_commoninfo,
        )
        bulkUp_btn.pack(side="left", padx=5, pady=5, anchor="nw")

    def update_target_byRef(self, dbKey, target_qItemName):
        state = self.state

        selected_qItem_name = target_qItemName
        data = state.team_std_info["std-calcdict"]["children"]
        selected_node = next(
            (node for node in data if node["name"] == selected_qItem_name),
            None,
        )
        print(f"selected_node : {str(selected_node)}")

        refData = state.team_std_info[dbKey]
        targetData = selected_node
        # Flatten A data into a dictionary with 'name' as key and 'values' as value
        a_values_map = {}

        def flatten_a_data(data):
            if isinstance(data, list):
                for item in data:
                    flatten_a_data(item)
            elif isinstance(data, dict):
                a_values_map[data["name"]] = data["values"]
                flatten_a_data(data.get("children", []))

        # Flatten A data for easy lookup
        flatten_a_data(refData["children"])
        print(f'flatten_a_data : {flatten_a_data(refData["children"])}')

        # Update B data
        for child in targetData["children"]:
            if child["name"] in a_values_map:
                # Get the fourth value of the matched A data
                a_value_to_update = (
                    a_values_map[child["name"]][3]
                    if len(a_values_map[child["name"]]) > 3
                    else None
                )
                if a_value_to_update:
                    # Update the second value of B data
                    # child["values"].insert(2, a_value_to_update)
                    child["values"][2] = a_value_to_update

        for idx, node in enumerate(state.team_std_info[self.data_kind]["children"]):
            if node["name"] == selected_qItem_name:
                state.team_std_info["std-calcdict"]["children"].remove(node)
                state.team_std_info["std-calcdict"]["children"].insert(idx, targetData)
                state.log_widget.write(str(targetData))

        # state.log_widget.write(str(targetData))
        self.update(view_level=self.view_level)
        return targetData

    def bulkUpdate_commoninfo(self):
        state = self.state
        # 일괄 적용 안내 메시지 구간
        will_act = messagebox.askyesno(
            "변경 반영 확인", "현재 세팅값을 산출유형 전체에 반영하시겠습니까?"
        )

        all_keys = go(
            state.team_std_info["std-calcdict"]["children"],
            map(lambda x: x["name"]),
            # lambda x: tap(print, x),
            list,
        )
        print(f"all_keys {all_keys}")
        # 안내 메시지 결과가 참인 경우 수행할 코드 구간
        if will_act:
            for k in all_keys:
                try:
                    self.update_target_byRef("common-input", k)
                except:
                    pass
            state.observer_manager.notify_observers(state)
            messagebox.showinfo("업데이트", "업데이트 완료")

    def rollback_commoninfo(self):
        state = self.state
        # 롤백 안내 메시지 구간
        will_act = messagebox.askyesno(
            "팀표준 복원 확인", "현재 세팅값을 팀표준 값으로 복원하시겠습니까?"
        )
        std_calcdict_path = "resource/PlantArch_BIM Standard.bnote"
        print("현재 작업 디렉토리:", os.getcwd())
        print("file_path:", std_calcdict_path)
        print("파일 존재 여부:", os.path.exists(std_calcdict_path))
        # 안내 메시지 결과가 참인 경우 수행할 코드 구간
        if will_act:
            try:
                with open(std_calcdict_path, "r", encoding="utf-8") as file:
                    print(f"file ! {file}")
                    team_std_filedata = json.load(file)
                    # print(f"템플릿 로딩? {team_std_filedata}")
            except Exception as e:
                print(f"Error loading data from JSON: {e}\n")
                return False
        if team_std_filedata:
            state.team_std_info["common-input"] = deepcopy(
                team_std_filedata["common-input"]
            )
            # self.update()
            state.observer_manager.notify_observers(state)
            messagebox.showinfo("common-input 복원", "common-input 복원 완료")

    def update(self, event=None, view_level=None):
        state = self.state
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")

        selected_item_id = self.treeview.tree.focus()
        origin_indices = None

        # Extract origin indices if a valid item is focused
        if selected_item_id:
            try:
                origin_indices = self.treeview.get_item_indices(selected_item_id)
                print(origin_indices, "!!!")
            except Exception as e:
                print(f"origin_indices 추출 실패: {e}")

        """Update the TreeView whenever the state changes."""
        if self.data_kind in state.team_std_info:
            # Clear the TreeView and reload data from the updated state
            data = state.team_std_info[self.data_kind]["children"]
            self.treeview.clear_treeview()
            self.treeview.insert_data_with_levels(data)

            self.treeview.expand_tree_to_level(level=view_level)

        # Reselect the item if possible
        if origin_indices:
            try:
                self.treeview.select_item_by_indices(origin_indices)
            except Exception as e:
                self.state.log_widget.write(f"Item selection failed: {e}")

        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        try:
            if self.last_selected_item:
                self.treeview.tree.item(self.last_selected_item, tags=("normal",))
        except Exception as e:
            self.state.log_widget.write(f"Error resetting last selected item tag: {e}")

        # Get the currently selected item
        selected_item_id = self.treeview.tree.focus()
        self.last_selected_item = selected_item_id

        # Get the parent of the selected item based on its structure
        parent_item_id = self.treeview.tree.parent(selected_item_id)

        try:
            # Extract the names from the values, ensuring each level is handled appropriately
            selected_values = self.treeview.tree.item(selected_item_id, "values")
            parent_values = self.treeview.tree.item(parent_item_id, "values")

            selected_item_name = next((v for v in selected_values if v), None)
            parent_item_name = next((v for v in parent_values if v), None)

            # Format the selected path as "parent | selected"
            formatted_value = " | ".join(
                filter(None, [parent_item_name, selected_item_name])
            )

            # Update the selected item in the state
            self.selected_item.set(formatted_value)

            # Register the last selected item
            self.last_selected_item = selected_item_id

        except IndexError as e:
            self.state.log_widget.write(f"IndexError: {e}")
        except Exception as e:
            self.state.log_widget.write(f"Error processing selected item details: {e}")

    def get_parent_ids(self, selected_item_id):
        parent_ids = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.treeview.tree.parent(current_item)
            if parent_id:  # 부모 항목이 있을 경우에만 리스트에 추가
                parent_ids.append(parent_id)
            current_item = parent_id

        # Debug: print the final list of parent IDs
        self.state.log_widget.write(
            f"Parent IDs for selected item '{selected_item_id}': {parent_ids}"
        )
        return list(reversed(parent_ids))

    def add_top_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Top Item", "Enter the name of the new item:"
        )
        if new_item_name:
            self.treeDataManager.add_top_level_node(self.data_kind, new_item_name)
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)

    def add_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)

    def delete_item(self):
        state = self.state

        # Get the selected item ID
        selected_item_id = self.treeview.tree.selection()
        if not selected_item_id:
            print("No item selected.")
            return
        selected_item_id = selected_item_id[0]  # Handle single selection case

        # Build the full path to the selected item
        def get_full_path(item_id):
            path = []
            current_item_id = item_id
            while current_item_id:
                # Get the name of the current item
                item_values = self.treeview.tree.item(current_item_id, "values")
                item_name = next(
                    (v for v in item_values if v), None
                )  # Get the first non-empty value
                if item_name:
                    path.insert(0, item_name)  # Add to the beginning of the path
                # Move to the parent item
                current_item_id = self.treeview.tree.parent(current_item_id)
            return path

        # Get the full path of the selected item
        full_path = get_full_path(selected_item_id)
        print(f"Deleting item with path: {full_path}")

        # Validate the constructed path
        if not full_path:
            print("Could not construct a valid path for the selected item.")
            return

        # Pass the full path to the delete_node method
        try:
            self.treeDataManager.delete_node(self.data_kind, full_path)
            print(f"Successfully deleted item with path: {full_path}")
        except Exception as e:
            print(f"Error deleting item with path {full_path}: {e}")

        # Notify observers about the state update
        state.observer_manager.notify_observers(state)


################################


class TeamStd_FamlistTreeView:
    def __init__(
        self,
        state,
        parent,
        title="Standard Family List",
        showmode="team",
        view_level=2,
        relate_widget=None,
    ):
        self.state = state
        self.data_kind = "std-familylist"
        self.showmode = showmode
        self.view_level = view_level
        self.title = title
        self.relate_widget = relate_widget

        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(
            state, lambda e: self.update(e, view_level), tag="famlist"
        )

        self.selected_item = tk.StringVar()
        self.selected_item.trace_add("write", state._notify_selected_change)
        headers = [
            "최상위",
            "분류",
            "No",
            "Family Name",
            "GWM/SWM",
            "Item",
            "표준산출 수식",
            "Description",
            "표준산출유형 번호",
        ]
        # hdr_widths = [0, 60, 20, 150, 100, 100, 100, 230, 20]
        hdr_widths = [0, 60, 5, 150, 100, 100, 100, 230, 20]

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(parent, width=600, height=2000)
        tree_frame.pack(fill="both")
        self.tree_frame = tree_frame

        button_frame = ttk.Frame(tree_frame, width=600)
        button_frame.pack(pady=10)

        def set_tree_row(event=None):
            # rowheight = round(rowheight_scalebar.get())
            # state._rowheight = tk.IntVar()
            state._rowheight.set(
                round(rowheight_scalebar.get()),
            )

            DefaultTreeViewStyleManager.apply_style(
                self.treeview.tree,
                state._rowheight.get(),
            )
            self.treeview.tree.update_idletasks()

        rowheight_scalebar = ttk.Scale(
            button_frame,
            bootstyle="info",
            variable=state._rowheight,
            command=set_tree_row,
            from_=20,
            to=35,
            length=5,
        )
        rowheight_scalebar.pack(padx=10, pady=5, side="left")

        # Add filter button
        if showmode != "team":
            filter_button = ttk.Button(
                button_frame,
                text="Filter Unused",
                command=lambda: self.update(self, filter_mode=True),
                bootstyle="info-outline",
            )
            filter_button.pack(padx=10, pady=5, side="left")

            # Add Reset button
            reset_button = ttk.Button(
                button_frame,
                text="Reset TreeView",
                command=lambda: self.update(self, filter_mode=False),
                bootstyle="info-outline",
            )
            reset_button.pack(padx=10, pady=5, side="left")

        # Add ComboBox for selecting view level
        if showmode == "project_assign":
            combo_range = list(range(1, 4))
        else:
            combo_range = list(range(1, 5))

        self.level_combobox = ttk.Combobox(
            button_frame,
            values=combo_range,
            state="readonly",
            width=7,
        )
        self.level_combobox.current(view_level - 1)  # Set default level
        self.level_combobox.bind("<<ComboboxSelected>>", self.on_level_selected)
        self.level_combobox.pack(padx=10, pady=5, side="left")

        # export 버튼
        self.export_btn = ttk.Button(
            button_frame,
            text="Export to Excel",
            command=self.export_visible_treeview_to_excel,
            bootstyle="success-outline",
        )
        self.export_btn.pack(padx=10, pady=5, side="left")

        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.tree = self.treeview.tree
        self.treeview.tree.config(height=3000)

        # self.state.on_level_selected = self.on_level_selected

        # config selection mode
        self.treeview.tree.config(selectmode="browse")
        # Tag styles
        # self.treeview.tree.tag_configure("bold", font=("Arial", 10, "bold"))
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(tree_frame)

        # 검색 매니저 붙이기
        self.search = TreeviewSearchManager(self.treeview.tree, tree_frame)

        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)
        self.treeview.tree.column("GWM/SWM", anchor="center")  # Center-align

        # set treeview_editor class
        # self.treeviewEditor = TreeviewEditor(state, self)
        self.treeviewEditor = TreeviewEditor_forFamilyList(state, self)

        # Track the last selected item with an instance attribute
        self.last_selected_item = None
        self.state.on_level_selected = self.on_level_selected

        # Bind selection events
        self.treeview.tree.bind("<<TreeviewSelect>>", self.on_item_selected)
        self.treeview.tree.bind(
            "<<TreeviewOpen>>",
            DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(
                self.treeview.tree
            ),
        )
        self.treeview.tree.bind(
            "<<TreeviewClose>>",
            DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(
                self.treeview.tree
            ),
        )
        self.treeview.tree.bind(
            "<<TreeviewSelect>>",
            DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(
                self.treeview.tree
            ),
        )

        # Create and integrate context menu
        self.context_menu = TreeViewContextMenu_FamilyList(
            state,
            self.treeview,
            data_kind=self.data_kind,
            add_top=self.add_top_item,
            add=self.add_item_forLv1,
            delete=self.delete_item,
        )
        # state.edit_mode_manager.register_widgets(treeCtxtMenu=[self.context_menu])

    def export_visible_treeview_to_excel(self):
        """Exports only expanded (visible) items from the treeview to an Excel file with formatting."""
        treeview = self.tree

        # Ask user for save location
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
        )

        if not file_path:
            return  # User canceled save dialog

        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Family List"

        # Extract column headers
        columns = [treeview.heading(col, "text") for col in treeview["columns"]]
        ws.append(columns)  # Write headers to first row

        # Apply header styles (bold, center alignment, fill color)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Define cell formatting (alignment & border)
        center_align = Alignment(horizontal="center", vertical="center")
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Recursive function to write only visible treeview data
        def write_visible_data(parent="", level=0, row_index=2):
            green_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            pink_fill = PatternFill(
                start_color="f173ff", end_color="f173ff", fill_type="solid"
            )
            mild_pink_fill = PatternFill(
                start_color="f6a6ff", end_color="f6a6ff", fill_type="solid"
            )

            def check_str_inRow(row, str_):
                res = False
                for c in row:
                    if str_ in c:
                        res = True
                return res

            for item in treeview.get_children(parent):
                if not treeview.item(
                    item, "open"
                ):  # Folded items should not be included
                    name = treeview.item(item, "text")
                    values = treeview.item(item, "values")

                    # Process values (avoid Excel formula issues with '=')
                    values_ = [v.replace("=", "'=") if "=" in v else v for v in values]

                    # Write row to Excel
                    ws.append(values_)

                    # Apply formatting to each row
                    for col_num, value in enumerate(values_, 1):
                        cell = ws.cell(row=row_index, column=col_num)
                        # cell.alignment = center_align
                        # if col_num > 5:
                        #     cell.border = border_style

                    # print(f"현재레벨 - {self.level_combobox.get()}")
                    if self.level_combobox.get() == "4":
                        ws.cell(row=row_index, column=6).border = border_style
                        ws.cell(row=row_index, column=7).border = border_style

                    row_index += 1  # Move to next row
                    continue  # Skip folded items

                name = treeview.item(item, "text")
                values = treeview.item(item, "values")
                values_ = [v.replace("=", "'=") if "=" in v else v for v in values]

                # print(f"row ~ {values_}")
                if check_str_inRow(values, "GWM") or check_str_inRow(values, "SWM"):
                    ws.append([""])
                    row_index += 1  # Move to next row

                # Write row to Excel
                ws.append(values_)

                # # Apply formatting to each row

                row_data = list(enumerate(values_, 1))
                for col_num, value in row_data:
                    if re.match(r"\d+\.[A-Za-z]+", value):
                        for c in range(col_num, len(columns)):
                            cell = ws.cell(row=row_index, column=c)
                            cell.fill = green_fill
                    elif re.match(r"^\d+", value):
                        # print(
                        #     f"check_asterisk_inRow(value) {check_asterisk_inRow(value)}"
                        # )
                        if check_str_inRow(values_, "***"):
                            for c in range(col_num, len(columns)):
                                cell = ws.cell(row=row_index, column=c)
                                cell.fill = pink_fill
                        else:
                            for c in range(col_num, len(columns)):
                                cell = ws.cell(row=row_index, column=c)
                                cell.fill = mild_pink_fill

                row_index += 1

                # Recursively process only expanded child nodes
                row_index = write_visible_data(item, level + 1, row_index)

            return row_index  # Return the current row index for recursion tracking

        # Start writing from root nodes
        write_visible_data()

        # Auto-adjust column widths based on content length
        for col_num, column_title in enumerate(columns, 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col_num).value) or "")
                for row in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = (
                max_length + 2
            )

        ########## std-calcdict 출력 부 #########
        ws2 = wb.create_sheet(title="표준산출유형사전")
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )

        row_idx = 1  # 엑셀의 행 번호는 1부터 시작
        data = go(
            self.state.team_std_info["std-calcdict"]["children"],
            deepcopy,
            lambda x: sorted(x, key=lambda i: sort_func_forCalctype(i)),
        )
        for node in data:
            # calcType 추가 + 스타일 적용
            calcType = [node["name"], "", ""]
            ws2.append(calcType)
            ws2.cell(row=row_idx, column=1).fill = green_fill
            ws2.cell(row=row_idx, column=2).fill = green_fill
            ws2.cell(row=row_idx, column=3).fill = green_fill
            row_idx += 1

            for childnode in node["children"]:
                ws2.append(childnode["values"])
                row_idx += 1

            ws2.append([""])  # 빈 줄
            row_idx += 1

        # Auto-adjust column widths based on content length
        for col_num, column_title in enumerate(columns, 1):
            max_length = max(
                len(str(ws2.cell(row=row, column=col_num).value) or "")
                for row in range(1, ws2.max_row + 1)
            )
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = (
                max_length + 2
            )

        # Save the Excel file
        wb.save(file_path)
        print(f"Formatted Treeview data exported successfully to {file_path}")

        # 6. 메시지 박스 띄우기
        open_file = messagebox.askyesno(
            "저장 완료", "엑셀 파일 저장이 완료되었습니다.\n파일을 여시겠습니까?"
        )

        # 7. 사용자가 "예(Yes)"를 선택한 경우 엑셀 파일 열기
        if open_file:
            try:
                os.startfile(file_path)  # Windows에서 엑셀 파일 열기
                # subprocess.Popen(["start", file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다.\n{e}")

    def set_title(self, parent, text=None):
        if not text:
            text = self.title
        title_font = ttk.font.Font(family="맑은 고딕", size=11)
        title_label = ttk.Label(parent, text=text, font=title_font)
        title_label.pack(padx=5, pady=5, anchor="w")

    def remove_items_unused(self, _data_list, _depth, _rulelist):
        """
        Removes items with an underscore in their 'name' at depth 2, along with their children.

        Parameters:
            data_list (list): A list of hierarchical data structures.

        Returns:
            list: Cleaned data with specified items removed.
        """
        data_list = deepcopy(_data_list)

        def clean_children(children, depth):
            cleaned_children = []
            for child in children:
                # At _depth , check for _rule in the 'name' field
                if depth == _depth and child.get("name", "") not in _rulelist:
                    continue  # Skip this item and its children
                # Recursively clean if 'children' field exists
                if "children" in child:
                    child["children"] = clean_children(child["children"], depth + 1)
                cleaned_children.append(child)
            return cleaned_children

        # Process each top-level item in the list
        for item in data_list:
            if "children" in item:
                # item["children"] = clean_children(item["children"], _depth)
                item["children"] = clean_children(item["children"], 1)

        return data_list

    def update(self, event=None, view_level=None, filter_mode=False):
        state = self.state
        # self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")

        self.valid_names = []
        if (
            state.team_std_info.get("project-assigntype")
            and state.current_building.get() != "건물을 선택하세요"
        ):
            self.valid_names = go(
                state.team_std_info.get("project-assigntype").get("children"),
                filter(lambda x: x["values"][1] == state.current_building.get()),
                map(lambda x: x["values"][-1]),
                map(lambda x: x.split(" | ")[-1]),
                list,
            )

        selected_item_id = self.treeview.tree.focus()
        origin_indices = None

        # Extract origin indices if a valid item is focused
        if selected_item_id:
            try:
                origin_indices = self.treeview.get_item_indices(selected_item_id)
                print(origin_indices, "!!!")
            except Exception as e:
                print(f"origin_indices 추출 실패: {e}")

        """Update the TreeView whenever the state changes."""
        if self.data_kind in state.team_std_info:
            if self.showmode == "team":
                data = self.treeview.remove_items_with_rule(
                    state.team_std_info[self.data_kind]["children"],
                    _depth=3,
                    _rule="::",
                )
                # print(f"\n걸러진데이터!\n")
            else:
                if filter_mode:
                    data = self.remove_items_unused(
                        state.team_std_info[self.data_kind]["children"],
                        _depth=2,
                        _rulelist=self.valid_names,
                    )
                else:
                    data = state.team_std_info[self.data_kind]["children"]
            # Clear the TreeView and reload data from the updated state
            self.treeview.clear_treeview()
            self.treeview.insert_data_with_levels(data)

            self.on_level_selected(event=None)

        # Reselect the item if possible
        if origin_indices:
            try:
                self.treeview.select_item_by_indices(origin_indices)
            except Exception as e:
                self.state.log_widget.write(f"Item selection failed: {e}")

        # Apply row colors dynamically
        DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(
            self.treeview.tree
        )

        try:
            self.place_selected_item_at_top()
        except:
            pass

        # self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        # Get the currently selected item
        state = self.state

        # 레빗 타입 별 WM 입력 오류 방지를 위한 초기화
        state.selected_rvtTypes.set(None)
        state.selected_rvtTypes_forLabel.set(None)
        # 표준타입 미입력시 직전 선택 내용 잔존 방지지
        state.typeAssign_treeview.treeview.clear_treeview()
        state.project_WM_perRVT_SheetView.sheet.set_sheet_data([])

        selected_item_id = self.treeview.tree.focus()
        try:
            if (
                self.last_selected_item
                and self.last_selected_item not in selected_item_id
            ):
                original_tag = self.treeview.tree.item(self.last_selected_item, "tags")[
                    0
                ]
                self.treeview.tree.item(self.last_selected_item, tags=(original_tag,))

            # 현재 선택된 항목의 스타일을 변경
            if selected_item_id:
                for item in selected_item_id:
                    self.treeview.tree.item(item, tags=("selected",))

        except Exception as e:
            self.state.log_widget.write(f"Error resetting last selected item tag: {e}")

        self.treeview.last_selected_item = selected_item_id
        self.state.log_widget.write(
            f"select tree item!!! {self.treeview.tree.item(selected_item_id, 'text')}"
        )

        # Get the parent and grandparent of the selected item
        parent_item_id = self.treeview.tree.parent(selected_item_id)
        grand_parent_item_id = self.treeview.tree.parent(parent_item_id)
        # Extract the names from the values, ensuring each level is handled appropriately
        try:
            # Get the name of the selected item, parent, and grandparent
            selected_item_name = go(
                self.treeview.tree.item(selected_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            parent_item_name = go(
                self.treeview.tree.item(parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0],
            )
            grand_parent_item_name = go(
                self.treeview.tree.item(grand_parent_item_id, "values"),
                filter(lambda x: x != ""),
                list,
                lambda x: x[0] or "",
            )

            # Format the selected path as "grandparent | parent | selected"
            formatted_value = " | ".join(
                filter(
                    # None, [grand_parent_item_name, parent_item_name, selected_item_name]
                    None,
                    [grand_parent_item_name, parent_item_name, selected_item_name],
                )
            )
            self.state.log_widget.write(
                f"패밀리리스트 선택 항목 포맷 밸류 {formatted_value}"
            )

            # Update the selected item in the state
            self.selected_item.set(formatted_value)
            self.state.log_widget.write(
                f"패밀리리스트 선택 항목 {self.selected_item.get()}"
            )
            # Register the last selected item
            self.last_selected_item = selected_item_id

            def select_first_item(tree):
                children = tree.get_children()  # Get all children of the root
                if children:
                    first_item = children[0]
                    tree.selection_set(first_item)  # Select the first item
                    tree.focus(first_item)  # Set focus to the first item
                    tree.see(first_item)  # Ensure the first item is visible
                    # tree.focus_set()

            select_first_item(state.typeAssign_treeview.treeview.tree)

        except Exception as e:
            self.state.log_widget.write(f"Error processing selected item details: {e}")

    def place_selected_item_at_top(self):
        tree = self.treeview.tree
        # Handle focused item visibility
        focused_item = self.treeview.tree.focus()
        if focused_item:
            self.treeview.bring_item_to_top(focused_item)
            self.treeview.tree.focus(focused_item)
        else:
            # Optionally, focus on the first visible item if no item is focused
            first_root = self.treeview.tree.get_children("")[0]
            self.treeview.bring_item_to_top(first_root)
            self.treeview.tree.focus(first_root)

    def on_level_selected(self, event):
        """Handle combo box selection and expand the tree to the selected level."""
        try:
            selected_level = int(self.level_combobox.get())

            # Expand the tree to the selected level
            self.treeview.expand_tree_to_level(level=selected_level)

            # Apply row colors dynamically
            DefaultTreeViewStyleManager.apply_dynamic_alternate_row_colors(
                self.treeview.tree
            )

            # Handle focused item visibility
            self.place_selected_item_at_top()

        except ValueError as e:
            self.state.log_widget.write(f"Invalid level selected: {e}")
        except Exception as e:
            self.state.log_widget.write(
                f"Error expanding tree to level {selected_level}: {e}"
            )

    def get_parent_ids(self, selected_item_id):
        parent_ids = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.treeview.tree.parent(current_item)
            if parent_id:  # 부모 항목이 있을 경우에만 리스트에 추가
                parent_ids.append(parent_id)
            current_item = parent_id

        # Debug: print the final list of parent IDs
        self.state.log_widget.write(
            f"Parent IDs for selected item '{selected_item_id}': {parent_ids}"
        )
        return list(reversed(parent_ids))

    def add_top_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Top Item", "Enter the name of the new item:"
        )
        if new_item_name:
            self.treeDataManager.add_top_level_node(self.data_kind, new_item_name)
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)

    def add_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)

    def add_item_forLv1(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node_forFamilyList(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)

    def delete_item(self):
        state = self.state

        # Get the selected item ID
        selected_item_id = self.treeview.tree.selection()
        if not selected_item_id:
            print("No item selected.")
            return
        selected_item_id = selected_item_id[0]  # Handle single selection case

        # Build the full path to the selected item
        def get_full_path(item_id):
            path = []
            current_item_id = item_id
            while current_item_id:
                # Get the name of the current item
                item_values = self.treeview.tree.item(current_item_id, "values")
                item_name = next(
                    (v for v in item_values if v), None
                )  # Get the first non-empty value
                if item_name:
                    path.insert(0, item_name)  # Add to the beginning of the path
                # Move to the parent item
                current_item_id = self.treeview.tree.parent(current_item_id)
            return path

        # Get the full path of the selected item
        full_path = get_full_path(selected_item_id)
        print(f"Deleting item with path: {full_path}")

        # Validate the constructed path
        if not full_path:
            print("Could not construct a valid path for the selected item.")
            return

        # Pass the full path to the delete_node method
        try:
            self.treeDataManager.delete_node(self.data_kind, full_path)
            print(f"Successfully deleted item with path: {full_path}")
        except Exception as e:
            print(f"Error deleting item with path {full_path}: {e}")

        # Notify observers about the state update
        state.observer_manager.notify_observers(state)


class TeamStd_calcDict_TreeView:
    def __init__(self, state, parent, relate_widget, data_kind=None, view_level=3):
        self.state = state
        # self.data_kind = data_kind
        self.data_kind = "std-calcdict"
        self.relate_widget = relate_widget
        self.view_level = view_level
        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(
            state,
            lambda e: self.update(event=e, view_level=self.view_level),
            tag="calcDict",
        )
        self.selected_item_relate_widget = relate_widget.selected_item
        headers = [
            "표준산출유형 번호",
            "심벌키",
            "심벌값",
        ]
        # hdr_widths = [100, 50, 100]
        hdr_widths = [50, 50, 200]

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(parent, width=600, height=2000)
        self.update_button = self.add_update_button(tree_frame)
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=3000)

        # config selection mode
        self.treeview.tree.config(selectmode="extended")

        # Tag styles
        self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        self.set_title(tree_frame)
        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        self.treeviewEditor = TreeviewEditor(state, self)

        # Track the last selected item with an instance attribute
        self.last_selected_item = None
        # Bind selection events
        self.treeview.tree.bind("<<TreeviewSelect>>", self.on_item_selected)

        self.context_menu = TreeViewContextMenu(
            state,
            self.treeview,
            data_kind=self.data_kind,
            add_top=self.add_top_item,
            add=self.add_item,
            delete=self.delete_item,
        )

    def set_title(self, parent):
        title_font = ttk.font.Font(family="맑은 고딕", size=9)
        title_label = ttk.Label(
            parent, text="Matched WMs for Selected Standard Types", font=title_font
        )
        title_label.pack(padx=5, pady=5, anchor="w")

    def add_update_button(self, parent):
        """Adds an Update button to the tree view."""
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill="x", pady=5)

        update_button = ttk.Button(
            button_frame,
            text="Update from common-input",
            command=lambda: self.update_target_byRef("common-input"),
        )
        update_button.pack(padx=10, side="left")
        return update_button

    def update_target_byRef(self, dbKey):
        state = self.state

        # pattern = r"^\d+\.\d+$"
        pattern = r"^\d+\.\d+[a-zA-Z]?$"
        selected_NoItem = go(
            self.selected_item_relate_widget.get().split(" | "),
            filter(lambda x: re.match(pattern, x)),
            next,
        )
        print(f"selected_NoItem{selected_NoItem}")
        selected_qItem_name = f"Q{selected_NoItem}"
        print(f"selected_qItem_name{selected_qItem_name}")
        # Ensure the data kind exists in the team standard information
        data = state.team_std_info[self.data_kind]["children"]
        selected_node = next(
            (node for node in data if node["name"] == selected_qItem_name),
            None,
        )
        print(f"selected_node{selected_node}")
        # state.log_widget.write(str(selected_node))

        refData = state.team_std_info[dbKey]
        targetData = selected_node
        # Flatten A data into a dictionary with 'name' as key and 'values' as value
        a_values_map = {}

        def flatten_a_data(data):
            if isinstance(data, list):
                for item in data:
                    flatten_a_data(item)
            elif isinstance(data, dict):
                a_values_map[data["name"]] = data["values"]
                flatten_a_data(data.get("children", []))

        # Flatten A data for easy lookup
        flatten_a_data(refData["children"])

        # Update B data
        for child in targetData["children"]:
            if child["name"] in a_values_map:
                # Get the fourth value of the matched A data
                a_value_to_update = (
                    a_values_map[child["name"]][3]
                    if len(a_values_map[child["name"]]) > 3
                    else None
                )
                if a_value_to_update:
                    # Update the second value of B data
                    # child["values"].insert(2, a_value_to_update)
                    child["values"][2] = a_value_to_update

        for idx, node in enumerate(state.team_std_info[self.data_kind]["children"]):
            if node["name"] == selected_qItem_name:
                state.team_std_info[self.data_kind]["children"].remove(node)
                state.team_std_info[self.data_kind]["children"].insert(idx, targetData)
                state.log_widget.write(str(targetData))

        # state.log_widget.write(str(targetData))
        self.update(view_level=self.view_level)
        return targetData

    def get_selected_row_values(self, event):
        """Handle right-click on a Treeview row to get its values."""
        # Identify the item under the mouse pointer
        item_id = self.treeview.tree.selection()
        if item_id:  # If an item is found
            values = self.treeview.tree.item(
                item_id, "values"
            )  # Get the values of the clicked row
            print(f"Right-clicked row values: {values}")
            return values

    def update(self, event=None, view_level=None):
        view_level = self.view_level
        self.state.log_widget.write(f"view_level {view_level}")
        state = self.state
        """Update the TreeView whenever the state changes."""
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")
        print(f"관련위젯 선택항목 출력 : {self.selected_item_relate_widget.get()}")

        try:
            # # Split the selected item path to find the grandparent, parent, and selected item names
            # pattern = r"^\d+\.\d+$"
            pattern = r"^\d+\.\d+[a-zA-Z]?$"
            _selected_NoItem = go(
                self.selected_item_relate_widget.get().split(" | "),
                reversed,
                list,
            )
            selected_NoItem = go(
                _selected_NoItem,
                filter(lambda x: re.match(pattern, x)),
                next,
            )
            selected_idx = _selected_NoItem.index(selected_NoItem)
            print(f"selected_idx::~   {selected_idx}")

            std_famtree = self.relate_widget.treeview.tree
            std_famtree_select = std_famtree.selection()

            if selected_idx == 0:
                selected_qItem_name = std_famtree.item(std_famtree_select, "values")[-1]
            elif selected_idx == 1:
                tgt = std_famtree.parent(std_famtree_select)
                selected_qItem_name = std_famtree.item(tgt, "values")[-1]
            elif selected_idx == 2:
                _parent = std_famtree.parent(std_famtree_select)
                tgt = std_famtree.parent(_parent)
                selected_qItem_name = std_famtree.item(tgt, "values")[-1]
                pass

            print(f"selected_qItem_name::~   {selected_qItem_name}")

            # selected_qItem_name = f"Q{selected_NoItem}"

            # Ensure the data kind exists in the team standard information
            if self.data_kind in state.team_std_info:
                data = state.team_std_info[self.data_kind]["children"]
                selected_node = next(
                    (node for node in data if node["name"] == selected_qItem_name),
                    None,
                )
                # print(selected_node)

                if selected_node:
                    # Clear the TreeView and insert the data for the selected node
                    self.treeview.clear_treeview()
                    self.treeview.insert_data_with_levels([selected_node])

                    # self.treeview.expand_tree_to_level(level=view_level)

        except Exception as e:
            self.state.log_widget.write(
                f"{self.__class__.__name__} > update 메소드 진입 안됩니다~: {e}"
            )

        # self.treeview.expand_all_items()
        self.treeview.expand_tree_to_level(level=view_level)
        # self.treeview.expand_tree_to_level(level=3)
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")

    def on_item_selected(self, event):
        # if self.last_selected_item:
        #     self.treeview.tree.item(self.treeview.last_selected_item, tags=("normal",))

        self.state.log_widget.write("on_item_selected_시작")
        selected_items_id = self.treeview.tree.selection()
        selected_items_name = go(
            selected_items_id,
            map(lambda x: list(self.treeview.tree.item(x, "values"))),
            list,
            map(lambda x: x[0]),
            list,
        )
        self.state.log_widget.write(str(selected_items_name))
        self.state.selected_matchedWMs = selected_items_name
        # # 마지막 선택항목으로 재등록
        # self.last_selected_item = selected_items_id[0]
        self.state.log_widget.write("on_item_selected_종료")

    def get_parent_ids(self, selected_item_id):
        parent_ids = []
        current_item = selected_item_id

        while current_item:
            parent_id = self.treeview.tree.parent(current_item)
            if parent_id:  # 부모 항목이 있을 경우에만 리스트에 추가
                parent_ids.append(parent_id)
            current_item = parent_id

        # Debug: print the final list of parent IDs
        self.state.log_widget.write(
            f"Parent IDs for selected item '{selected_item_id}': {parent_ids}"
        )
        return list(reversed(parent_ids))

    def add_top_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Top Item", "Enter the name of the new item:"
        )
        if new_item_name:
            self.treeDataManager.add_top_level_node(self.data_kind, new_item_name)
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["calcDict"])

    def add_item(self):
        state = self.state
        # Prompt the user for the new item name
        new_item_name = simpledialog.askstring(
            "Add Item", "Enter the name of the new item:"
        )
        if new_item_name:
            selected_item_id = self.treeview.tree.selection()
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]
            self.treeDataManager.add_child_node(
                self.data_kind, selected_item_name, new_item_name
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state, targets=["calcDict"])

    def delete_item(self):
        state = self.state

        pattern = r"^\d+\.\d+$"
        _selected_NoItem = go(
            self.selected_item_relate_widget.get().split(" | "),
            reversed,
            list,
        )
        selected_NoItem = go(
            _selected_NoItem,
            filter(lambda x: re.match(pattern, x)),
            next,
        )
        selected_idx = _selected_NoItem.index(selected_NoItem)
        print(f"selected_idx::~   {selected_idx}")

        std_famtree = self.relate_widget.treeview.tree
        std_famtree_select = std_famtree.selection()

        if selected_idx == 0:
            selected_qItem_name = std_famtree.item(std_famtree_select, "values")[-1]
        elif selected_idx == 1:
            tgt = std_famtree.parent(std_famtree_select)
            selected_qItem_name = std_famtree.item(tgt, "values")[-1]
        elif selected_idx == 2:
            _parent = std_famtree.parent(std_famtree_select)
            tgt = std_famtree.parent(_parent)
            selected_qItem_name = std_famtree.item(tgt, "values")[-1]
            pass

        selected_item_id = self.treeview.tree.selection()
        selected_item_names = go(
            selected_item_id,
            lambda x: self.treeview.tree.item(x, "values"),
            filter(lambda x: x != ""),
            list,
        )

        if selected_item_names:
            selected_item_name = selected_item_names[0]
        else:
            selected_item_name = ""

        self.treeDataManager.delete_node(
            self.data_kind,
            [selected_qItem_name, selected_item_name],
        )
        # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
        state.observer_manager.notify_observers(state, targets=["calcDict"])


class BuildingList_TreeView:
    def __init__(self, state, parent, view_level=2, *args, **kwargs):
        # super().__init__(parent, *args, **kwargs)
        # self.pack(fill="both", expand=True)
        self.state = state
        self.data_kind = "project-buildinglist"
        self.treeDataManager = TreeDataManager_treeview(state, self)
        self.state_observer = StateObserver(state, lambda e: self.update(e, view_level))
        self.selected_item = tk.StringVar()
        self.selected_item.trace_add("write", state._notify_selected_change)

        self.frame = ttk.Frame(parent)
        self.frame.pack(fill="both", expand=True)
        # Title Label
        ttk.Label(self.frame, text="Building List", font=("Arial", 16)).pack(pady=10)

        # Input Field
        ttk.Label(self.frame, text="Enter Building Name:", font=("Arial", 12)).pack(
            pady=5
        )
        self.entry = ttk.Entry(self.frame, font=("Arial", 12), bootstyle=SECONDARY)
        self.entry.pack(pady=5, fill="x", padx=10)

        # Buttons for Add and Delete
        button_frame = ttk.Frame(self.frame)
        button_frame.pack(pady=10)

        ttk.Button(
            button_frame,
            text="Add Building",
            command=self.add_building,
            bootstyle=SUCCESS,
        ).pack(side="left", padx=5)
        ttk.Button(
            button_frame,
            text="Delete Building",
            # command=self.delete_item,
            command=self.delete_building,
            bootstyle=DANGER,
        ).pack(side="left", padx=5)

        # Status Label
        self.status_label = ttk.Label(self.frame, text="", font=("Arial", 10))
        self.status_label.pack(pady=5)

        headers = ["Building Name"]
        hdr_widths = [400]

        # Compose TreeView, Style Manager, and State Observer
        tree_frame = ttk.Frame(self.frame, width=600, height=2000)
        tree_frame.pack()
        self.tree_frame = tree_frame
        self.treeview = BaseTreeView(state, tree_frame, headers)
        self.treeview.tree.config(height=3000)

        # config selection mode
        self.treeview.tree.config(selectmode="browse")
        # Tag styles
        self.treeview.tree.tag_configure("bold", font=("Arial", 10, "bold"))
        # self.treeview.tree.tag_configure("normal", font=("Arial Narrow", 10))
        # Set up UI
        # self.set_title(tree_frame)
        self.scroll_widget = ScrollbarWidget(tree_frame, self.treeview.tree)
        self.treeview.tree.pack(expand=True, fill="both", side="left")
        self.treeview.setup_columns(headers, hdr_widths)

        # set treeview_editor class
        # self.treeviewEditor = TreeviewEditor(state, self)
        self.treeviewEditor = TreeviewEditor_forBuildingList(state, self)

    def add_building(self):
        """Add a building name to the TreeView."""
        state = self.state
        building_name = self.entry.get().strip()
        state.log_widget.write(f"building_name: {building_name}")
        if building_name:
            self.treeDataManager.add_top_level_node(self.data_kind, building_name)

            self.entry.delete(0, "end")  # Clear the entry field
            self.status_label.config(
                text=f"Building '{building_name}' added successfully.",
                bootstyle=SUCCESS,
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)
        else:
            self.status_label.config(
                text="Please enter a building name.", bootstyle=DANGER
            )
        state.log_widget.write(
            self.treeview.tree.item(self.treeview.tree.get_children(""))
        )

    def delete_building(self):
        """Delete the selected building from the TreeView."""
        state = self.state
        selected_item_id = self.treeview.tree.selection()

        delete_will = messagebox.askokcancel(
            "건물삭제",
            "해당 건물에 대한 레빗 할당 데이터도 함께 지워집니다. 정말 삭제하시겠습니까?",
        )

        if selected_item_id and delete_will:
            selected_item_name = go(
                selected_item_id,
                lambda x: self.treeview.tree.item(x, "values"),
                filter(lambda x: x != ""),
                list,
            )[0]

            self.treeDataManager.delete_node(
                self.data_kind,
                [selected_item_name],
            )
            # 상태가 업데이트되었을 때 모든 관찰자에게 알림을 보냄
            state.observer_manager.notify_observers(state)
            self.status_label.config(text="Successfully deleted.", bootstyle=DANGER)

            pjt_assign_db = state.team_std_info["project-assigntype"]["children"]
            new_assign_db = []
            for idx, dic in enumerate(pjt_assign_db):
                if dic["values"][1] != selected_item_name:
                    new_assign_db.append(dic)
            state.team_std_info["project-assigntype"]["children"] = new_assign_db
            # Notify observers about the state update
            state.observer_manager.notify_observers(state)
        else:
            self.status_label.config(
                text="Please select a building to delete.", bootstyle=DANGER
            )

    def delete_item(self):
        state = self.state

        # Get the selected item ID
        selected_item_id = self.treeview.tree.selection()
        if not selected_item_id:
            state.log_widget.write("No item selected.")
            self.status_label.config(
                text="Please select a building to delete.", bootstyle=DANGER
            )
            return
        selected_item_id = selected_item_id[0]  # Handle single selection case

        # Build the full path to the selected item
        def get_full_path(item_id):
            path = []
            current_item_id = item_id
            while current_item_id:
                # Get the name of the current item
                item_values = self.treeview.tree.item(current_item_id, "values")
                item_name = next(
                    (v for v in item_values if v), None
                )  # Get the first non-empty value
                if item_name:
                    path.insert(0, item_name)  # Add to the beginning of the path
                # Move to the parent item
                current_item_id = self.treeview.tree.parent(current_item_id)
            return path

        # Get the full path of the selected item
        full_path = get_full_path(selected_item_id)
        print(f"Deleting item with path: {full_path}")

        # Validate the constructed path
        if not full_path:
            print("Could not construct a valid path for the selected item.")
            return

        # Pass the full path to the delete_node method
        try:
            self.treeDataManager.delete_node(self.data_kind, full_path)
            state.log_widget.write(f"Successfully deleted item with path: {full_path}")
            self.status_label.config(text="Successfully deleted.", bootstyle=DANGER)
        except Exception as e:
            print(f"Error deleting item with path {full_path}: {e}")

        # Notify observers about the state update
        state.observer_manager.notify_observers(state)

    def update(self, event=None, view_level=None):
        state = self.state
        self.state.log_widget.write(f"{self.__class__.__name__} > update 메소드 시작")

        selected_item_id = self.treeview.tree.focus()
        origin_indices = None

        # Extract origin indices if a valid item is focused
        if selected_item_id:
            try:
                origin_indices = self.treeview.get_item_indices(selected_item_id)
                print(origin_indices, "!!!")
            except Exception as e:
                print(f"origin_indices 추출 실패: {e}")

        """Update the TreeView whenever the state changes."""
        if self.data_kind in state.team_std_info:
            data = state.team_std_info[self.data_kind]["children"]

            # Clear the TreeView and reload data from the updated state
            self.treeview.clear_treeview()
            self.treeview.insert_data_with_levels(data)

            self.treeview.expand_tree_to_level(level=view_level)

        # Reselect the item if possible
        if origin_indices:
            try:
                self.treeview.select_item_by_indices(origin_indices)
            except Exception as e:
                state.log_widget.write(f"Item selection failed: {e}")

        state.log_widget.write(f"{self.__class__.__name__} > update 메소드 종료")
