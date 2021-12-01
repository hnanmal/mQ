# Load the Python Standard and DesignScript Libraries
import os
python_rootpath = os.getenv('LOCALAPPDATA')+'\Programs\Python'
file_list = os.listdir(python_rootpath)
python_version = file_list[0]
import sys
sys.path.append(os.getenv('LOCALAPPDATA').replace('\\','\\\\') + f'\Programs\Python\{python_version}\Lib\site-packages')
import clr
clr.AddReference('RevitAPI')
import Autodesk

clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *

clr.AddReference('RevitNodes')
import Revit
clr.ImportExtensions(Revit.Elements)

clr.ImportExtensions(Revit.GeometryConversion)
clr.ImportExtensions(Revit.GeometryReferences)

clr.AddReference('RevitServices')
import RevitServices
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

doc = DocumentManager.Instance.CurrentDBDocument

import random
import itertools
from functools import reduce
from collections import Iterable
import datetime
import openpyxl

# The inputs to this node will be stored as a list in the IN variables.
dataEnteringNode = IN
inputFilePath = IN[0]
inputSheetName = IN[1]

def importExcel(path, SheetNameIn):
    wb_obj = openpyxl.load_workbook(str(path), data_only=True)
    sheet_obj = wb_obj.get_sheet_by_name(str(SheetNameIn))

    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    listOut = []

    for i in range(1, m_row +1):
        listTemp = []
        for j in range(1, m_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            listTemp.append(cell_obj.value)
            if len(listTemp) == m_col:
                listOut.append(listTemp)
    return listOut

def checkAllElement(list, target):
    tmp=[]
    for i in list:
        if i == target:
            tmp.append(True)
        else:
            tmp.append(False)
    return all(tmp)


# Place your code below this line

result = importExcel(inputFilePath, inputSheetName)

for i in result:## null만 있는 Excel Row 제거
    if checkAllElement(i, None):
        result.remove(i)
    else:
        pass
        
result_tr = list(zip(*result))

for i in result_tr:## null만 있는 Excel Column 제거
    if checkAllElement(i, None):
        result_tr.remove(i)
    else:
        pass

result = list(zip(*result_tr))

# Assign your output to the OUT variable.
OUT = result