import tkinter as tk
from tkinter import ttk
from tksheet import Sheet
from openpyxl import load_workbook


def load_and_display_excel_with_search(
    parent,
    wm_group_manager,
    logging_text_widget,
    callback_after_load,
):
    """Read the WorkMaster_DB.xlsx file using openpyxl and display it in tksheet with a search feature."""
    try:
        # Load the Excel file using openpyxl
        file_path = "resources/WorkMaster_DB.xlsx"
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Extract data from the active sheet
        data_ = []
        for row in sheet.iter_rows(values_only=True):
            data_.append(list(row))
        data = list(filter(lambda x: any(x) != False, data_))

        # Create a frame for the search box and buttons
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill=tk.X, padx=5, pady=5)

        # Create the search entry box
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var)
        search_entry.pack(side=tk.LEFT, padx=5)

        # Create the search button
        search_button = ttk.Button(
            search_frame,
            text="Search",
            command=lambda: search_excel_data(
                sheet_widget, data, search_var.get(), wm_group_manager, shadow_var.get()
            ),
        )
        search_button.pack(side=tk.LEFT, padx=5)

        # Create a tksheet widget for the Excel data
        sheet_widget = Sheet(parent)
        sheet_widget.pack(fill=tk.BOTH, expand=True)

        # Set headers and data
        headers = [str(cell) for cell in data[4]]  # First row as headers
        sheet_widget.headers(headers)
        sheet_widget.set_sheet_data(data[5:])  # Populate the data, skipping headers

        # Optionally hide specific columns
        sheet_widget.hide_columns([1, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22])

        # Enable sheet functionality
        sheet_widget.disable_bindings()
        sheet_widget.enable_bindings(
            "single_select",  # Allow single cell selection
            "row_select",  # Allow row selection
            "column_select",  # Allow column selection
            "drag_select",  # Allow drag selection
            "column_width_resize",
            "double_click_column_resize",
            "copy",
            "ctrl_click_select",
        )

        # Create a variable to track the shadow feature status
        shadow_var = tk.BooleanVar(value=False)

        # Create a button to toggle shadowing of used WM type
        def toggle_shadow():
            if shadow_var.get():
                highlight_matched_rows(sheet_widget, data[5:], wm_group_manager)
            else:
                reset_highlighting(sheet_widget, data[5:])

        shadow_button = ttk.Checkbutton(
            search_frame,
            text="Shadow used WM type",
            variable=shadow_var,
            command=toggle_shadow,
        )
        shadow_button.pack(side=tk.RIGHT, padx=5)

        # Bind the Enter key to trigger search functionality
        search_entry.bind(
            "<Return>",
            lambda event: search_excel_data(
                sheet_widget,
                data,
                search_var.get(),
                wm_group_manager,
                shadow_var.get(),
            ),
        )

        sheet_widget.update_idletasks()
        return sheet_widget  # Return the sheet widget for use in other functions
    except FileNotFoundError:
        tk.Label(parent, text="WorkMaster_DB.xlsx not found").pack(pady=20)
    except Exception as e:
        tk.Label(parent, text=f"Error: {e}").pack(pady=20)


def highlight_matched_rows(sheet_widget, excel_data, wm_group_manager):
    """Highlight rows in the Excel sheet that match any items in wm_group_match.json."""
    wm_group_data = wm_group_manager.get_wm_group_data()
    for row_index, row in enumerate(excel_data):
        first_column_value = row[0]  # Get the first column value as a string
        if first_column_value in str(wm_group_data):
            sheet_widget.highlight_rows(
                row_index, bg="#e2e2e2"
            )  # Highlight matching rows
        else:
            sheet_widget.highlight_rows(row_index, bg="white")


def reset_highlighting(sheet_widget, excel_data):
    """Reset row highlighting to default (white background)."""
    for row_index, _ in enumerate(excel_data):
        sheet_widget.highlight_rows(row_index, bg="white")


def search_excel_data(
    sheet_widget, original_data, search_text, wm_group_manager, shadow_enabled
):
    """Filter the Excel rows based on the search text and update the sheet."""
    # reset_highlighting(sheet_widget, original_data[5:])
    if not search_text:
        sheet_widget.set_sheet_data(original_data[5:])
        if shadow_enabled:
            highlight_matched_rows(sheet_widget, original_data[5:], wm_group_manager)
    else:
        filtered_data = [
            row
            for row in original_data[5:]
            if any(search_text.lower() in str(cell).lower() for cell in row)
        ]
        sheet_widget.set_sheet_data(filtered_data)
        if shadow_enabled:
            highlight_matched_rows(sheet_widget, filtered_data, wm_group_manager)
