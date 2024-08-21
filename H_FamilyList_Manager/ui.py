# Version: 1.0.12

import time
import tkinter as tk
from tkinter import ttk, font
from search import SearchManager
from openpyxl import load_workbook


def create_single_area_tab(app, name):
    # Create a frame for the tab
    # root = app.root

    main_frame = ttk.Frame(app.notebook)
    main_frame.pack(fill="both", expand=True)

    # tab_frame = ttk.Frame(app.notebook)
    tab_frame = ttk.Frame(main_frame)
    tab_frame.pack(fill="both", expand=True)

    # Frame for buttons at the top
    button_frame = ttk.Frame(tab_frame)
    button_frame.pack(fill="x", padx=10, pady=10)

    # Frame for buttons at the top
    bottom_frame = ttk.Frame(tab_frame)
    bottom_frame.pack(fill="x", padx=10, pady=10)

    # Add Item Button
    add_button = ttk.Button(
        button_frame, text="Add Item", command=app.treeview_operations.add_item
    )
    add_button.pack(side="left", padx=2, pady=5)

    # Remove Selected Button
    remove_button = ttk.Button(
        button_frame,
        text="Remove Selected",
        command=app.treeview_operations.remove_selected_item,
    )
    remove_button.pack(side="left", padx=2, pady=5)

    # Save Button
    save_button = ttk.Button(
        button_frame, text="Save", command=app.config_manager.save_configuration
    )
    save_button.pack(side="left", padx=2, pady=5)

    # Load Button
    load_button = ttk.Button(
        button_frame, text="Load", command=app.config_manager.load_configuration
    )
    load_button.pack(side="left", padx=2, pady=5)

    # Collapse All Button
    collapse_button = ttk.Button(
        button_frame, text="Collapse All", command=app.treeview_operations.collapse_all
    )
    collapse_button.pack(side="left", padx=2, pady=5)

    # Level limit combobox
    level_limit_var = tk.StringVar()  # Corrected to use tk.StringVar() directly
    app.level_limit_var = level_limit_var  # Save to the app instance
    level_limit_combo = ttk.Combobox(
        button_frame, textvariable=level_limit_var, state="readonly", width=10
    )
    level_limit_combo["values"] = [str(i) for i in range(1, 11)]  # Levels 1 to 10
    # level_limit_combo.current(2)  # Default to level 3
    level_limit_combo.current(6)  # Default to level 7
    level_limit_combo.bind(
        "<<ComboboxSelected>>", app.treeview_operations.update_level_limit
    )
    level_limit_combo.pack(side="left", padx=2, pady=5)

    # Search entry and button (added back from v1.4.4)
    app.search_var = tk.StringVar()
    app.search_manager.setup_search(button_frame)  ### mk

    # Create a frame for the treeview and scrollbar
    tree_frame = ttk.Frame(tab_frame)
    tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # Create a vertical scrollbar and associate it with the treeview
    tree_scroll = ttk.Scrollbar(tree_frame)
    tree_scroll.pack(side="right", fill="y")

    # Create and pack the Treeview widget inside this frame
    app.tree = ttk.Treeview(
        tree_frame,
        columns=("indented_name", "description"),
        show="tree headings",
        yscrollcommand=tree_scroll.set,
        style="Treeview",
    )
    app.tree.column("#0", width=50, anchor="center")  # Tree structure column
    app.tree.column("indented_name", width=300, minwidth=200)
    app.tree.column("description", width=400, minwidth=200)
    app.tree.heading("#0", text="#")
    app.tree.heading("indented_name", text="Item")
    app.tree.heading("description", text="Description")
    app.tree.pack(padx=20, pady=10, fill="both", expand=True)

    # Configure the scrollbar
    tree_scroll.config(command=app.tree.yview)

    # Define custom fonts
    app.font_level_0 = font.Font(family="Arial", size=12, weight="bold")
    app.font_level_4 = font.Font(family="Arial", size=11)
    app.default_font = font.Font(family="Arial", size=9)

    # Define tag for top-level items with light green background and bold font
    app.tree.tag_configure("top_level", background="light green", font=app.font_level_0)
    # Define tag for level 4 items with blue foreground and custom font
    app.tree.tag_configure("level_4", foreground="blue", font=app.font_level_4)

    # Add initial top-level items to the tree
    for i, item in enumerate(app.top_level_items):
        app.treeview_operations.add_numbered_item(
            "", f"{i}", item, "", top_level=True, track_undo=False
        )

    # Bind double-click event for editing item names
    app.tree.bind("<Double-1>", app.treeview_operations.on_item_double_click)

    # Bind single-click event for editing descriptions
    app.tree.bind("<Button-1>", app.treeview_operations.on_item_single_click)

    # Add the tab to the notebook
    # app.notebook.add(tab_frame, text=f"   {name}   ")
    app.notebook.add(main_frame, text=f"   {name}   ")

    # Add "Order Adjustment Mode" button
    app.order_adjustment_mode = tk.BooleanVar(value=False)
    # app.order_adjustment_button = ttk.Button(
    #     root,
    #     text="Order Adjustment Mode: OFF",
    #     command=app.drag_and_drop_manager.toggle_order_adjustment_mode,
    # )
    app.order_adjustment_button = ttk.Button(
        # button_frame,
        main_frame,
        text="Order Adjustment Mode: OFF",
        command=app.drag_and_drop_manager.toggle_order_adjustment_mode,
    )
    app.order_adjustment_button.pack(side="bottom", anchor="w", padx=10, pady=5)


# ui.py


def create_wm_matching_by_group_tab(app):
    # Create a frame for the new tab using tk.Frame
    tab_frame = ttk.Frame(app.notebook)

    # Split the frame into left, center, and right areas using tk.Frame
    left_frame = tk.Frame(tab_frame, width=250)
    center_frame = tk.Frame(tab_frame, width=300)
    right_frame = tk.Frame(tab_frame)

    # Configure grid layout for the tab_frame
    tab_frame.grid_columnconfigure(0, weight=0)
    tab_frame.grid_columnconfigure(1, weight=0)
    tab_frame.grid_columnconfigure(2, weight=1)
    tab_frame.grid_rowconfigure(0, weight=1)

    left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
    center_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
    right_frame.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)

    left_frame.grid_rowconfigure(0, weight=1)
    left_frame.grid_columnconfigure(0, weight=1)

    center_frame.grid_rowconfigure(0, weight=0)
    center_frame.grid_rowconfigure(1, weight=0)
    center_frame.grid_rowconfigure(2, weight=1)
    center_frame.grid_columnconfigure(0, weight=1)

    right_frame.grid_rowconfigure(0, weight=1)
    right_frame.grid_columnconfigure(0, weight=1)

    # Create a Treeview to display the unique names without indents in the left area
    app.wm_group_treeview = ttk.Treeview(left_frame, columns=("name",), show="headings")
    app.wm_group_treeview.heading("name", text="Name")
    app.wm_group_treeview.column("name", width=240, anchor="w")
    app.wm_group_treeview.grid(row=0, column=0, sticky="nsew")

    # Create a fixed title label in the center area
    fixed_title_label = ttk.Label(
        center_frame, text="          WM Matching          ", font=("Arial", 16, "bold")
    )
    fixed_title_label.grid(row=0, column=0, sticky="nw", padx=10, pady=5)

    # Create a dynamic title label in the center area
    app.center_title_label = ttk.Label(center_frame, text="", font=("Arial", 14))
    app.center_title_label.grid(row=1, column=0, sticky="nw", padx=10, pady=5)

    # Create a white display area under the title
    app.drop_area = tk.Frame(center_frame, bg="white", relief="sunken", bd=2)
    app.drop_area.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)

    # Create a Frame for Excel Treeview in the right area
    excel_frame = tk.Frame(right_frame)
    excel_frame.grid(row=0, column=0, sticky="nsew")

    excel_frame.grid_rowconfigure(0, weight=1)
    excel_frame.grid_columnconfigure(0, weight=1)

    app.excel_treeview = ttk.Treeview(excel_frame, show="headings")
    app.excel_treeview.grid(row=0, column=0, sticky="nsew")

    scrollbar_x = ttk.Scrollbar(
        excel_frame, orient="horizontal", command=app.excel_treeview.xview
    )
    scrollbar_x.grid(row=1, column=0, sticky="ew")
    app.excel_treeview.configure(xscrollcommand=scrollbar_x.set)

    scrollbar_y = ttk.Scrollbar(
        excel_frame, orient="vertical", command=app.excel_treeview.yview
    )
    scrollbar_y.grid(row=0, column=1, sticky="ns")
    app.excel_treeview.configure(yscrollcommand=scrollbar_y.set)

    display_excel_data_openpyxl(app)

    app.notebook.add(tab_frame, text="WM 그룹별 매칭")

    app.wm_group_treeview.bind("<<TreeviewSelect>>", app.update_center_title_label)

    # Enable drag-and-drop functionality
    enable_drag_and_drop(app)


def enable_drag_and_drop(app):
    # Handle drag start
    def on_drag_start(event, treeview):
        # Delay to ensure the selection process completes
        app.after(50, lambda: initiate_drag(event, treeview))

    def initiate_drag(event, treeview):
        # Clear any previous drag data
        if hasattr(app, "dragged_items"):
            del app.dragged_items

        selected_items = treeview.selection()
        if selected_items:
            print("drag on")  # Debugging: print when drag starts
            app.dragged_items = [(treeview, item) for item in selected_items]
            item_texts = [treeview.item(item, "values")[0] for item in selected_items]

            # Create a label to follow the cursor
            if hasattr(app, "drag_label"):
                app.drag_label.destroy()  # Ensure previous label is destroyed
            app.drag_label = tk.Label(app, text=", ".join(item_texts), bg="yellow")
            app.drag_label.place(x=event.x_root, y=event.y_root)
            app.bind(
                "<B1-Motion>", on_drag_motion
            )  # Bind motion event to follow the cursor
            app.bind("<ButtonRelease-1>", on_drop)  # Bind drop event to the root window

    # Handle dragging motion to move the drag_label
    def on_drag_motion(event):
        if hasattr(app, "drag_label"):
            app.drag_label.place(
                x=event.x_root + 10, y=event.y_root + 10
            )  # Slight offset for visibility

    # Handle drop action
    def on_drop(event):
        print("drag off")  # Debugging: print when drag ends

        if not hasattr(app, "dragged_items"):
            return

        # Check if drop area is under the mouse
        widget_under_mouse = app.winfo_containing(event.x_root, event.y_root)
        if widget_under_mouse == app.drop_area:
            # Display dragged items in the drop area
            for treeview, item in app.dragged_items:
                item_text = treeview.item(item, "values")[
                    0
                ]  # Get the first column value
                label = ttk.Label(
                    app.drop_area, text=item_text, background="white", anchor="w"
                )
                label.pack(anchor="w", padx=5, pady=2)

        # Cleanup: remove drag label
        if hasattr(app, "drag_label"):
            app.drag_label.destroy()
            del app.drag_label

        # Cleanup: clear dragged items
        if hasattr(app, "dragged_items"):
            del app.dragged_items

        # Unbind the motion and drop events
        app.unbind("<B1-Motion>")
        app.unbind("<ButtonRelease-1>")

    # Bind the drag start to both Treeviews
    app.wm_group_treeview.bind(
        "<ButtonPress-1>", lambda e: on_drag_start(e, app.wm_group_treeview)
    )
    app.excel_treeview.bind(
        "<ButtonPress-1>", lambda e: on_drag_start(e, app.excel_treeview)
    )

    # Handle shift-click and control-click for multi-selection
    app.wm_group_treeview.bind("<Shift-ButtonPress-1>", lambda e: None)
    app.wm_group_treeview.bind("<Control-ButtonPress-1>", lambda e: None)


def display_excel_data_openpyxl(app):
    # Load the Excel file using openpyxl
    file_path = "WorkMaster_DB.xlsx"  # Assuming the file is in the project directory
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Set up the Treeview columns
    columns = [cell.value for cell in sheet[1]]  # Assuming the first row is the header
    app.excel_treeview["columns"] = columns
    for col in columns:
        app.excel_treeview.heading(col, text=col)
        app.excel_treeview.column(
            col, anchor="w", width=150
        )  # Set a default width for each column

    # Insert rows into the Treeview
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip rows 2 through 5
        if 2 <= i <= 6:
            continue
        app.excel_treeview.insert("", "end", values=row)


def create_three_area_tab(app, name):
    # Create a frame for the tab
    tab_frame = ttk.Frame(app.notebook)

    # Create left, middle, and right frames
    left_frame = ttk.Frame(tab_frame)
    middle_frame = ttk.Frame(tab_frame)
    right_frame = ttk.Frame(tab_frame)

    # Left content
    left_label = ttk.Label(left_frame, text=f"Left content of {name}")
    left_button = ttk.Button(
        left_frame,
        text="Left Button",
        command=lambda: app.context_menu_manager.on_button_click(
            f"Left Button on {name}"
        ),
    )
    left_text = ttk.Entry(left_frame, width=30)

    # Middle content
    middle_label = ttk.Label(middle_frame, text=f"Middle content of {name}")
    middle_button = ttk.Button(
        middle_frame,
        text="Middle Button",
        command=lambda: app.context_menu_manager.on_button_click(
            f"Middle Button on {name}"
        ),
    )
    middle_text = ttk.Entry(middle_frame, width=30)

    # Right content
    right_label = ttk.Label(right_frame, text=f"Right content of {name}")
    right_button = ttk.Button(
        right_frame,
        text="Right Button",
        command=lambda: app.context_menu_manager.on_button_click(
            f"Right Button on {name}"
        ),
    )
    right_text = ttk.Entry(right_frame, width=30)

    # Pack left content with padding
    left_label.pack(padx=20, pady=10)
    left_button.pack(padx=20, pady=10)
    left_text.pack(padx=20, pady=10)

    # Pack middle content with padding
    middle_label.pack(padx=20, pady=10)
    middle_button.pack(padx=20, pady=10)
    middle_text.pack(padx=20, pady=10)

    # Pack right content with padding
    right_label.pack(padx=20, pady=10)
    right_button.pack(padx=20, pady=10)
    right_text.pack(padx=20, pady=10)

    # Layout left, middle, and right frames
    left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    middle_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
    right_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

    # Add the tab to the notebook with increased padding
    app.notebook.add(tab_frame, text=f"   {name}   ")
