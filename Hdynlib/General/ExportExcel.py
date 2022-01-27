# Load the Python Standard and DesignScript Libraries
import sys
import os
#sys.path.append(r'C:\Users\HEC\AppData\Local\Programs\Python\Python39\Lib\site-packages')
#sys.path.append(r'C:\Users\pjmk0\AppData\Local\Programs\Python\Python39\Lib\site-packages')
sys.path.append(os.getenv('LOCALAPPDATA').replace('\\','\\\\') + r'\Programs\Python\Python39\Lib\site-packages')
import clr
clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *
from collections import Iterable

from itertools import chain
def iterable(obj):
    return isinstance(obj, Iterable)


# The inputs to this node will be stored as a list in the IN variables.
dataEnteringNode = IN

exportBln = IN[0]
oldfilepath = IN[1]
pjtinfo = IN[2][0]
PjtName = pjtinfo[0]
BldgName = pjtinfo[1]
SheetName = IN[2][1]
InData = IN[3]

# Place your code below this line
import datetime
from datetime import date, timedelta, datetime
Today = date.today()
now = datetime.now()

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color, numbers

sample_items = InData

### 기존 워크북 불러오기
#wb = openpyxl.load_workbook(oldfilepath)

def write_to_Newfile(filepath, SheetName):
    def findWMset(InData):
        result = []
        tmp = []
        for i in InData:
            tmp.append(i[1])
        tmp = set(tmp)
        for i in tmp:
            result.append([i])
        return result
    
    WMset = findWMset(sample_items)
    
    wb = openpyxl.Workbook()
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    ws_sum = wb.create_sheet(title= "Summary", index=0)
    
    ws = wb.create_sheet(title= f'{SheetName}')
    
##################ws_sum#########
    ws_sum.merge_cells('A1:L1')
    ws_sum['A1'] = f'Summary Sheet ({PjtName}_{BldgName})'
    cell = ws_sum['A1']
    cell.font = openpyxl.styles.Font(color='000000', size=20)
    cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='fff3b0'))

    ws_sum['A2'] = 'Item Number'
    ws_sum['A2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['B2'] = 'WorkMaster Code'
    ws_sum['B2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['C2'] = 'Description'
    ws_sum['C2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['D2'] = 'Quantities'
    ws_sum['D2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['E2'] = 'Unit'
    ws_sum['E2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['F2'] = 'Unit Price'
    ws_sum['F2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws_sum['G2'] = 'Total Price'
    ws_sum['G2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)


## 칼럼별 너비 조정 구간
#    for col in range(ws.max_column):
#        ws.column_dimensions[chr(ord('A') + col)].width = 15
    ws_sum.column_dimensions['A'].width = 5
    ws_sum.column_dimensions['B'].width = 20
    ws_sum.column_dimensions['C'].width = 75
    ws_sum.column_dimensions['D'].width = 11
    ws_sum.column_dimensions['E'].width = 9.5
    ws_sum.column_dimensions['F'].width = 11
    ws_sum.column_dimensions['G'].width = 11


    start_row = 4
    ws_sum.merge_cells(f'A{start_row-1}:G{start_row-1}')
    ws_sum[f'A{start_row-1}'] = f'Summary for_{SheetName}'
    cell = ws_sum[f'A{start_row-1}']
    cell.font = openpyxl.styles.Font(color='000000', size=12.5)
    cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='aaf2d1'))
    for (row, item) in enumerate(WMset, start_row):
        cell = ws_sum.cell(row=row, column=1, value=row - 3)
        #cell.number_format = openpyxl.styles.numbers.builtin_format_code(0)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        ws_sum.cell(row=row, column=2, value=item[0]) ## WorkMaster Code 칼럼
        
        cell = ws_sum.cell(row=row, column=3, value= f"= VLOOKUP(B{cell.row},'{SheetName}'!$C${start_row}:$F${start_row+len(sample_items)},2,FALSE)".format(current_row=cell.row))
        
        cell = ws_sum.cell(row=row, column=4, value= f"= SUMIF('{SheetName}'!$C${start_row}:$C${start_row+len(sample_items)},B{cell.row},'{SheetName}'!$E${start_row}:$E${start_row+len(sample_items)})".format(current_row=cell.row))
        
        cell = ws_sum.cell(row=row, column=5, value= f"= VLOOKUP(B{cell.row},'{SheetName}'!$C${start_row}:$F${start_row+len(sample_items)},4,FALSE)".format(current_row=cell.row))
        
        
        cell = ws_sum.cell(row=row, column=7, value='= D{current_row}*F{current_row}'.format(current_row=cell.row))
        cell.number_format = '#,##0.00'

    total_row = ws_sum.max_row + 1
    if total_row > start_row:
        cell1 = ws_sum.cell(row=total_row, column=6, value='Total')
        cell1.font = openpyxl.styles.Font(b=True, color='000000', size=15)
        
        cell2 = ws_sum.cell(row=total_row, column=7, value='= SUM(G{}:G{})'.format(start_row, total_row - 1))
        cell2.number_format = openpyxl.styles.numbers.builtin_format_code(3)
        cell2.border = Border(top=double, left=double, right=double, bottom=double)
        cell2.font = openpyxl.styles.Font(b=True, color='ff03ce', size=15)
    
    
    
##################ws#########
    ws.merge_cells('A1:L1')
    ws['A1'] = f'{SheetName} Category Calculation Sheet ({PjtName}_{BldgName})'
    cell = ws['A1']
    cell.font = openpyxl.styles.Font(color='000000', size=20)
    cell.fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='fff3b0'))

    ws['A2'] = 'Item Number'
    ws['A2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['B2'] = 'Family Type'
    ws['B2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['C2'] = 'WorkMaster Code'
    ws['C2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['D2'] = 'Description'
    ws['D2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['E2'] = 'Quantities'
    ws['E2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['F2'] = 'Unit'
    ws['F2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['G2'] = 'Unit Price'
    ws['G2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)
    ws['H2'] = 'Total Price'
    ws['H2'].font = openpyxl.styles.Font(b=True, color='000000', size=12)


## 칼럼별 너비 조정 구간
#    for col in range(ws.max_column):
#        ws.column_dimensions[chr(ord('A') + col)].width = 15
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20



    start_row = 3
    for (row, item) in enumerate(sample_items, start_row):
        cell = ws.cell(row=row, column=1, value=row - 2)
        #cell.number_format = openpyxl.styles.numbers.builtin_format_code(0)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

#        cell = ws.cell(row=row, column=2, value=item[0])
#        cell.number_format = openpyxl.styles.numbers.FORMAT_DATE_YYYYMMDD2  # 'yyyy-mm-dd'
        ws.cell(row=row, column=2, value=item[0]) ## Family Type 칼럼

        ws.cell(row=row, column=3, value=item[1]) ## WorkMaster Code 칼럼
        
        ws.cell(row=row, column=4, value=item[2]) ## WorkMaster Code 칼럼
        
        ws.cell(row=row, column=5, value=item[3]) ## Quantities 칼럼
        cell.number_format = '#,##0.00'
        
        ws.cell(row=row, column=6, value=item[4]) ## Unit 칼럼
        

        cell = ws.cell(row=row, column=8, value='= E{current_row}*G{current_row}'.format(current_row=cell.row))
        cell.number_format = '#,##0.00'

    total_row = ws.max_row + 1
    if total_row > start_row:
        cell1 = ws.cell(row=total_row, column=7, value='Total')
        cell1.font = openpyxl.styles.Font(b=True, color='000000', size=15)
        
        cell2 = ws.cell(row=total_row, column=8, value='= SUM(H{}:H{})'.format(start_row, total_row - 1))
        cell2.number_format = openpyxl.styles.numbers.builtin_format_code(3)
        cell2.border = Border(top=double, left=double, right=double, bottom=double)
        cell2.font = openpyxl.styles.Font(b=True, color='ff03ce', size=15)


    wb.save(filepath)
    wb.close()


##if __name__ == '__main__':
#write_to_file('C:\\Users\\pjmk0\\AppData\\Roaming\\test.xlsx')
if exportBln:
    if not oldfilepath:
        save_name = f'{SheetName} CalcSheet ({PjtName}_{BldgName})_' + now.strftime("%Y-%m-%d_%H%M%S") + '.xlsx'
        path = os.getenv('APPDATA').replace('\\','\\\\')
        filePath = path + '\\' + save_name
        
        write_to_Newfile(filePath, SheetName)
    else:
#        #write_to_Oldfile(filePath)
#        wb = openpyxl.load_workbook(oldfilepath)
#        ws_sum = wb['Summary']
#        
#        ws_sum.append(range(10))
#        wb.save(oldfilepath)
#        wb.close()
        pass
else:
    pass

# Assign your output to the OUT variable.
OUT = os.getenv('APPDATA')