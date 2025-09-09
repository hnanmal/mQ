import os
from tkinter import filedialog
from tkinter import messagebox

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from src.controllers.tree_data_navigator import TreeDataManager_treesheet
from src.core.fp_utils import *
import tkinter as tk
import tkinter.font
from tksheet import Sheet

# from tkinter import ttk
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

from src.views.widget.widget import StateObserver


class pjt_interior_matrix_widget:
    def __init__(self, state, parent):
        self.state = state
        self.data_kind = "project-assigntype"
        self.treeDataManager = TreeDataManager_treesheet(state, self)
        self.state_observer = StateObserver(state, lambda e: self.update(e))

        self.widget_filtermode = False

        # ✅ 대분류(Materials) + 하위 항목(Sub-Materials) 데이터 구조
        self.materials = {}

        # ✅ 룸 목록 (열)
        self.all_rooms = []

        # ✅ 현재 표시할 룸 목록 (필터 적용 가능)
        self.filtered_rooms = self.all_rooms[:]

        self.material_rows = []
        self.category_rows = set()  # ✅ 대분류 행 인덱스를 저장

        self.combobox_area = ttk.Frame(parent)
        self.combobox_area.pack(anchor="nw", fill="x")
        self.sheet_area = ttk.Frame(parent)
        self.sheet_area.pack(anchor="nw", fill="both", expand=True)

        self.select_lv_label = ttk.Label(
            self.combobox_area,
            text="층을 선택해주세요 >>",
        )
        self.select_lv_label.config(foreground="#ff0080")
        self.select_lv_label.pack(side="left", padx=10)
        self.fl_combo = ttk.Combobox(
            self.combobox_area,
            values=[],
            bootstyle="info",
            state="readonly",
        )
        self.fl_combo.pack(side="left")
        # Bind the <<ComboboxSelected>> event to the handler
        self.fl_combo.bind(
            "<<ComboboxSelected>>", lambda event: self.on_combobox_select(event, state)
        )

        # Add project-only filter button
        filter_button_ptOnly = ttk.Button(
            self.combobox_area,
            text="Filter Project Type Only",
            command=self.on_click_filter_btn_pjtOnly,
            bootstyle="info-outline",
        )
        filter_button_ptOnly.pack(padx=10, pady=5, side="left")

        # Add filter button
        filter_button = ttk.Button(
            self.combobox_area,
            text="Filter Unused",
            command=self.on_click_filter_btn,
            bootstyle="info-outline",
        )
        filter_button.pack(padx=10, pady=5, side="left")

        # Add Reset button
        reset_button = ttk.Button(
            self.combobox_area,
            text="Reset SheetView",
            command=self.on_click_reset_btn,
            bootstyle="info-outline",
        )
        reset_button.pack(padx=10, pady=5, side="left")

        self.def_row_height = 15

        # ✅ tksheet 생성 (가로 스크롤 지원)
        self.sheet = Sheet(
            self.sheet_area,
            data=[],
            default_column_width=100,
            # default_header=[],
            # headers=self.filtered_rooms,
            # row_index=self.material_rows,
            align="c",
            index_align="w",
            default_row_height=self.def_row_height,
        )
        self.sheet.set_options(
            header_font=("Arial", 8, "normal"),
        )
        self.sheet.set_index_width(300)
        self.sheet.enable_bindings()
        self.sheet.pack(fill="both", expand=True)

        # Create Extract Excel Button
        export_btn = ttk.Button(
            self.combobox_area,
            text="Export to Excel",
            command=self.export_tksheet_to_excel,
            bootstyle="success-outline",
        )
        export_btn.pack(side="left", padx=100)

        # ✅ 스타일 다시 적용
        self.apply_styles()

        # ✅ 클릭 이벤트 바인딩 (셀 클릭 시 체크박스 토글)
        self.sheet.extra_bindings("begin_edit_cell", self.toggle_checkbox)

        # Right Click pop up - SWM 이동 메뉴 추가
        self.sheet.popup_menu_add_command(
            label="선택항목 SWM 탭 에서 조회",
            func=self.move_toTgtSWM,
        )

    def move_toTgtSWM(self):
        state = self.state
        current_selection = self.sheet.get_currently_selected()
        print(f"current_selection :: {current_selection}")
        print(f"current_selection.row :: {current_selection.row}")
        print(f"self.material_rows :: {self.material_rows}")

        crnt_row = current_selection.row

        for cat_row in sorted(list(self.category_rows)):
            if crnt_row - cat_row > 0:
                tgt_father_row = cat_row

        tgt_grand_name = "00 Room Finish"
        tgt_father_name = self.material_rows[tgt_father_row]
        tgt_item_name = self.material_rows[crnt_row]

        state.main_notebook.select(2)
        state.pjtStd_tab.select(2)

        tgt_tree = state.pjtStdSWM_treeview.treeview.tree
        grand = go(
            tgt_tree.get_children(""),
            filter(lambda x: tgt_tree.item(x, "text") == tgt_grand_name),
            next,
        )
        swm = go(
            tgt_tree.get_children(grand),
            filter(lambda x: tgt_tree.item(x, "text") == tgt_father_name),
            next,
        )
        tgt_item = go(
            tgt_tree.get_children(swm),
            filter(lambda x: tgt_tree.item(x, "text") == tgt_item_name),
            next,
        )
        print(f"tgt_item {tgt_item}")
        tgt_tree.selection_set(tgt_item)
        tgt_tree.focus(tgt_item)
        tgt_tree.see(tgt_item)

    def setup_sheet(self, headers):
        # 헤더 설정

        self.sheet.headers(
            headers,
        )
        # self.setup_column_style()

        # self.sheet.set_sheet_data([])

    def get_level_fromRoomName(self, name):
        res = go(
            name,
            lambda x: x.split("_"),
            lambda x: x[0][0],
        )
        return res

    def set_floor_commbovalues(self):
        state = self.state
        if state.current_building:
            floors = ["WM 할당 확인", "All"] + go(
                self.all_rooms,
                map(self.get_level_fromRoomName),
                set,
                list,
                sorted,
            )
            print(f"[int matrix] floors : {floors}")
        else:
            floors = ["건물선택필요"]

        self.fl_combo.config(values=floors)
        # self.fl_combo.set(floors[0])

    def set_row_idx(self):
        # ✅ 테이블을 위한 행 구성 (대분류 & 하위 항목)
        self.material_rows = []
        self.category_rows = set()  # ✅ 대분류 행 인덱스를 저장
        row_index = 0

        # for category, sub_materials in self.materials.items():
        order_map = {
            "Floor": 0,
            "Skirt": 1,
            "Wall": 2,
            "Ceiling": 3,
        }

        # for category, sub_materials in sorted(self.materials.items()):
        for category, sub_materials in sorted(
            self.materials.items(), key=lambda x: order_map.get(list(x)[0], 4)
        ):
            self.material_rows.append(category)  # ✅ 대분류 (체크 불가)
            self.category_rows.add(row_index)  # ✅ 대분류 행 인덱스 저장
            row_index += 1
            self.material_rows.extend(sorted(sub_materials))  # ✅ 하위 항목 (체크 가능)
            row_index += len(sub_materials)

        self.sheet.row_index(self.material_rows)

    def apply_styles(self):
        # ✅ 폰트 크기 조정 (튜플 형식 수정)
        self.sheet.set_options(font=("Arial", 8, "normal"))
        self.sheet.set_options(default_row_height=self.def_row_height)
        self.sheet.dehighlight_all()
        """✅ 대분류(Materials) 스타일 적용: 글자색 회색, 배경 보라색"""
        for row_idx in self.category_rows:
            self.sheet.highlight_rows(
                row_idx, bg="#D8BFD8", fg="gray"
            )  # ✅ 연한 보라색 배경, 회색 글자 (underline 제거)
        print(f"self.category_rows : {self.category_rows}")

    # Event handler for combo box selection
    def on_combobox_select(self, event, state):
        selected_value = event.widget.get()  # Get the selected value from the combo box
        self.set_floor_commbovalues()

        state.observer_manager.notify_observers(state)

    def save_crnt_scroll(self, event=None):
        state = self.state
        root = state.root

        # 현재 스크롤 위치 저장
        self.y_scroll = self.sheet.yview()
        print(f"yscroll 위치 저장 : {self.y_scroll}")
        self.x_scroll = self.sheet.xview()
        print(f"xscroll 위치 저장 : {self.x_scroll}")

    def restore_scroll(self, event=None):
        state = self.state
        root = state.root
        """셀을 수정할 때 스크롤 위치를 유지"""

        # 수정 후 스크롤 위치 복원 (약간의 지연을 줘야 동작함)
        if avg(*self.y_scroll) < 0.5:
            y_scroll_moveto = min(self.y_scroll)
        else:
            y_scroll_moveto = max(self.y_scroll)

        if avg(*self.x_scroll) < 0.5:
            x_scroll_moveto = min(self.x_scroll)
        else:
            x_scroll_moveto = max(self.x_scroll)

        root.after(10, lambda: self.sheet.set_yview(y_scroll_moveto))
        root.after(10, lambda: self.sheet.set_xview(x_scroll_moveto))

    def on_click_filter_btn(self, event=None):
        self.widget_filtermode = True
        self.category_rows = []
        self.update(event=None, mode="db_update", filter_mode=self.widget_filtermode)
        self.apply_styles()
        self.sheet.update()

    def on_click_filter_btn_pjtOnly(self, event=None):
        self.widget_filtermode = True
        self.category_rows = []
        self.update(
            event=None, mode="db_update_pjtOnly", filter_mode=self.widget_filtermode
        )
        self.apply_styles()
        self.sheet.update()

    def on_click_reset_btn(self, event=None):
        self.widget_filtermode = False
        self.category_rows = []
        self.update(event=None, mode="db_update", filter_mode=self.widget_filtermode)
        self.apply_styles()
        self.sheet.update()

    def update(self, event=None, mode=None, filter_mode=None):
        state = self.state
        self.sheet.set_sheet_data([])

        std_data = state.team_std_info.get("std-familylist")
        pjt_data = state.team_std_info.get(self.data_kind)
        # self.filtered_rooms = []

        rooms_for_selectedBuilding = go(
            pjt_data["children"],
            filter(lambda x: x["values"][1] == state.current_building.get()),
            filter(lambda x: x["values"][-1] == "Top | 0.Room | 0.1"),
            map(lambda x: x["name"]),
            map(lambda x: x.replace("\t", "_")),
            list,
            sorted,
        )
        self.all_rooms = rooms_for_selectedBuilding
        print(f"pjt_data {pjt_data["children"][0]}")
        print(f"rooms_for_selectedBuilding {rooms_for_selectedBuilding}")
        # ✅ 레벨 콤보박스 업데이트
        self.set_floor_commbovalues()

        selected_lv = self.fl_combo.get()

        rooms_for_selectedLevel = go(
            rooms_for_selectedBuilding,
            filter(lambda x: self.get_level_fromRoomName(x) == selected_lv),
            list,
            sorted,
        )

        if selected_lv == "WM 할당 확인":
            self.filtered_rooms = ["WM info"]
        elif selected_lv == "All":
            self.filtered_rooms = rooms_for_selectedBuilding
        else:
            self.filtered_rooms = rooms_for_selectedLevel
        print(f"[int matrix] rooms_for_selectedLevel : {rooms_for_selectedLevel}")

        SWM = go(
            std_data["children"][0]["children"],
            filter(lambda x: x["name"] == "0.Room"),
            lambda x: list(x)[0]["children"],
            filter(lambda x: x["name"] == "0.1"),
            lambda x: list(x)[0]["children"],
            list,
        )

        SWM_names = go(
            SWM,
            map(lambda x: x["name"]),
            list,
        )
        SWMitem_names_ = go(
            SWM,
            map(lambda x: x["children"]),
            map(lambda x: list(map(lambda y: y["name"], x))),
            list,
        )

        if state.current_building.get():
            # SWM item 사용 여부 체크 기준 리스트
            usedChecklist = go(
                pjt_data["children"],
                filter(lambda x: x["values"][1] == state.current_building.get()),
                filter(lambda x: x["values"][-1] == "Top | 0.Room | 0.1"),
                map(lambda x: x.get("children", ["", "", ""])),
                lambda x: chain(*x),
                list,
                filter(lambda x: x),
                map(lambda x: x[2]),
                map(lambda x: x.split(" | ")[-1]),
                list,
            )
            pjtOnly_Checklist = go(
                SWMitem_names_,
                map(lambda x: list(filter(lambda y: "::" in y, x))),
                lambda x: chain(*x),
                list,
            )
            # print(f"SWMitem_names_ : {SWMitem_names_}")
            # print(f"pjtOnly_Checklist : {pjtOnly_Checklist}")

        # floor_usedChecklist, skirt_usedChecklist 등 분리해서 filter_mode 적용
        if filter_mode and mode == "db_update":
            SWMitem_names = go(
                SWMitem_names_,
                map(lambda x: list(filter(lambda i: i in usedChecklist, x))),
                list,
            )
            # print(f"SWMitem_names - {SWMitem_names}")
        elif filter_mode and mode == "db_update_pjtOnly":
            SWMitem_names = go(
                SWMitem_names_,
                map(lambda x: list(filter(lambda i: i in pjtOnly_Checklist, x))),
                list,
            )
        else:
            SWMitem_names = SWMitem_names_

        self.materials = dict(zip(SWM_names, SWMitem_names))
        self.set_row_idx()

        # print(f"[int matrix] self.materials : {self.materials}")

        if mode == "db_update" or "db_update_pjtOnly":
            self.apply_styles()
        else:
            self.sheet.pack_forget()
            self.sheet = Sheet(
                self.sheet_area,
                data=[],
                default_column_width=100,
                headers=self.filtered_rooms,
                row_index=self.material_rows,
                align="c",
                index_align="w",
            )
            self.sheet.enable_bindings()
            self.sheet.pack(fill="both", expand=True)
            # ✅ 스타일 다시 적용
            self.apply_styles()

        # ✅ 클릭 이벤트 바인딩 (셀 클릭 시 체크박스 토글)
        self.sheet.extra_bindings("begin_edit_cell", self.toggle_checkbox)
        self.sheet.extra_bindings("end_paste", self.toggle_forPaste)

        # Right Click pop up - SWM 이동 메뉴 추가
        self.sheet.popup_menu_add_command(
            label="선택항목 SWM 탭 에서 조회",
            func=self.move_toTgtSWM,
        )

        # ✅ 테이블을 위한 행 구성 (대분류 & 하위 항목)
        self.set_row_idx()
        self.sheet.set_index_width(300)

        # ✅ DB 체크 상태 반영
        def get_db_status_forRoom(room, row_idx):
            col = self.filtered_rooms.index(room)

            roomName_asDBformat = room.replace("_", "\t")

            tgt_room_db = go(
                state.team_std_info[self.data_kind]["children"],
                filter(lambda x: x["name"] == roomName_asDBformat),
                next,
                lambda x: x["children"],
            )
            tgt_room_material_names = go(
                tgt_room_db,
                map(lambda x: x[2]),
                list,
            )
            SWM_names = [
                "Floor",
                "Skirt",
                "Wall",
                "Ceiling",
            ]

            SWM_idxes = [self.material_rows.index(x) for x in SWM_names]

            row_dist = [x - row_idx for x in SWM_idxes]
            min_row_dist = go(
                row_dist,
                filter(lambda x: x < 0),
                max,
            )
            parent_SWM_idx = row_dist.index(min_row_dist)
            SWM_name = SWM_names[parent_SWM_idx]

            material_name = self.material_rows[row_idx]  # ✅ 원래 데이터 가져오기
            material_full_name = " | ".join([SWM_name, material_name])

            res = material_full_name in tgt_room_material_names
            return res

        swmDB = state.team_std_info["project-SWM"]
        new_data = []
        crnt_father = None
        for row_idx, row_name in enumerate(self.material_rows):
            if row_idx in self.category_rows:
                new_data.append(
                    [""] * len(self.filtered_rooms)
                )  # ✅ 대분류 행은 체크 없음
                crnt_father = row_name
                print(f"crnt_father {crnt_father}")
            else:
                if selected_lv == "WM 할당 확인":
                    swm_key_name = " | ".join(
                        [
                            "00 Room Finish",
                            crnt_father,
                            row_name,
                        ]
                    )
                    row_data = [
                        (
                            " | ".join(swmDB.get(swm_key_name))
                            if swmDB.get(swm_key_name)
                            else ""
                        )
                    ]
                else:
                    row_data = [
                        "✅" if get_db_status_forRoom(room, row_idx) else ""
                        for room in self.filtered_rooms
                    ]

                new_data.append(row_data)
        # print(f"new_data {new_data}")

        self.sheet.set_header_data(self.filtered_rooms)
        self.sheet.set_sheet_data(new_data)  # ✅ 데이터 업데이트
        self.sheet.set_options(
            header_font=("Arial", 8, "normal"),
            index_font=("Arial", 8, "normal"),
        )
        self.sheet.set_all_cell_sizes_to_text()

        # ✅ 스타일 다시 적용
        self.apply_styles()
        # self.sheet.redraw()  # ✅ 화면 갱신

    def delete_assign_forRoom(self, roomName, material_full_name):
        state = self.state
        # print(f"[delete_assign_forRoom] roomName {roomName}")
        # print(f"[delete_assign_forRoom] deleting start")
        db_data = state.team_std_info[self.data_kind]["children"]

        for i in db_data:
            if i["name"] == roomName and i["values"][1] == state.current_building.get():
                for idx, c in enumerate(i["children"]):
                    # print(f"[delete_assign_forRoom] deleting stt {c}")
                    if c[2] == material_full_name:
                        res = i["children"].pop(idx)
                        # print(f"[delete_assign_forRoom] deleting end {res}")

    # def go_toSWMtab(self):
    #     state = self.state
    #     state.main_notebook.select(state.pjtStd_tab)
    #     state.pjtStd_tab.select(state.s_wm_tab)

    def add_assign_forRoom(self, roomName, material_full_name):
        state = self.state
        pjt_SWM_key = " | ".join(["00 Room Finish", material_full_name])
        # print(f"[add_assign_forRoom] pjt_SWM_key {pjt_SWM_key}")
        pjt_swm_data = state.team_std_info.get("project-SWM")
        tgt_wm_data = pjt_swm_data.get(pjt_SWM_key)
        if tgt_wm_data:
            tgt_wm_code = tgt_wm_data[0]
            tgt_wm_gauge = tgt_wm_data[1]
            tgt_wm_unit = tgt_wm_data[2]
            tgt_wm_spec = tgt_wm_data[3]
        else:
            tab_move = messagebox.askyesno(
                "add_assign_forRoom",
                f"[ {pjt_SWM_key} ] SWM 항목이 wm 지정이 되지 않았습니다.\nProject Single Work Master 탭에서 해당항목의 wm을 지정하시겠습니까?",
            )
            print(f"tab_move {tab_move}")
            if tab_move:
                self.move_toTgtSWM()
            return False

        WMs = state.team_std_info.get("WMs")

        WMsStr = go(
            WMs,
            map(lambda x: list(map(str, x))),
            map(lambda x: filter(lambda x: x != "0", x)),
            map(lambda x: filter(lambda x: x != "", x)),
            map(lambda x: filter(lambda x: x != " ", x)),
            map(lambda x: filter(lambda x: x != "ㅤ", x)),  #  공백 특수 문자
            map(lambda x: " | ".join(x)),
            list,
        )

        tgt_wm_str_list = go(
            WMsStr,
            filter(lambda x: tgt_wm_code in x),
            list,
        )
        try:
            tgt_wm_str = tgt_wm_str_list[0]
        except:
            tgt_wm_str = ""

        std_famlist_room_item = go(
            state.team_std_info["std-familylist"]["children"],
            list,
            lambda x: x[0]["children"],
            filter(lambda x: x["name"] == "0.Room"),
            list,
            lambda x: x[0]["children"],
            filter(lambda x: x["name"] == "0.1"),
            list,
        )

        tgt_SWM_name, tgt_item_name = material_full_name.split(" | ")
        # print(f"[int matrix] tgt_item_name {tgt_item_name}")

        tgt_item_ = go(
            std_famlist_room_item,
            lambda x: x[0]["children"],
            filter(lambda x: x["name"] == tgt_SWM_name),
            list,
            lambda x: x[0]["children"],
            filter(lambda x: x["name"] == tgt_item_name),
            # list,
            # lambda x: x[0]["values"][6],
            next,
        )
        try:
            tgt_std_formula = go(
                tgt_item_,
                lambda x: x["values"],
                filter(lambda x: x.startswith("=")),
                next,
            )
        except:
            tgt_std_formula = ""
        # print(f"[int matrix] std_famlist_room_item {std_famlist_room_item}")
        # print(f"[int matrix] tgt_std_formula {tgt_std_formula}")

        tgt_calc_type = go(
            std_famlist_room_item[0],
            lambda x: x["values"][-1],
        )

        add_data_form = [
            "SWM",
            "Room Name",
            material_full_name,
            tgt_wm_str,
            tgt_wm_gauge,
            tgt_wm_spec,
            tgt_std_formula,  # 수식,
            tgt_wm_unit,
            tgt_calc_type,  # 산출 유형,
            # "",  # 노트
        ]

        db_data = state.team_std_info[self.data_kind]["children"]

        for i in db_data:
            if i["name"] == roomName and i["values"][1] == state.current_building.get():
                i["children"].append(add_data_form)
                i["children"].sort()

        # print(f"[add_assign_forRoom] add_data_form {add_data_form}")
        return True

    def toggle_forPaste(self, event):
        state = self.state
        print(f"toggle_forPaste {event}")

        box_stt = list(event["selection_boxes"].keys())
        # print(f"toggle_forPaste : box_stt {box_stt}")
        # print(f"toggle_forPaste : box_sttType {type(box_stt[0])}")
        # print(f"toggle_forPaste : box_sttType {dir(box_stt[0])}")
        stt_row = box_stt[0].from_r
        stt_col = box_stt[0].from_c
        paste_data_src = event["data"]
        # print(f"toggle_forPaste : paste_data_src {paste_data_src}")

        oneCol_Cellcount = len(self.materials.keys()) + len(
            list(chain(*self.materials.values()))
        )
        # print(f"col copy event {event}")
        # print(f"oneCol_Cellcount {oneCol_Cellcount}")
        # print(f"len(paste_data_src) {len(paste_data_src)}")
        from_c = go(
            event["selection_boxes"],
            lambda x: x.items(),
            list,
            lambda x: list(x[0])[0],
            lambda x: x.from_c,
        )
        upto_c = go(
            event["selection_boxes"],
            lambda x: x.items(),
            list,
            lambda x: list(x[0])[0],
            lambda x: x.upto_c,
        )
        # print(f"from_c!! {from_c}")
        # print(f"upto_c!! {upto_c}")
        if upto_c - from_c > 1:
            messagebox.showinfo(
                "toggle_forPaste",
                "열 복사 기능은 열에서만 작동합니다.(1번에 1열씩)",
            )
            self.update(
                event=None,
                mode="db_update",
                filter_mode=self.widget_filtermode,
            )
            return

        SWM_names = [
            "Floor",
            "Skirt",
            "Wall",
            "Ceiling",
        ]

        SWM_idxes = [self.material_rows.index(x) for x in SWM_names]

        room = self.filtered_rooms[stt_col]
        roomName_asDBformat = room.replace("_", "\t")

        db_data = state.team_std_info[self.data_kind]["children"]
        pjt_swm_data = state.team_std_info.get("project-SWM")

        chkd_material_names = []
        chkd_material_full_names_effBln = []
        for _row, v in enumerate(paste_data_src):
            # print(f"_row {_row}")
            row = _row + stt_row
            # print(f"row로우 {row}")
            # print(f"v \ {v}")
            try:
                row_dist = [x - row for x in SWM_idxes]
                min_row_dist = go(
                    row_dist,
                    filter(lambda x: x < 0),
                    max,
                )
                parent_SWM_idx = row_dist.index(min_row_dist)

                SWM_name = SWM_names[parent_SWM_idx]
                material_name = self.material_rows[row]  # ✅ 원래 데이터 가져오기
                material_full_name = " | ".join([SWM_name, material_name])

                if v == ["✅"]:
                    chkd_material_names.append(material_name)
                    pjt_SWM_key = " | ".join(["00 Room Finish", material_full_name])
                    chkd_material_full_names_effBln.append(
                        pjt_swm_data.get(pjt_SWM_key)
                    )
            except:
                material_full_name = "기준열"
        # print(f"chkd_material_names : {chkd_material_names}")
        # print(f"chkd_material_full_names_effBln : {chkd_material_full_names_effBln}")

        if not all(chkd_material_full_names_effBln):
            messagebox.showinfo(
                "toggle_forPaste",
                "열 복사 기능은 해당 열의 체크된 모든 항목에 wm 지정이 되었을때만 가능합니다.",
            )
            self.update(
                event=None,
                mode="db_update",
                filter_mode=self.widget_filtermode,
            )
            return

        for _row, v in enumerate(paste_data_src):
            # print(f"_row {_row}")
            row = _row + stt_row
            # print(f"row로우 {row}")
            try:
                row_dist = [x - row for x in SWM_idxes]
                min_row_dist = go(
                    row_dist,
                    filter(lambda x: x < 0),
                    max,
                )
                parent_SWM_idx = row_dist.index(min_row_dist)

                SWM_name = SWM_names[parent_SWM_idx]
                material_name = self.material_rows[row]  # ✅ 원래 데이터 가져오기
                material_full_name = " | ".join([SWM_name, material_name])
            except:
                material_full_name = "기준열"
            # print(f"material_full_name!_!  {material_full_name}")
            # print(self.sheet.get_cell_data(row, stt_col))

            if v == ["✅"] and material_full_name != "기준열":
                # add
                add_bln = None
                for i in db_data:
                    if (
                        i["name"] == roomName_asDBformat
                        and i["values"][1] == state.current_building.get()
                    ):
                        for c in i["children"]:
                            if c[2] == material_full_name:
                                add_bln = False
                if not add_bln:
                    try:
                        pjt_SWM_key = " | ".join(["00 Room Finish", material_full_name])
                        self.add_assign_forRoom(roomName_asDBformat, material_full_name)
                        # print(f"{row}행 데이터 {v} 추가 완료")
                    except:
                        # print(f"{row}행 데이터 {v} 추가 실패")
                        messagebox.showinfo(
                            "toggle_forPaste",
                            f"{row}행 데이터 {v} 추가 실패",
                        )
                else:
                    print(f"{row}행 데이터 add_bln 대상 아님")
            else:
                # del
                try:
                    self.delete_assign_forRoom(roomName_asDBformat, material_full_name)
                except:
                    pass

        self.update(event=None, mode="db_update", filter_mode=self.widget_filtermode)

    def toggle_checkbox(self, event, r_c=None):
        """✅ 셀 클릭 시 체크박스를 토글하는 함수"""
        state = self.state

        self.save_crnt_scroll()

        selected_cells = list(
            self.sheet.get_selected_cells()
        )  # ✅ 'set'을 리스트로 변환
        if not selected_cells:
            return

        if r_c:
            row, col = r_c
        else:
            row, col = selected_cells[0]
        # print(f"[int matrix] self.material_rows : {self.material_rows}")

        SWM_names = [
            "Floor",
            "Skirt",
            "Wall",
            "Ceiling",
        ]

        SWM_idxes = [self.material_rows.index(x) for x in SWM_names]

        row_dist = [x - row for x in SWM_idxes]
        min_row_dist = go(
            row_dist,
            filter(lambda x: x < 0),
            max,
        )
        parent_SWM_idx = row_dist.index(min_row_dist)

        SWM_name = SWM_names[parent_SWM_idx]
        material_name = self.material_rows[row]  # ✅ 원래 데이터 가져오기
        material_full_name = " | ".join([SWM_name, material_name])
        room = self.filtered_rooms[col]
        roomName_asDBformat = room.replace("_", "\t")

        # ✅ 대분류(바닥, 천장, 걸레받이, 벽)는 체크 불가능
        if row in self.category_rows:
            return

        # ✅ 현재 상태 반전 (체크/해제)
        tgt_room_db = go(
            state.team_std_info[self.data_kind]["children"],
            filter(lambda x: x["name"] == roomName_asDBformat),
            next,
            lambda x: x["children"],
        )

        tgt_room_material_names = go(
            tgt_room_db,
            map(lambda x: x[2]),
            list,
        )

        # print(f"[int matrix] tgt_room_material_names : {tgt_room_material_names}")
        # print(f"[int matrix] material_name : {material_name}")

        # DB 상태에 따라 체크하는 코드를 update로 옮기고,-> 완료
        # 여기서는 더블 클릭에 따라 DB를 조작하도록
        if material_full_name in tgt_room_material_names:
            # DB에서 해당 항목 삭제
            self.delete_assign_forRoom(roomName_asDBformat, material_full_name)
        else:
            # DB에 해당 항목 추가
            self.add_assign_forRoom(roomName_asDBformat, material_full_name)

        self.update(event=None, mode="db_update", filter_mode=self.widget_filtermode)
        # self.sheet.redraw()

        self.restore_scroll()
        # state.observer_manager.notify_observers(state)

    def export_tksheet_to_excel(self):
        """tksheet 위젯 데이터를 엑셀 파일로 저장하며 서식도 적용"""
        state = self.state
        sheet = self.sheet

        brief_pjtName = state.team_std_info.get("project-info").get("abbr")
        brief_buildingName = go(
            state.current_building.get(),
            lambda x: x.split(" "),
            map(lambda x: x[0]),
            lambda x: "".join(x),
        )

        # 파일 저장 경로 선택
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save as Excel",
            initialfile=f"{brief_pjtName}_{brief_buildingName}-Interior Matrix",
        )

        if not file_path:  # 사용자가 취소한 경우
            return

        # 새로운 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.title = f"{brief_buildingName}-Interior Matrix"
        # 틀 고정 추가 (Index 열과 Header 행을 동시에 고정)
        ws.freeze_panes = "B3"  # ✅ A열과 1행 고정

        purple_fill = PatternFill(
            # start_color="D9B3FF", end_color="D9B3FF", fill_type="solid"
            start_color="D8BFD8",
            end_color="D8BFD8",
            fill_type="solid",
        )
        special_index_keywords = {"Floor", "Skirt", "Wall", "Ceiling"}
        thin_border = Border(
            right=Side(style="thin", color="CCCCCC")
        )  # 오른쪽 세로선만

        # 헤더 가져오기
        headers = sheet.headers()
        max_header_length = 0  # 헤더 중 가장 긴 텍스트 길이 저장용

        # 행 인덱스 라벨 가져오기
        row_indices = sheet.row_index()

        # 0. 엑셀 첫 번째 행에 타이틀 입력
        ws.cell(
            row=1,
            column=1,
            value=f"{brief_pjtName} - {state.current_building.get()} Interior Matrix",
        )
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(headers) + 1
        )
        ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.row_dimensions[1].height = 20

        # 1. 헤더 (2행부터)
        header_row = 2
        ws.cell(row=header_row, column=1, value="SWM for Rooms")  # 인덱스 열
        ws.cell(row=header_row, column=1).border = (
            thin_border  # ✅ 인덱스 열도 세로선 추가
        )
        # for col_idx, header in enumerate(headers, start=1):
        for col_idx, header in enumerate(headers, start=2):
            max_header_length = max(max_header_length, len(str(header)))  # 길이 체크
            cell = ws.cell(row=2, column=col_idx, value=header)

            # 헤더 스타일 (Bold, 가운데 정렬)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom",
                textRotation=90,  # ✅ 90도 회전
                wrap_text=True,
            )

            # 헤더 배경색 (회색)
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )
            cell.border = thin_border  # ✅ 세로 구분선 추가

        # 헤더 행 높이 조정 (길이에 따라 적당히 계산, 약간 여유있게)
        ws.row_dimensions[2].height = max(150, max_header_length * 4)

        # 2. tksheet 데이터는 3행부터
        data = sheet.get_sheet_data()
        column_widths = sheet.get_column_widths()

        # tksheet 셀 스타일 가져오기 (배경색, 글꼴, 정렬)
        # for row_idx, row_data in enumerate(data, start=2):
        for row_idx, (row_data, row_label) in enumerate(
            zip(data, row_indices), start=3
        ):
            # 인덱스 열 작성
            index_cell = ws.cell(row=row_idx, column=1, value=row_label)
            index_cell.font = Font(name="Calibri", size=9, bold=True)
            index_cell.alignment = Alignment(horizontal="left", vertical="center")
            index_cell.border = thin_border  # ✅ 세로 구분선 추가

            is_special_index = row_label in special_index_keywords

            if is_special_index:
                index_cell.fill = purple_fill
            #####

            # 데이터 셀 작성
            for col_idx, cell_value in enumerate(row_data, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # 정렬 설정
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                # 기본 폰트 설정 (Calibri, Bold)
                cell.font = Font(name="Calibri", size=7, bold=False)

                cell.border = thin_border  # ✅ 오른쪽 경계선 추가

                if is_special_index:
                    cell.fill = purple_fill

        # 컬럼 너비 조정
        # print(f"열 너비 들 {column_widths}")
        # ws.column_dimensions["A"].width = column_widths[0] / 3
        ws.column_dimensions["A"].width = 30
        for col_idx, width in enumerate(column_widths, start=2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 3.3

        #### Transpose 시트 생성 ####
        ws_transposed = wb.create_sheet(title=f"{brief_buildingName}-Transposed Matrix")
        ws_transposed.freeze_panes = "B3"  # ✅ A열과 1행 고정

        # 1행: 타이틀
        title_transposed = f"{brief_pjtName} - {state.current_building.get()} Interior Matrix (Transposed)"
        ws_transposed.cell(row=1, column=1, value=title_transposed)
        ws_transposed.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(data) + 1
        )
        ws_transposed.cell(row=1, column=1).font = Font(
            name="Calibri", size=14, bold=True
        )
        ws_transposed.cell(row=1, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws_transposed.row_dimensions[1].height = 20

        # 2행: 인덱스 헤더 (수직 텍스트 적용!)
        ws_transposed.cell(row=2, column=1, value="Rooms")
        ws_transposed.cell(row=2, column=1).font = Font(
            name="Calibri", size=10, bold=True
        )
        ws_transposed.cell(row=2, column=1).alignment = Alignment(
            horizontal="left", vertical="bottom"
        )
        ws_transposed.cell(row=2, column=1).border = thin_border

        for idx, col in enumerate(sheet.get_sheet_data(), start=2):
            ws_transposed.cell(row=2, column=idx).fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

        highlighted_columns = set()
        for col, row_label in enumerate(row_indices, start=2):
            cell = ws_transposed.cell(row=2, column=col, value=row_label)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="bottom",
                textRotation=90,  # ✅ 헤더를 90도 회전시킴
                wrap_text=False,
            )
            cell.border = thin_border

            if row_label in special_index_keywords:
                cell.fill = purple_fill
                highlighted_columns.add(col)

        # 헤더 행 높이 조정 (길이에 따라 적당히 계산, 약간 여유있게)
        ws_transposed.row_dimensions[2].height = max(150, max_header_length * 4)

        # # 2행 높이도 늘리기
        # ws_transposed.row_dimensions[2].height = max(
        #     40, max(len(r) for r in row_indices) * 3.5
        # )

        # 3행부터: 헤더와 데이터
        for row_idx, header in enumerate(headers, start=3):

            header_cell = ws_transposed.cell(row=row_idx, column=1, value=header)
            header_cell.font = Font(name="Calibri", size=10, bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = thin_border

            if row_idx > 2:
                ws_transposed.cell(row=row_idx, column=1).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )

            for col_idx, row_data in enumerate(data, start=2):
                cell_value = data[col_idx - 2][row_idx - 3]  # Transpose
                cell = ws_transposed.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = Font(name="Calibri", size=7)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = thin_border

                # 💡 강조된 열이면 배경색 칠함
                if col_idx in highlighted_columns:
                    cell.fill = purple_fill

        # 컬럼 너비 조정
        # print(f"열 너비 들 {column_widths}")
        column_tr = sheet.get_sheet_data()
        # ws_transposed.column_dimensions["A"].width = column_widths[0] / 3
        ws_transposed.column_dimensions["A"].width = 30
        for col_idx, width in enumerate(column_tr, start=2):
            ws_transposed.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = 3.3

        #### Transpose 시트 생성 ####

        # 엑셀 파일 저장
        wb.save(file_path)
        print(f"엑셀 파일 저장 완료: {file_path}")

        # 6. 메시지 박스 띄우기
        open_file = messagebox.askyesno(
            "저장 완료", "엑셀 파일 저장이 완료되었습니다.\n파일을 여시겠습니까?"
        )

        # 7. 사용자가 "예(Yes)"를 선택한 경우 엑셀 파일 열기
        if open_file:
            try:
                os.startfile(file_path)  # Windows에서 엑셀 파일 열기
                # subprocess.Popen(["start", file_path])
            except Exception as e:
                messagebox.showerror("오류", f"파일을 열 수 없습니다.\n{e}")
