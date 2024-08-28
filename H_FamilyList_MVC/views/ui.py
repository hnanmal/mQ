# import time
import json
import tkinter as tk
from tkinter import ttk, font
from search import SearchManager
from openpyxl import load_workbook

from models.wm_group import (
    load_wm_group_match_data,
    save_lock_status_to_json,
    save_current_matching_to_json,
    filter_excel_data,
    save_configuration,
    load_configuration,
)
from controllers.main_controller import (
    on_item_double_click,
    on_item_single_click,
    update_center_title_label,
)
from controllers.ui_controller import (
    add_item_to_center,
    remove_item_from_center,
    toggle_lock,
)


class UIManager:
    def __init__(self, app):
        self.app = app

    def create_single_area_tab(self, app, name):
        # Create a frame for the tab
        main_frame = ttk.Frame(app.notebook)
        main_frame.pack(fill="both", expand=True)

        # tab_frame = ttk.Frame(app.notebook)
        tab_frame = ttk.Frame(main_frame)
        tab_frame.pack(fill="both", expand=True)

        # Frame for buttons at the top
        button_frame = ttk.Frame(tab_frame)
        button_frame.pack(fill="x", padx=10, pady=10)

        # Frame for buttons at the top
        bottom_frame = ttk.Frame(tab_frame)
        bottom_frame.pack(fill="x", padx=10, pady=10)

        # Add Item Button
        add_button = ttk.Button(
            button_frame, text="Add Item", command=app.treeview_operations.add_item
        )
        add_button.pack(side="left", padx=2, pady=5)

        # Remove Selected Button
        remove_button = ttk.Button(
            button_frame,
            text="Remove Selected",
            command=app.treeview_operations.remove_selected_item,
        )
        remove_button.pack(side="left", padx=2, pady=5)

        # Save Button
        save_button = ttk.Button(button_frame, text="Save", command=save_configuration)
        save_button.pack(side="left", padx=2, pady=5)

        # Load Button
        load_button = ttk.Button(button_frame, text="Load", command=load_configuration)
        load_button.pack(side="left", padx=2, pady=5)

        # Collapse All Button
        collapse_button = ttk.Button(
            button_frame,
            text="Collapse All",
            command=app.treeview_operations.collapse_all,
        )
        collapse_button.pack(side="left", padx=2, pady=5)

        # Level limit combobox
        level_limit_var = tk.StringVar()  # Corrected to use tk.StringVar() directly
        app.level_limit_var = level_limit_var  # Save to the app instance
        level_limit_combo = ttk.Combobox(
            button_frame, textvariable=level_limit_var, state="readonly", width=10
        )
        level_limit_combo["values"] = [str(i) for i in range(1, 11)]  # Levels 1 to 10
        # level_limit_combo.current(2)  # Default to level 3
        level_limit_combo.current(6)  # Default to level 7
        level_limit_combo.bind(
            "<<ComboboxSelected>>", app.treeview_operations.update_level_limit
        )
        level_limit_combo.pack(side="left", padx=2, pady=5)

        # Search entry and button (added back from v1.4.4)
        app.search_var = tk.StringVar()
        app.search_manager.setup_search(button_frame)  ### mk

        # Create a frame for the treeview and scrollbar
        tree_frame = ttk.Frame(tab_frame)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Create a vertical scrollbar and associate it with the treeview
        tree_scroll = ttk.Scrollbar(tree_frame)
        tree_scroll.pack(side="right", fill="y")

        # Create and pack the Treeview widget inside this frame
        app.tree = ttk.Treeview(
            tree_frame,
            columns=("indented_name", "description"),
            show="tree headings",
            yscrollcommand=tree_scroll.set,
            style="Treeview",
        )
        app.tree.column("#0", width=50, anchor="center")  # Tree structure column
        app.tree.column("indented_name", width=300, minwidth=200)
        app.tree.column("description", width=400, minwidth=200)
        app.tree.heading("#0", text="#")
        app.tree.heading("indented_name", text="Item")
        app.tree.heading("description", text="Description")
        app.tree.pack(padx=20, pady=10, fill="both", expand=True)

        # Configure the scrollbar
        tree_scroll.config(command=app.tree.yview)

        # Define custom fonts
        app.font_level_0 = font.Font(family="Arial", size=12, weight="bold")
        app.font_level_4 = font.Font(family="Arial", size=11)
        app.default_font = font.Font(family="Arial", size=9)

        # Define tag for top-level items with light green background and bold font
        app.tree.tag_configure(
            "top_level", background="light green", font=app.font_level_0
        )
        # Define tag for level 4 items with blue foreground and custom font
        app.tree.tag_configure("level_4", foreground="blue", font=app.font_level_4)

        # Add initial top-level items to the tree
        for i, item in enumerate(app.top_level_items):
            app.treeview_operations.add_numbered_item(
                "", f"{i}", item, "", top_level=True, track_undo=False
            )

        # Bind double-click event for editing item names
        app.tree.bind("<Double-1>", lambda event: on_item_double_click(app, event))

        # Bind single-click event for editing descriptions
        app.tree.bind("<Button-1>", lambda event: on_item_single_click(app, event))

        # Add the tab to the notebook
        # app.notebook.add(tab_frame, text=f"   {name}   ")
        app.notebook.add(main_frame, text=f"   {name}   ")

        # Add "Order Adjustment Mode" button
        app.order_adjustment_mode = tk.BooleanVar(value=False)
        app.order_adjustment_button = ttk.Button(
            # button_frame,
            main_frame,
            text="Order Adjustment Mode: OFF",
            command=app.drag_and_drop_manager.toggle_order_adjustment_mode,
        )
        app.order_adjustment_button.pack(side="bottom", anchor="w", padx=10, pady=5)

    # ui.py

    def create_wm_matching_by_group_tab(self, app):
        # Create a frame for the new tab using tk.Frame
        tab_frame = ttk.Frame(app.notebook)

        # Split the frame into left, center, and right areas using tk.Frame
        left_frame = tk.Frame(tab_frame, width=250)
        center_frame = tk.Frame(tab_frame, width=300)
        right_frame = tk.Frame(tab_frame)
        center_button_frame = tk.Frame(center_frame, width=50)  # Frame for the buttons

        # Configure grid layout for the tab_frame
        tab_frame.grid_columnconfigure(0, weight=0)
        tab_frame.grid_columnconfigure(1, weight=0)
        tab_frame.grid_columnconfigure(2, weight=1)
        tab_frame.grid_rowconfigure(0, weight=1)

        left_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        center_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        center_button_frame.grid(
            row=2, column=1, sticky="nsew", padx=5, pady=5
        )  # Place the button frame
        right_frame.grid(row=0, column=2, sticky="nsew", padx=5, pady=5)

        left_frame.grid_rowconfigure(0, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)

        center_frame.grid_rowconfigure(0, weight=0)
        center_frame.grid_rowconfigure(1, weight=0)
        center_frame.grid_rowconfigure(2, weight=1)
        center_frame.grid_columnconfigure(0, weight=1)

        center_button_frame.grid_rowconfigure(
            0, weight=0
        )  # Align with the top of the area
        center_button_frame.grid_columnconfigure(0, weight=1)

        right_frame.grid_rowconfigure(0, weight=0)  # Adjusted to make space for search
        right_frame.grid_rowconfigure(1, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        # Create a Treeview to display the unique names without indents in the left area
        app.wm_group_treeview = ttk.Treeview(
            left_frame, columns=("name",), show="headings"
        )
        app.wm_group_treeview.heading("name", text="Name")
        app.wm_group_treeview.column("name", width=240, anchor="w")
        app.wm_group_treeview.grid(row=0, column=0, sticky="nsew")

        # Add a vertical scrollbar to the left area
        left_scrollbar = ttk.Scrollbar(
            left_frame, orient="vertical", command=app.wm_group_treeview.yview
        )
        app.wm_group_treeview.configure(yscrollcommand=left_scrollbar.set)
        left_scrollbar.grid(row=0, column=1, sticky="ns")

        # Create a fixed title label in the center area
        fixed_title_label = ttk.Label(
            center_frame,
            text="          WM Matching          ",
            font=("Arial", 16, "bold"),
        )
        fixed_title_label.grid(row=0, column=0, sticky="nw", padx=10, pady=5)

        # Create a dynamic title label in the center area
        app.center_title_label = ttk.Label(center_frame, text="", font=("Arial", 14))
        app.center_title_label.grid(row=1, column=0, sticky="nw", padx=10, pady=5)

        # Create a Listbox for the drop area to hold multiple items and allow selection
        app.drop_area = tk.Listbox(
            center_frame, bg="white", relief="sunken", bd=2, selectmode=tk.SINGLE
        )
        app.drop_area.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)

        # Add horizontal and vertical scrollbars to the center area
        # center_scrollbar_y = ttk.Scrollbar(
        #     center_frame, orient="vertical", command=app.drop_area.yview
        # )
        center_scrollbar_x = ttk.Scrollbar(
            center_frame, orient="horizontal", command=app.drop_area.xview
        )
        app.drop_area.configure(
            # yscrollcommand=center_scrollbar_y.set,
            xscrollcommand=center_scrollbar_x.set
        )
        # center_scrollbar_y.grid(row=2, column=1, sticky="ns")
        center_scrollbar_x.grid(row=3, column=0, sticky="ew")

        # Add + and - buttons to the center_button_frame
        app.add_button = ttk.Button(
            center_button_frame,
            text="+",
            command=lambda: add_item_to_center(app),
        )
        app.add_button.grid(row=0, column=0, padx=5, pady=5, sticky="n")

        app.remove_button = ttk.Button(
            center_button_frame,
            text="-",
            command=lambda: remove_item_from_center(app),
        )
        app.remove_button.grid(row=1, column=0, padx=5, pady=5, sticky="n")

        # Add Lock button
        app.lock_button = ttk.Button(
            center_button_frame,
            text="Lock",
            command=lambda: toggle_lock(
                app, app.add_button, app.remove_button, app.lock_button
            ),
        )
        app.lock_button.grid(row=2, column=0, padx=5, pady=5, sticky="n")

        # Create a Frame for Excel Treeview in the right area
        search_frame = tk.Frame(right_frame)
        search_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

        search_label = ttk.Label(search_frame, text="Search:")
        search_label.pack(side="left", padx=(0, 5))

        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var)
        search_entry.pack(side="left", fill="x", expand=True)

        search_button = ttk.Button(
            search_frame,
            text="Search",
            command=lambda: self.filter_excel_data(app, search_var.get()),
        )
        search_button.pack(side="left", padx=(5, 0))

        #######
        # Bind the Enter key to trigger the search when pressed in the search box
        search_entry.bind(
            "<Return>", lambda event: filter_excel_data(app, search_var.get())
        )
        #######

        excel_frame = tk.Frame(right_frame)
        excel_frame.grid(row=1, column=0, sticky="nsew")

        excel_frame.grid_rowconfigure(0, weight=1)
        excel_frame.grid_columnconfigure(0, weight=1)

        app.excel_treeview = ttk.Treeview(excel_frame, show="headings")
        app.excel_treeview.grid(row=0, column=0, sticky="nsew")

        scrollbar_x = ttk.Scrollbar(
            excel_frame, orient="horizontal", command=app.excel_treeview.xview
        )
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        app.excel_treeview.configure(xscrollcommand=scrollbar_x.set)

        scrollbar_y = ttk.Scrollbar(
            excel_frame, orient="vertical", command=app.excel_treeview.yview
        )
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        app.excel_treeview.configure(yscrollcommand=scrollbar_y.set)

        self.display_excel_data_openpyxl(app)

        app.notebook.add(tab_frame, text="WM 그룹별 매칭")

        app.wm_group_treeview.bind("<<TreeviewSelect>>", app.update_center_title_label)

        # Define styles for locked and unlocked items in the left area
        app.wm_group_treeview.tag_configure("locked", background="gray")
        app.wm_group_treeview.tag_configure("unlocked", background="white")

        # Enable drag-and-drop functionality
        self.enable_drag_and_drop(app)

        # Initialize the lock state
        app.lock_status = {}

        # Now load the WM Group Match data to set initial states
        load_wm_group_match_data(app)

        # Set the initial state based on the selected item or disable if none selected
        app.update_center_title_label(None)

    def save_lock_status_to_json(app):
        try:
            with open("wm_group_match.json", "r", encoding="utf-8") as f:
                data = json.load(f)
        except FileNotFoundError:
            data = {}

        if "lock_status" not in data:
            data["lock_status"] = {}

        # Update the lock status for the current item
        data["lock_status"] = app.lock_status

        with open("wm_group_match.json", "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

    def load_wm_group_match_data(app):
        try:
            with open("wm_group_match.json", "r", encoding="utf-8") as f:
                data = json.load(f)

            # Load the lock status from the JSON file
            app.lock_status = data.get("lock_status", {})

            # Set the initial state of each item in the left area based on the lock status
            for item in app.wm_group_treeview.get_children():
                item_name = app.wm_group_treeview.item(item, "values")[0]
                is_locked = app.lock_status.get(item_name, False)
                if is_locked:
                    app.wm_group_treeview.item(item, tags=("locked",))
                else:
                    app.wm_group_treeview.item(item, tags=("unlocked",))

        except FileNotFoundError:
            app.lock_status = {}

    def reapply_lock_status(app):
        """Reapply the lock status to the items in the left area based on their saved lock state."""
        for item_id in app.wm_group_treeview.get_children():
            item_name = app.wm_group_treeview.item(item_id, "values")[
                0
            ].lower()  # Normalize the name
            is_locked = app.lock_status.get(item_name, False)
            if is_locked:
                app.wm_group_treeview.item(item_id, tags=("locked",))
            else:
                app.wm_group_treeview.item(item_id, tags=("unlocked",))

    def save_to_json(data, filename="wm_group_match.json"):
        """Saves the given dictionary to a JSON file."""
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def enable_drag_and_drop(self, app):
        # Handle drag start
        def on_drag_start(event, treeview):
            # Ensure dragging is not initiated from the left area
            if treeview == app.wm_group_treeview:
                return  # Do nothing if dragging is initiated from the left area

            app.selected_items = treeview.selection()
            # Save the initial position
            app.drag_start_x = event.x_root
            app.drag_start_y = event.y_root

            # Start listening for mouse movement
            app.bind("<B1-Motion>", lambda e: initiate_drag(e, treeview))
            app.bind("<ButtonRelease-1>", on_drop)

        def initiate_drag(event, treeview):
            # Calculate distance moved
            dx = abs(event.x_root - app.drag_start_x)
            dy = abs(event.y_root - app.drag_start_y)

            selected_items = app.selected_items

            # Only start dragging if the mouse has moved a significant distance
            if dx > 5 or dy > 5:
                # Clear any previous drag data
                if hasattr(app, "dragged_items"):
                    del app.dragged_items

                # selected_items = treeview.selection()
                print(selected_items)  # Debugging
                if selected_items:
                    print("drag on")  # Debugging: print when drag starts
                    app.dragged_items = [(treeview, item) for item in selected_items]

                    item_texts = [
                        treeview.item(item, "values")[0] for item in selected_items
                    ]

                    # Create a label to follow the cursor
                    if hasattr(app, "drag_label"):
                        app.drag_label.destroy()  # Ensure previous label is destroyed
                    app.drag_label = ttk.Label(
                        app,
                        text=",\n ".join(item_texts),  # bg="yellow"
                    )
                    app.drag_label.place(x=event.x_root, y=event.y_root)

                    app.bind("<B1-Motion>", on_drag_motion)

        def on_drag_motion(event):
            # print(f"Motion event at x={event.x_root}, y={event.y_root}")  # Debugging
            if hasattr(app, "drag_label"):
                app.drag_label.place(
                    x=event.x_root,
                    y=event.y_root,
                )  # Slight offset for visibility

        def on_drop(event):
            print("drag off")  # Debugging: print when drag ends

            if not hasattr(app, "dragged_items"):
                return

            # Check if drop area is under the mouse
            widget_under_mouse = app.winfo_containing(event.x_root, event.y_root)
            if widget_under_mouse == app.drop_area:
                # Display dragged items in the drop area
                for treeview, item in app.dragged_items:
                    item_values = treeview.item(item, "values")
                    # Get the 0th, 7th, and 9th column values
                    item_texts = app.config_manager.get_item_texts(item_values)
                    item_string = " - ".join(item_texts)

                    # Insert all selected values into the Listbox
                    app.drop_area.insert(
                        tk.END,
                        item_string,
                    )

            # Cleanup: remove drag label
            if hasattr(app, "drag_label"):
                app.drag_label.destroy()
                del app.drag_label

            # Cleanup: clear dragged items
            if hasattr(app, "dragged_items"):
                del app.dragged_items

            # Unbind the motion and drop events
            app.unbind("<B1-Motion>")
            app.unbind("<ButtonRelease-1>")

        # Bind the drag start only to the Treeview in the right area (Excel data)
        app.excel_treeview.bind(
            "<ButtonPress-1>", lambda e: on_drag_start(e, app.excel_treeview)
        )

    def display_excel_data_openpyxl(self, app):
        # Load the Excel file using openpyxl
        file_path = (
            "WorkMaster_DB.xlsx"  # Assuming the file is in the project directory
        )
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Set up the Treeview columns
        columns = [
            cell.value for cell in sheet[1]
        ]  # Assuming the first row is the header
        app.excel_treeview["columns"] = columns

        # Columns to be hidden (2, 3, 5, 7, 9, 11, 13)
        hidden_columns = [1, 2, 4, 6, 8, 10, 12]  # Zero-based index

        for idx, col in enumerate(columns):
            if idx in hidden_columns:
                app.excel_treeview.column(col, width=0, stretch=tk.NO)
            else:
                app.excel_treeview.heading(col, text=col)
                app.excel_treeview.column(
                    col, anchor="w", width=150
                )  # Set a default width for visible columns

        # Insert rows into the Treeview
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Skip rows 2 through 5
            if 2 <= i <= 6:
                continue
            app.excel_treeview.insert("", "end", values=row)

    def create_three_area_tab(self, app, name):
        # Create a frame for the tab
        tab_frame = ttk.Frame(app.notebook)

        # Create left, middle, and right frames
        left_frame = ttk.Frame(tab_frame)
        middle_frame = ttk.Frame(tab_frame)
        right_frame = ttk.Frame(tab_frame)

        # Left content
        left_label = ttk.Label(left_frame, text=f"Left content of {name}")
        left_button = ttk.Button(
            left_frame,
            text="Left Button",
            command=lambda: app.context_menu_manager.on_button_click(
                f"Left Button on {name}"
            ),
        )
        left_text = ttk.Entry(left_frame, width=30)

        # Middle content
        middle_label = ttk.Label(middle_frame, text=f"Middle content of {name}")
        middle_button = ttk.Button(
            middle_frame,
            text="Middle Button",
            command=lambda: app.context_menu_manager.on_button_click(
                f"Middle Button on {name}"
            ),
        )
        middle_text = ttk.Entry(middle_frame, width=30)

        # Right content
        right_label = ttk.Label(right_frame, text=f"Right content of {name}")
        right_button = ttk.Button(
            right_frame,
            text="Right Button",
            command=lambda: app.context_menu_manager.on_button_click(
                f"Right Button on {name}"
            ),
        )
        right_text = ttk.Entry(right_frame, width=30)

        # Pack left content with padding
        left_label.pack(padx=20, pady=10)
        left_button.pack(padx=20, pady=10)
        left_text.pack(padx=20, pady=10)

        # Pack middle content with padding
        middle_label.pack(padx=20, pady=10)
        middle_button.pack(padx=20, pady=10)
        middle_text.pack(padx=20, pady=10)

        # Pack right content with padding
        right_label.pack(padx=20, pady=10)
        right_button.pack(padx=20, pady=10)
        right_text.pack(padx=20, pady=10)

        # Layout left, middle, and right frames
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        middle_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        right_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        # Add the tab to the notebook with increased padding
        app.notebook.add(tab_frame, text=f"   {name}   ")
